# -*- coding: utf-8 -*-
"""
Peer-distance research webapp (Flask + PostgreSQL)

Design intent:
- This app does NOT diagnose/evaluate students.
- It collects perceived relationship structure via spatial placements and summarizes structure via
  distance-based analysis, MDS, and (optional) K-means.

Operational notes:
- PostgreSQL is assumed (JSONB used).
- init_db() includes a minimal schema versioning + migrations skeleton.
"""

from __future__ import annotations

import io
import json
import math
import os
import random
import string
from datetime import datetime, timedelta
from typing import Any, Dict, List, Optional, Tuple

import requests
from flask import Flask, flash, jsonify, redirect, render_template, request, send_file, session
from io import BytesIO
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.units import mm
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont

from datetime import datetime
from zoneinfo import ZoneInfo

from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4, landscape
from reportlab.pdfbase.pdfmetrics import stringWidth

from sqlalchemy import create_engine, text
from sqlalchemy.orm import sessionmaker
from urllib.parse import quote, unquote
from werkzeug.security import check_password_hash, generate_password_hash

try:
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ModuleNotFoundError:  # openpyxl 미설치 환경
    Workbook = None  # type: ignore
    get_column_letter = None  # type: ignore
    OPENPYXL_AVAILABLE = False

    def get_column_letter(_n: int) -> str:  # pragma: no cover
        # openpyxl이 없을 때는 _autosize_columns가 호출되지 않도록 가드함
        return "A"



# -------------------------
# Database bootstrap
# -------------------------

DATABASE_URL = (os.environ.get("DATABASE_URL") or "").strip()
if DATABASE_URL.startswith("postgres://"):
    # Render/Heroku style legacy scheme
    DATABASE_URL = DATABASE_URL.replace("postgres://", "postgresql://", 1)

engine = None
SessionLocal = None

if DATABASE_URL:
    engine = create_engine(DATABASE_URL, pool_pre_ping=True)
    SessionLocal = sessionmaker(bind=engine)


def init_db() -> None:
    """
    DB schema initialization + minimal migration skeleton.

    How it works:
    - A single-row table schema_migrations stores the current schema version.
    - migrations is a list of (version, [sql...]) applied in order.
    - To evolve schema later WITHOUT dropping data:
        - Add a new (version+1, [...ALTER/CREATE...]) entry
        - Deploy; init_db will apply missing migrations.

    Dev mode (data not important):
    - Dropping tables and letting init_db recreate is simplest.
    """
    if not engine:
        return

    with engine.begin() as conn:
        # 0) schema_migrations
        conn.execute(text("""
        CREATE TABLE IF NOT EXISTS schema_migrations (
            id INTEGER PRIMARY KEY,
            version INTEGER NOT NULL,
            updated_at TIMESTAMP DEFAULT NOW()
        );
        """))
        conn.execute(text("""
        INSERT INTO schema_migrations (id, version)
        VALUES (1, 0)
        ON CONFLICT (id) DO NOTHING;
        """))

        row = conn.execute(text("SELECT version FROM schema_migrations WHERE id = 1")).fetchone()
        current_version = int(row.version) if row and row.version is not None else 0

        def set_version(v: int) -> None:
            conn.execute(text("""
            UPDATE schema_migrations
            SET version = :v,
                updated_at = NOW()
            WHERE id = 1
            """), {"v": int(v)})

        migrations: List[Tuple[int, List[str]]] = []

        migrations: List[Tuple[int, List[str]]] = []

        # -------------------
        # Migration v1: canonical schema for current app
        # -------------------
        migrations.append((1, [
            """
            CREATE TABLE IF NOT EXISTS teachers (
                id SERIAL PRIMARY KEY,
                username TEXT UNIQUE NOT NULL,
                created_at TIMESTAMP DEFAULT NOW()
            );
            """,
            """
            CREATE TABLE IF NOT EXISTS classes (
                id SERIAL PRIMARY KEY,
                code TEXT UNIQUE NOT NULL,
                name TEXT NOT NULL,
                teacher_username TEXT NOT NULL,
                created_at TIMESTAMP DEFAULT NOW()
            );
            """,
            """
            CREATE TABLE IF NOT EXISTS students (
                id SERIAL PRIMARY KEY,
                class_code TEXT NOT NULL,
                student_no TEXT,
                name TEXT NOT NULL
            );
            """,
            """
            CREATE TABLE IF NOT EXISTS student_sessions (
                id SERIAL PRIMARY KEY,
                class_code TEXT NOT NULL,
                sid TEXT NOT NULL,
                student_name TEXT NOT NULL,
                placements JSONB,
                placements_json TEXT,
                confidence INTEGER,
                priority INTEGER,
                submitted BOOLEAN DEFAULT FALSE,
                created_at TIMESTAMP DEFAULT NOW()
            );
            """,
            """
            CREATE UNIQUE INDEX IF NOT EXISTS uq_student_sessions
            ON student_sessions (class_code, sid, student_name);
            """,
            """
            CREATE TABLE IF NOT EXISTS teacher_placement_runs (
                id SERIAL PRIMARY KEY,
                class_code TEXT NOT NULL,
                session_id TEXT NOT NULL,
                teacher_username TEXT NOT NULL,
                condition TEXT,
                tool_run_id INTEGER,
                placements JSONB,
                placements_json TEXT,
                submitted BOOLEAN DEFAULT FALSE,
                started_at TIMESTAMP DEFAULT NOW(),
                ended_at TIMESTAMP,
                duration_ms INTEGER,
                confidence_score INTEGER,
                created_at TIMESTAMP DEFAULT NOW()
            );
            """,
            """
            CREATE INDEX IF NOT EXISTS ix_teacher_runs_class_session
            ON teacher_placement_runs (class_code, session_id);
            """,
            """
            CREATE TABLE IF NOT EXISTS teacher_decisions (
                id SERIAL PRIMARY KEY,
                run_id INTEGER NOT NULL,
                target_student_name TEXT NOT NULL,
                priority_rank INTEGER NOT NULL,
                decision_confidence INTEGER,
                reason_tags JSONB,
                created_at TIMESTAMP DEFAULT NOW()
            );
            """,
            """
            CREATE INDEX IF NOT EXISTS ix_teacher_decisions_run
            ON teacher_decisions (run_id);
            """,
            """
            CREATE TABLE IF NOT EXISTS analysis_cache (
                id SERIAL PRIMARY KEY,
                class_code TEXT NOT NULL,
                session_id TEXT NOT NULL,
                cache_key TEXT NOT NULL,
                payload JSONB,
                updated_at TIMESTAMP DEFAULT NOW(),
                UNIQUE(class_code, session_id, cache_key)
            );
            """,
        ]))

        # -------------------
        # Migration v2: roster + finalize + exclusions
        # -------------------
        migrations.append((2, [
            "ALTER TABLE students ADD COLUMN IF NOT EXISTS gender TEXT;",
            "ALTER TABLE students ADD COLUMN IF NOT EXISTS pin_code TEXT;",
            "ALTER TABLE students ADD COLUMN IF NOT EXISTS active BOOLEAN DEFAULT TRUE;",
            """
            CREATE UNIQUE INDEX IF NOT EXISTS uq_students_class_pin
            ON students (class_code, pin_code);
            """,
            """
            CREATE TABLE IF NOT EXISTS session_finalizations (
                id SERIAL PRIMARY KEY,
                class_code TEXT NOT NULL,
                session_id TEXT NOT NULL,
                teacher_username TEXT NOT NULL,
                exclusions_resolved BOOLEAN DEFAULT FALSE,
                survey_submitted BOOLEAN DEFAULT FALSE,
                finalized BOOLEAN DEFAULT FALSE,
                updated_at TIMESTAMP DEFAULT NOW(),
                UNIQUE(class_code, session_id)
            );
            """,
            """
            CREATE TABLE IF NOT EXISTS session_exclusions (
                id SERIAL PRIMARY KEY,
                class_code TEXT NOT NULL,
                session_id TEXT NOT NULL,
                student_name TEXT NOT NULL,
                excluded BOOLEAN DEFAULT TRUE,
                reason TEXT,
                created_at TIMESTAMP DEFAULT NOW(),
                UNIQUE(class_code, session_id, student_name)
            );
            """,
        ]))

        # -------------------
        # Migration v3: teacher_surveys (fix missing table on old DBs)
        # -------------------
        migrations.append((3, [
            """
            CREATE TABLE IF NOT EXISTS teacher_surveys (
                id SERIAL PRIMARY KEY,
                class_code TEXT NOT NULL,
                session_id TEXT NOT NULL,
                teacher_username TEXT NOT NULL,
                q1_help INTEGER,
                q2_new TEXT,
                q2_detail TEXT,
                q3_use INTEGER,
                q4_cmp TEXT,
                q4_detail TEXT,
                q5_conf TEXT,
                q6_next TEXT,
                q7_feedback TEXT,
                submitted_at TIMESTAMP DEFAULT NOW(),
                UNIQUE(class_code, session_id)
            );
            """,
            """
            CREATE INDEX IF NOT EXISTS ix_teacher_surveys_class_session
            ON teacher_surveys (class_code, session_id);
            """,
        ]))

        # -------------------
        # Migration v4: student creation timestamp
        # -------------------
        migrations.append((4, [
            "ALTER TABLE students ADD COLUMN IF NOT EXISTS joined_at TIMESTAMP DEFAULT NOW();",
        ]))

        # -------------------
        # Migration v5: archived reset snapshots
        # -------------------
        migrations.append((5, [
            """
            CREATE TABLE IF NOT EXISTS archived_session_resets (
                id SERIAL PRIMARY KEY,
                class_code TEXT NOT NULL,
                session_id TEXT NOT NULL,
                reset_by TEXT NOT NULL,
                reset_reason TEXT,
                archived_payload JSONB,
                created_at TIMESTAMP DEFAULT NOW()
            );
            """,
            """
            CREATE INDEX IF NOT EXISTS ix_archived_session_resets_class_session
            ON archived_session_resets (class_code, session_id);
            """,
        ]))

        # -------------------
        # Migration v6: class archive + light admin editing
        # -------------------
        migrations.append((6, [
            "ALTER TABLE classes ADD COLUMN IF NOT EXISTS archived BOOLEAN DEFAULT FALSE;",
            "ALTER TABLE classes ADD COLUMN IF NOT EXISTS archived_at TIMESTAMP;",
            "ALTER TABLE classes ADD COLUMN IF NOT EXISTS archived_by TEXT;",
            "ALTER TABLE classes ADD COLUMN IF NOT EXISTS archived_reason TEXT;",
            "ALTER TABLE classes ADD COLUMN IF NOT EXISTS updated_at TIMESTAMP DEFAULT NOW();",
        ]))



        migrations.sort(key=lambda x: int(x[0]))
        for target_version, stmts in migrations:
            if current_version >= int(target_version):
                continue
            for stmt in stmts:
                if not stmt or not str(stmt).strip():
                    continue
                conn.execute(text(stmt))
            set_version(int(target_version))
            current_version = int(target_version)


# Initialize DB tables on startup (best-effort; app still boots for debug)
try:
    init_db()
except Exception as e:
    print("init_db failed:", e)

# -------------------------
# Flask bootstrap
# -------------------------

app = Flask(__name__)
app.config["PROPAGATE_EXCEPTIONS"] = True
app.secret_key = os.environ.get("SECRET_KEY", "dev-only-change-me")

app.config["PERMANENT_SESSION_LIFETIME"] = timedelta(days=30)
app.config["SESSION_COOKIE_HTTPONLY"] = True
app.config["SESSION_COOKIE_SAMESITE"] = "Lax"


@app.before_request
def make_session_permanent() -> None:
    session.permanent = True

# -------------------------
# Research admin (owner-only)
# -------------------------
ADMIN_USERS = {
    u.strip()
    for u in (os.environ.get("ADMIN_USERS", "") or "").split(",")
    if u.strip()
}

def require_admin():
    """
    Owner-only research/admin pages.
    - Must be logged in as teacher.
    - Username must be in ADMIN_USERS (env var, comma-separated).
    """
    if "teacher" not in session:
        return redirect("/teacher/login")

    if ADMIN_USERS and session.get("teacher") not in ADMIN_USERS:
        # 권한 없음
        return "forbidden", 403

    # ADMIN_USERS가 비어있으면(환경변수 미설정) 안전하게 막고 싶다면 아래 주석 해제:
    # if not ADMIN_USERS:
    #     return "forbidden (ADMIN_USERS not set)", 403

    return None


def is_research_admin_user(username: Optional[str]) -> bool:
    """Return whether this teacher account can use research/admin privileges."""
    if not username:
        return False
    if not ADMIN_USERS:
        # Matches require_admin(): when ADMIN_USERS is not configured, the
        # existing app treats any logged-in teacher as the local admin user.
        return True
    return str(username) in ADMIN_USERS

# -------------------------
# Research admin: XLSX helpers + overview fetch
# -------------------------

def _xlsx_response(wb, filename: str):
    """
    Safe XLSX response helper.
    - openpyxl 미설치/오작동 시: 라우트가 500으로 깔끔하게 실패하고, 앱 전체 부팅을 깨지 않도록 방어
    """
    if not OPENPYXL_AVAILABLE or Workbook is None:
        return "openpyxl not installed on server", 500

    if wb is None:
        return "workbook is None", 500

    bio = io.BytesIO()
    try:
        wb.save(bio)
    except Exception as e:
        return f"failed to generate xlsx: {e}", 500

    bio.seek(0)
    return send_file(
        bio,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=filename,
    )


def _autosize_columns(ws):
    # 단순 자동 폭 (완벽하진 않지만 연구용엔 충분)
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                v = "" if cell.value is None else str(cell.value)
            except Exception:
                v = ""
            if len(v) > max_len:
                max_len = len(v)
        ws.column_dimensions[col_letter].width = min(max_len + 2, 60)

def db_fetch_class_overview() -> List[Dict[str, Any]]:
    """Return class/session progress data for the research admin dashboard."""
    if not engine:
        return []

    with engine.connect() as conn:
        classes = conn.execute(text("""
            SELECT code, name, teacher_username, created_at,
                   COALESCE(archived, FALSE) AS archived,
                   archived_at, archived_by, archived_reason
            FROM classes
            ORDER BY COALESCE(archived, FALSE) ASC, id DESC
        """)).fetchall()

        out: List[Dict[str, Any]] = []
        for c in classes:
            student_cnt = conn.execute(text("""
                SELECT COUNT(*) AS n FROM students
                WHERE class_code = :code AND COALESCE(active, TRUE) = TRUE
            """), {"code": c.code}).fetchone().n

            sessions: List[Dict[str, Any]] = []
            for sid in VISIBLE_SESSION_IDS:
                ss = conn.execute(text("""
                    SELECT COUNT(*) AS total,
                           SUM(CASE WHEN submitted THEN 1 ELSE 0 END) AS submitted
                    FROM student_sessions
                    WHERE class_code = :code AND sid = :sid
                """), {"code": c.code, "sid": sid}).fetchone()

                tr = conn.execute(text("""
                    SELECT COUNT(*) AS total,
                           SUM(CASE WHEN submitted THEN 1 ELSE 0 END) AS submitted
                    FROM teacher_placement_runs
                    WHERE class_code = :code AND session_id = :sid
                """), {"code": c.code, "sid": sid}).fetchone()

                survey = conn.execute(text("""
                    SELECT submitted_at
                    FROM teacher_surveys
                    WHERE class_code = :code AND session_id = :sid
                    LIMIT 1
                """), {"code": c.code, "sid": sid}).fetchone()

                fin = conn.execute(text("""
                    SELECT exclusions_resolved, survey_submitted, finalized
                    FROM session_finalizations
                    WHERE class_code = :code AND session_id = :sid
                    LIMIT 1
                """), {"code": c.code, "sid": sid}).fetchone()

                reset_row = conn.execute(text("""
                    SELECT COUNT(*) AS n
                    FROM archived_session_resets
                    WHERE class_code = :code AND session_id = :sid
                """), {"code": c.code, "sid": sid}).fetchone()

                submitted = int((ss.submitted if ss and ss.submitted is not None else 0) or 0)
                teacher_submitted = int((tr.submitted if tr and tr.submitted is not None else 0) or 0)
                similarity = teacher_student_similarity_summary(str(c.code), sid)
                sessions.append({
                    "sid": sid,
                    "label": f"{int(sid) - 1}회차",
                    "student_total": int(student_cnt or 0),
                    "student_submitted": submitted,
                    "student_progress_pct": round((submitted / int(student_cnt or 1)) * 100),
                    "teacher_placement_done": teacher_submitted > 0,
                    "teacher_placement_count": teacher_submitted,
                    "pre_survey_done": bool(conn.execute(text("""
                        SELECT 1
                        FROM teacher_placement_runs
                        WHERE class_code = :code
                          AND session_id = :sid
                          AND submitted = TRUE
                          AND confidence_score IS NOT NULL
                        LIMIT 1
                    """), {"code": c.code, "sid": sid}).fetchone()),
                    "post_survey_done": bool(survey),
                    "exclusions_resolved": bool(getattr(fin, "exclusions_resolved", False)) if fin else False,
                    "finalized": bool(getattr(fin, "finalized", False)) if fin else False,
                    "reset_count": int(reset_row.n or 0) if reset_row else 0,
                    "similarity": similarity,
                })

            out.append({
                "code": c.code,
                "name": c.name,
                "teacher_username": c.teacher_username,
                "created_at": c.created_at,
                "archived": bool(getattr(c, "archived", False)),
                "archived_at": getattr(c, "archived_at", None),
                "archived_by": getattr(c, "archived_by", None),
                "archived_reason": getattr(c, "archived_reason", None),
                "student_count": int(student_cnt or 0),
                "sessions": sessions,
            })
        return out


def teacher_student_similarity_summary(class_code: str, sid: str) -> Dict[str, Any]:
    """Compare the latest teacher map with the student aggregate map for a session."""
    try:
        students = db_get_students_in_class(class_code)
        names = [(s.get("name") or "").strip() for s in students if (s.get("name") or "").strip()]
        if len(names) < 3:
            return {"ok": False, "label": "표본 부족", "score": None, "n_pairs": 0}

        with engine.connect() as conn:
            row = conn.execute(text("""
                SELECT placements
                FROM teacher_placement_runs
                WHERE class_code = :code AND session_id = :sid AND submitted = TRUE
                ORDER BY ended_at DESC NULLS LAST, created_at DESC
                LIMIT 1
            """), {"code": class_code, "sid": sid}).fetchone()
        if not row:
            return {"ok": False, "label": "교사 배치 없음", "score": None, "n_pairs": 0}

        teacher_placements = row.placements if isinstance(row.placements, dict) else _json_load_maybe(row.placements)
        if not isinstance(teacher_placements, dict):
            return {"ok": False, "label": "교사 배치 없음", "score": None, "n_pairs": 0}

        teacher_pts, teacher_valid = points_from_placements_all_students(teacher_placements, names)
        teacher_d = distance_matrix(teacher_pts, teacher_valid)

        avg_payload = student_avg_distance_payload(class_code, sid)
        avg_d = avg_payload.get("avg_distance_matrix") or []
        xs: List[float] = []
        ys: List[float] = []
        n = len(names)
        for i in range(n):
            for j in range(i + 1, n):
                tv = teacher_d[i][j] if i < len(teacher_d) and j < len(teacher_d[i]) else None
                sv = avg_d[i][j] if i < len(avg_d) and j < len(avg_d[i]) else None
                if isinstance(tv, (int, float)) and isinstance(sv, (int, float)):
                    xs.append(float(tv))
                    ys.append(float(sv))

        if len(xs) < 3:
            return {"ok": False, "label": "비교 부족", "score": None, "n_pairs": len(xs)}

        mx = sum(xs) / len(xs)
        my = sum(ys) / len(ys)
        vx = sum((x - mx) ** 2 for x in xs)
        vy = sum((y - my) ** 2 for y in ys)
        if vx <= 1e-12 or vy <= 1e-12:
            return {"ok": False, "label": "비교 부족", "score": None, "n_pairs": len(xs)}

        corr = sum((x - mx) * (y - my) for x, y in zip(xs, ys)) / math.sqrt(vx * vy)
        if corr >= 0.55:
            label = "높음"
        elif corr >= 0.25:
            label = "보통"
        else:
            label = "낮음"
        return {"ok": True, "label": label, "score": round(corr, 4), "n_pairs": len(xs)}
    except Exception as e:
        app.logger.exception("teacher_student_similarity_summary failed: %s %s", class_code, sid)
        return {"ok": False, "label": "계산 오류", "score": None, "n_pairs": 0, "error": str(e)}


SITE_TITLE = "내가 바라본 우리 반"
DEFAULT_VISIBLE_SESSION_ID = "2"
VISIBLE_SESSION_IDS = ["2", "3", "4"]


def normalize_visible_session_id(raw: Any) -> str:
    sid = str(raw or "").strip()
    return sid if sid in VISIBLE_SESSION_IDS else DEFAULT_VISIBLE_SESSION_ID

# JSON fallback file (only used if DB not configured)
DATA_FILE = os.environ.get("DATA_FILE", "data.json")

# Google Sheets integration
GOOGLE_WEBAPP_URL = os.environ.get(
    "GOOGLE_WEBAPP_URL",
    "https://script.google.com/macros/s/AKfycbwyjKC2JearJnySkxdG0oahMkMJ5V6uBqY5EYRGVVRa8KWZvRzHcskeVNY5hnlyiSw/exec",
)
GOOGLE_SECRET = (os.environ.get("GOOGLE_SECRET") or "").strip()

DEBUG_MODE = os.environ.get("DEBUG_MODE") == "1"


# -------------------------
# Utilities: Google Sheets
# -------------------------

def post_to_sheet(payload: Dict[str, Any]) -> Dict[str, Any]:
    payload = dict(payload)
    payload["secret"] = GOOGLE_SECRET

    try:
        r = requests.post(GOOGLE_WEBAPP_URL, json=payload, timeout=10)
    except Exception as e:
        return {"status": "error", "message": f"request failed: {e}"}

    if r.status_code != 200:
        return {"status": "error", "message": f"http {r.status_code}", "text": r.text[:300]}

    try:
        return r.json()
    except Exception:
        return {"status": "error", "message": "invalid json response", "text": r.text[:300]}


def sheet_list_results(class_code: str, sid: str) -> List[Dict[str, Any]]:
    """Google Sheets Results에서 특정 학급/회차 row 목록을 가져온다(테스트/동기화용)."""
    resp = post_to_sheet({
        "action": "results_list",
        "class_code": class_code,
        "session": str(sid),
    })
    if resp.get("status") != "ok":
        raise RuntimeError(f"sheet_list_results failed: {resp}")
    rows = resp.get("rows") or []
    if not isinstance(rows, list):
        return []
    return rows


def sheet_upsert_teacher_survey(
    class_code: str,
    sid: str,
    teacher_username: str,
    survey_type: str,
    payload: Dict[str, Any],
) -> Dict[str, Any]:
    """Send teacher pre/post survey data to the connected Google Sheet."""
    return post_to_sheet({
        "action": "teacher_survey_upsert",
        "survey_type": survey_type,
        "teacher": teacher_username,
        "class_code": class_code,
        "session": str(sid),
        "payload": payload or {},
    })


def sheet_upsert_research_session_summary(payload: Dict[str, Any]) -> Dict[str, Any]:
    """Send one class/session research summary row to the connected Google Sheet."""
    body = dict(payload or {})
    body["action"] = "research_session_summary_upsert"
    return post_to_sheet(body)


def sheet_append_archived_reset(payload: Dict[str, Any]) -> Dict[str, Any]:
    """Send a reset archive snapshot to the connected Google Sheet."""
    body = dict(payload or {})
    body["action"] = "archived_session_reset_append"
    return post_to_sheet(body)

      
# -------------------------
# Utilities: JSON file fallback
# -------------------------

def ensure_class_schema(cls: Optional[Dict[str, Any]]) -> Optional[Dict[str, Any]]:
    if not cls:
        return cls
    cls.setdefault("sessions", {})
    for i in range(1, 6):
        cls["sessions"].setdefault(str(i), {"label": f"{i}차", "active": i == 1})

    for name, sdata in cls.get("students_data", {}).items():
        sdata.setdefault("sessions", {})
        for sid in cls["sessions"]:
            sdata["sessions"].setdefault(sid, {"placements": {}, "submitted": False})
    return cls


def load_data() -> Dict[str, Any]:
    if not os.path.exists(DATA_FILE):
        return {"classes": {}}

    try:
        with open(DATA_FILE, "r", encoding="utf-8") as f:
            d = json.load(f)
    except Exception:
        return {"classes": {}}

    d.setdefault("classes", {})
    for code, cls in list(d.get("classes", {}).items()):
        d["classes"][code] = ensure_class_schema(cls)
    return d


def save_data(data: Dict[str, Any]) -> None:
    parent_dir = os.path.dirname(DATA_FILE) or "."
    os.makedirs(parent_dir, exist_ok=True)

    tmp_path = DATA_FILE + ".tmp"
    with open(tmp_path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    os.replace(tmp_path, DATA_FILE)


def save_data_safely(d: Dict[str, Any]) -> None:
    for code, cls in list(d.get("classes", {}).items()):
        d["classes"][code] = ensure_class_schema(cls)
    save_data(d)


def local_create_teacher(username: str, pw_hash: str, profile: Optional[Dict[str, Any]] = None) -> Dict[str, Any]:
    d = load_data()
    d.setdefault("teachers", {})
    if username in d["teachers"]:
        return {"status": "exists"}
    d["teachers"][username] = {
        "pw_hash": pw_hash,
        "profile": profile or {},
        "created_at": datetime.utcnow().isoformat() + "Z",
    }
    save_data_safely(d)
    return {"status": "ok"}


def local_get_teacher(username: str) -> Dict[str, Any]:
    d = load_data()
    teacher = (d.get("teachers") or {}).get(username)
    if not teacher:
        return {"status": "not_found"}
    return {
        "status": "ok",
        "pw_hash": teacher.get("pw_hash") or "",
        "profile": teacher.get("profile") or {},
    }


def make_code() -> str:
    return "".join(random.choices(string.ascii_uppercase + string.digits, k=6))
def normalize_gender(raw: str) -> str:
    """
    성별 입력 정규화:
    - 입력이 무엇이든 최종 저장은 "남" / "여" / ""(미입력) 중 하나로 통일
    - 이번 프로젝트에서는 현장 사용 패턴(엑셀/한글 명부)에 맞춰 한국어 표현만 허용
    """
    s = (raw or "").strip()
    if not s:
        return ""

    male = {"남", "남자", "남성", "남학생", "남아"}
    female = {"여", "여자", "여성", "여학생", "여아"}

    if s in male:
        return "남"
    if s in female:
        return "여"
    return ""


# -------------------------
# DB helpers (canonical schema)
# -------------------------

def _json_load_maybe(val: Any) -> Dict[str, Any]:
    if isinstance(val, dict):
        return val
    if isinstance(val, str):
        try:
            obj = json.loads(val)
            return obj if isinstance(obj, dict) else {}
        except Exception:
            return {}
    return {}


def db_list_classes_for_teacher(teacher_username: str) -> Dict[str, Dict[str, str]]:
    if not engine:
        raise RuntimeError("DB engine not initialized")
    with engine.connect() as conn:
        rows = conn.execute(text("""
            SELECT code, name, teacher_username
            FROM classes
            WHERE teacher_username = :t
              AND COALESCE(archived, FALSE) = FALSE
            ORDER BY id DESC
        """), {"t": teacher_username}).fetchall()

    out: Dict[str, Dict[str, str]] = {}
    for r in rows:
        out[r.code] = {"name": r.name, "teacher": r.teacher_username}
    return out


def db_update_class_name_admin(class_code: str, new_name: str, updated_by: str) -> bool:
    """Update only the class display name, leaving student/session data untouched."""
    if not engine:
        raise RuntimeError("DB engine not initialized")
    class_code = (class_code or "").upper().strip()
    new_name = (new_name or "").strip()
    if not class_code or not new_name:
        return False
    with engine.begin() as conn:
        res = conn.execute(text("""
            UPDATE classes
            SET name = :name,
                updated_at = NOW()
            WHERE code = :code
        """), {"code": class_code, "name": new_name})
    return bool(res.rowcount)


def db_set_class_archived_admin(class_code: str, archived: bool, updated_by: str, reason: str = "") -> bool:
    """Hide or restore a class without deleting any research data."""
    if not engine:
        raise RuntimeError("DB engine not initialized")
    class_code = (class_code or "").upper().strip()
    updated_by = (updated_by or "admin").strip()
    reason = (reason or "").strip()
    if not class_code:
        return False
    with engine.begin() as conn:
        if archived:
            res = conn.execute(text("""
                UPDATE classes
                SET archived = TRUE,
                    archived_at = NOW(),
                    archived_by = :by,
                    archived_reason = :reason,
                    updated_at = NOW()
                WHERE code = :code
            """), {"code": class_code, "by": updated_by, "reason": reason})
        else:
            res = conn.execute(text("""
                UPDATE classes
                SET archived = FALSE,
                    archived_at = NULL,
                    archived_by = NULL,
                    archived_reason = NULL,
                    updated_at = NOW()
                WHERE code = :code
            """), {"code": class_code})
    return bool(res.rowcount)


def db_create_class(teacher_username: str, class_code: str, class_name: str, students: List[Dict[str, str]]) -> None:
    if not engine:
        raise RuntimeError("DB engine not initialized")

    import random

    def make_pin(existing: set) -> str:
        # 6자리 숫자: 학급 내부에서만 유일하면 됨
        # (000000도 가능하지만 헷갈리므로 100000~999999 권장)
        for _ in range(2000):
            pin = str(random.randint(100000, 999999))
            if pin not in existing:
                existing.add(pin)
                return pin
        # 최악의 경우(거의 없음): 순차 발급
        pin = "100000"
        while pin in existing:
            pin = str(int(pin) + 1)
        existing.add(pin)
        return pin

    with engine.begin() as conn:
        conn.execute(text("""
            INSERT INTO classes (code, name, teacher_username)
            VALUES (:code, :name, :t)
        """), {"code": class_code, "name": class_name, "t": teacher_username})

        existing_pins: set = set()

        for s in students:
            nm = (s.get("name") or "").strip()
            if not nm:
                continue

            pin = make_pin(existing_pins)
            gender = normalize_gender(s.get("gender") or "")

            conn.execute(text("""
                INSERT INTO students (class_code, student_no, name, gender, pin_code, active, joined_at)
                VALUES (:code, :no, :name, :gender, :pin, TRUE, NOW())
            """), {
                "code": class_code,
                "no": str(s.get("no", "") or ""),
                "name": nm,
                "gender": gender,
                "pin": pin,
            })

def db_get_latest_teacher_run_id(class_code: str, teacher_username: str, sid: str) -> Optional[int]:
    if not engine:
        return None
    with engine.connect() as conn:
        row = conn.execute(text("""
            SELECT id
            FROM teacher_placement_runs
            WHERE class_code = :c AND teacher_username = :t AND session_id = :s
            ORDER BY started_at DESC, id DESC
            LIMIT 1
        """), {"c": class_code, "t": teacher_username, "s": sid}).fetchone()
    return int(row[0]) if row else None


def db_delete_class_for_teacher(class_code: str, teacher_username: str) -> bool:
    if not engine:
        raise RuntimeError("DB engine not initialized")

    class_code = (class_code or "").upper().strip()

    with engine.begin() as conn:
        row = conn.execute(text("""
            SELECT 1 FROM classes
            WHERE code = :code AND teacher_username = :t
            LIMIT 1
        """), {"code": class_code, "t": teacher_username}).fetchone()
        if not row:
            return False

        conn.execute(text("DELETE FROM teacher_decisions WHERE run_id IN (SELECT id FROM teacher_placement_runs WHERE class_code = :code)"), {"code": class_code})
        conn.execute(text("DELETE FROM teacher_placement_runs WHERE class_code = :code"), {"code": class_code})
        conn.execute(text("DELETE FROM student_sessions WHERE class_code = :code"), {"code": class_code})
        conn.execute(text("DELETE FROM students WHERE class_code = :code"), {"code": class_code})
        conn.execute(text("DELETE FROM analysis_cache WHERE class_code = :code"), {"code": class_code})
        conn.execute(text("DELETE FROM classes WHERE code = :code AND teacher_username = :t"), {"code": class_code, "t": teacher_username})

    return True


def db_get_class_name(class_code: str) -> Optional[str]:
    if not engine:
        return None
    with engine.connect() as conn:
        row = conn.execute(text("SELECT name FROM classes WHERE code = :code LIMIT 1"), {"code": class_code}).fetchone()
    return row.name if row else None


def db_get_class_for_teacher(class_code: str, teacher_username: str) -> Optional[Dict[str, Any]]:
    if not engine:
        raise RuntimeError("DB engine not initialized")

    with engine.connect() as conn:
        row = conn.execute(text("""
            SELECT code, name, teacher_username
            FROM classes
            WHERE code = :code
            LIMIT 1
        """), {"code": class_code}).fetchone()

    if not row:
        return None
    if row.teacher_username != teacher_username:
        return {"_forbidden": True}

    return {"code": row.code, "name": row.name, "teacher": row.teacher_username, "sessions": {}}


def db_get_class_for_viewer(class_code: str, viewer_username: str) -> Optional[Dict[str, Any]]:
    """Return a class if the viewer is its teacher or a research admin."""
    if not engine:
        raise RuntimeError("DB engine not initialized")

    class_code = (class_code or "").upper().strip()
    if is_research_admin_user(viewer_username):
        with engine.connect() as conn:
            row = conn.execute(text("""
                SELECT code, name, teacher_username
                FROM classes
                WHERE code = :code
                LIMIT 1
            """), {"code": class_code}).fetchone()
        if not row:
            return None
        return {
            "code": row.code,
            "name": row.name,
            "teacher": row.teacher_username,
            "sessions": {},
            "admin_view": row.teacher_username != viewer_username,
        }

    return db_get_class_for_teacher(class_code, viewer_username)


def db_get_students_in_class(class_code: str) -> List[Dict[str, str]]:
    if not engine:
        raise RuntimeError("DB engine not initialized")

    with engine.connect() as conn:
        rows = conn.execute(text("""
            SELECT student_no, name
            FROM students
            WHERE class_code = :code
              AND active = TRUE
            ORDER BY id ASC
        """), {"code": class_code}).fetchall()

    out: List[Dict[str, str]] = []
    for r in rows:
        out.append({
            "no": r.student_no or "",
            "name": r.name
        })
    return out



    out: List[Dict[str, str]] = []
    for r in rows:
        out.append({"no": (r.student_no or ""), "name": r.name})
    return out


def db_get_submitted_map(class_code: str, sid: str) -> Dict[str, bool]:
    if not engine:
        raise RuntimeError("DB engine not initialized")

    with engine.connect() as conn:
        rows = conn.execute(text("""
            SELECT student_name, submitted
            FROM student_sessions
            WHERE class_code = :code AND sid = :sid
        """), {"code": class_code, "sid": sid}).fetchall()

    m: Dict[str, bool] = {}
    for r in rows:
        m[r.student_name] = bool(r.submitted)
    return m



def db_get_session_state_v2(class_code: str, sid: str, teacher_username: str) -> Dict[str, Any]:
    """Return v2 finalize state + exclusions + survey flags for a given class/session."""
    if not engine:
        raise RuntimeError("DB engine not initialized")

    with engine.begin() as conn:
        # Ensure anchor row exists
        conn.execute(text("""
            INSERT INTO session_finalizations (class_code, session_id, teacher_username)
            VALUES (:code, :sid, :t)
            ON CONFLICT (class_code, session_id)
            DO UPDATE SET teacher_username = EXCLUDED.teacher_username,
                          updated_at = NOW()
        """), {"code": class_code, "sid": sid, "t": teacher_username})

        fin = conn.execute(text("""
            SELECT preview_seen, exclusions_resolved, survey_submitted, finalized,
                   preview_seen_at, exclusions_resolved_at, survey_submitted_at, finalized_at
            FROM session_finalizations
            WHERE class_code = :code AND session_id = :sid
        """), {"code": class_code, "sid": sid}).fetchone()

        try:
            exc_rows = conn.execute(text("""
                SELECT student_name, excluded, reason
                FROM session_exclusions
                WHERE class_code = :code AND session_id = :sid
                ORDER BY student_name
            """), {"code": class_code, "sid": sid}).fetchall()
        except Exception:
            # session_exclusions 테이블이 아직 없으면(초기/마이그레이션 누락) 빈 목록으로 처리
            exc_rows = []


        survey_row = conn.execute(text("""
            SELECT 1 AS ok
            FROM teacher_surveys
            WHERE class_code = :code AND session_id = :sid
        """), {"code": class_code, "sid": sid}).fetchone()

    exclusions = []
    for r in exc_rows:
        exclusions.append({
            "student_name": r.student_name,
            "excluded": bool(r.excluded),
            "reason": (r.reason or "")
        })

    state = {
        "class_code": class_code,
        "session_id": sid,
        "preview_seen": bool(getattr(fin, "preview_seen", False)),
        "exclusions_resolved": bool(getattr(fin, "exclusions_resolved", False)),
        "survey_submitted": bool(getattr(fin, "survey_submitted", False)) or bool(survey_row),
        "finalized": bool(getattr(fin, "finalized", False)),
        "exclusions": exclusions,
        "survey": db_get_survey_v2(class_code, sid),
    }
    submitted_map = db_get_submitted_map(class_code, sid)
    for item in state["exclusions"]:
        item["submitted"] = bool(submitted_map.get(item["student_name"], False))
    return state


def db_get_survey_v2(class_code: str, sid: str) -> Optional[Dict[str, Any]]:
    if not engine:
        return None

    with engine.connect() as conn:
        row = conn.execute(text("""
            SELECT q1_help, q2_new, q2_detail, q3_use, q4_cmp, q4_detail,
                   q5_conf, q6_next, q7_feedback, submitted_at
            FROM teacher_surveys
            WHERE class_code = :code AND session_id = :sid
            LIMIT 1
        """), {"code": class_code, "sid": sid}).fetchone()

    if not row:
        return None

    return {
        "q1_help": row.q1_help,
        "q2_new": row.q2_new,
        "q2_detail": row.q2_detail or "",
        "q3_use": row.q3_use,
        "q4_cmp": row.q4_cmp,
        "q4_detail": row.q4_detail or "",
        "q5_conf": row.q5_conf,
        "q6_next": row.q6_next,
        "q7_feedback": row.q7_feedback or "",
        "submitted_at": row.submitted_at.isoformat() if row.submitted_at else None,
    }


def db_set_preview_seen_v2(class_code: str, sid: str, teacher_username: str) -> None:
    if not engine:
        raise RuntimeError("DB engine not initialized")

    with engine.begin() as conn:
        conn.execute(text("""
            INSERT INTO session_finalizations (class_code, session_id, teacher_username, preview_seen, preview_seen_at)
            VALUES (:code, :sid, :t, TRUE, NOW())
            ON CONFLICT (class_code, session_id)
            DO UPDATE SET preview_seen = TRUE,
                          preview_seen_at = NOW(),
                          teacher_username = EXCLUDED.teacher_username,
                          updated_at = NOW()
        """), {"code": class_code, "sid": sid, "t": teacher_username})


def db_set_exclusions_v2(class_code: str, sid: str, teacher_username: str, mode: str, items: List[Dict[str, str]]) -> None:
    """mode:
      - 'no_exclude': mark exclusions_resolved true, keep exclusions table empty
      - 'exclude': upsert exclusions with reasons, mark exclusions_resolved true
    """
    if not engine:
        raise RuntimeError("DB engine not initialized")

    with engine.begin() as conn:
        # Clear previous exclusions for this session (simple and safe)
        conn.execute(text("""
            DELETE FROM session_exclusions
            WHERE class_code = :code AND session_id = :sid
        """), {"code": class_code, "sid": sid})

        if mode == "exclude":
            for it in items:
                nm = (it.get("student_name") or "").strip()
                rs = (it.get("reason") or "").strip()
                if not nm:
                    continue
                conn.execute(text("""
                    INSERT INTO session_exclusions (class_code, session_id, student_name, excluded, reason)
                    VALUES (:code, :sid, :nm, TRUE, :rs)
                    ON CONFLICT (class_code, session_id, student_name)
                    DO UPDATE SET excluded = TRUE, reason = EXCLUDED.reason
                """), {"code": class_code, "sid": sid, "nm": nm, "rs": rs})

        conn.execute(text("""
            INSERT INTO session_finalizations (class_code, session_id, teacher_username, exclusions_resolved, exclusions_resolved_at)
            VALUES (:code, :sid, :t, TRUE, NOW())
            ON CONFLICT (class_code, session_id)
            DO UPDATE SET exclusions_resolved = TRUE,
                          exclusions_resolved_at = NOW(),
                          teacher_username = EXCLUDED.teacher_username,
                          updated_at = NOW()
        """), {"code": class_code, "sid": sid, "t": teacher_username})


def db_remove_session_exclusion_v2(class_code: str, sid: str, teacher_username: str, student_name: str) -> None:
    """Remove one student from the session exclusion list."""
    if not engine:
        raise RuntimeError("DB engine not initialized")

    with engine.begin() as conn:
        conn.execute(text("""
            DELETE FROM session_exclusions
            WHERE class_code = :code
              AND session_id = :sid
              AND student_name = :name
        """), {"code": class_code, "sid": sid, "name": student_name})
        conn.execute(text("""
            INSERT INTO session_finalizations (class_code, session_id, teacher_username, exclusions_resolved, exclusions_resolved_at)
            VALUES (:code, :sid, :t, TRUE, NOW())
            ON CONFLICT (class_code, session_id)
            DO UPDATE SET exclusions_resolved = TRUE,
                          exclusions_resolved_at = NOW(),
                          teacher_username = EXCLUDED.teacher_username,
                          updated_at = NOW()
        """), {"code": class_code, "sid": sid, "t": teacher_username})


def db_get_excluded_student_names(class_code: str, sid: str) -> set:
    """Return students intentionally excluded from a session."""
    if not engine:
        return set()
    with engine.connect() as conn:
        rows = conn.execute(text("""
            SELECT student_name
            FROM session_exclusions
            WHERE class_code = :code
              AND session_id = :sid
              AND excluded = TRUE
        """), {"code": class_code, "sid": sid}).fetchall()
    return {(r.student_name or "").strip() for r in rows if (r.student_name or "").strip()}


def db_upsert_survey_v2(class_code: str, sid: str, teacher_username: str, payload: Dict[str, Any]) -> None:
    if not engine:
        raise RuntimeError("DB engine not initialized")

    def _to_int(v: Any) -> Optional[int]:
        try:
            if v is None or v == "":
                return None
            return int(v)
        except Exception:
            return None

    row = {
        "code": class_code,
        "sid": sid,
        "t": teacher_username,
        "q1_help": _to_int(payload.get("q1_help")),
        "q2_new": (payload.get("q2_new") or None),
        "q2_detail": (payload.get("q2_detail") or None),
        "q3_use": _to_int(payload.get("q3_use")),
        "q4_cmp": (payload.get("q4_cmp") or None),
        "q4_detail": (payload.get("q4_detail") or None),
        "q5_conf": (payload.get("q5_conf") or None),
        "q6_next": (payload.get("q6_next") or None),
        "q7_feedback": (payload.get("q7_feedback") or None),
    }

    with engine.begin() as conn:
        conn.execute(text("""
            INSERT INTO teacher_surveys
            (class_code, session_id, teacher_username,
             q1_help, q2_new, q2_detail, q3_use, q4_cmp, q4_detail, q5_conf, q6_next, q7_feedback,
             submitted_at)
            VALUES
            (:code, :sid, :t,
             :q1_help, :q2_new, :q2_detail, :q3_use, :q4_cmp, :q4_detail, :q5_conf, :q6_next, :q7_feedback,
             NOW())
            ON CONFLICT (class_code, session_id)
            DO UPDATE SET
              teacher_username = EXCLUDED.teacher_username,
              q1_help = EXCLUDED.q1_help,
              q2_new = EXCLUDED.q2_new,
              q2_detail = EXCLUDED.q2_detail,
              q3_use = EXCLUDED.q3_use,
              q4_cmp = EXCLUDED.q4_cmp,
              q4_detail = EXCLUDED.q4_detail,
              q5_conf = EXCLUDED.q5_conf,
              q6_next = EXCLUDED.q6_next,
              q7_feedback = EXCLUDED.q7_feedback,
              submitted_at = NOW()
        """), row)

        conn.execute(text("""
            INSERT INTO session_finalizations (class_code, session_id, teacher_username, survey_submitted, survey_submitted_at)
            VALUES (:code, :sid, :t, TRUE, NOW())
            ON CONFLICT (class_code, session_id)
            DO UPDATE SET survey_submitted = TRUE,
                          survey_submitted_at = NOW(),
                          teacher_username = EXCLUDED.teacher_username,
                          updated_at = NOW()
        """), {"code": class_code, "sid": sid, "t": teacher_username})


def db_finalize_session_v2(class_code: str, sid: str, teacher_username: str) -> Dict[str, Any]:
    """Mark a session complete once exclusion review and teacher survey are saved."""
    if not engine:
        raise RuntimeError("DB engine not initialized")

    with engine.begin() as conn:
        fin = conn.execute(text("""
            SELECT exclusions_resolved, survey_submitted, finalized
            FROM session_finalizations
            WHERE class_code = :code AND session_id = :sid
        """), {"code": class_code, "sid": sid}).fetchone()

        if not fin:
            raise ValueError("Finalize state not found")

        exclusions_resolved = bool(fin.exclusions_resolved)
        survey_submitted = bool(fin.survey_submitted) or bool(conn.execute(text("""
            SELECT 1 FROM teacher_surveys WHERE class_code=:code AND session_id=:sid
        """), {"code": class_code, "sid": sid}).fetchone())

        if not (exclusions_resolved and survey_submitted):
            return {
                "ok": False,
                "error": "NOT_READY",
                "exclusions_resolved": exclusions_resolved,
                "survey_submitted": survey_submitted,
            }

        conn.execute(text("""
            UPDATE session_finalizations
            SET finalized = TRUE,
                finalized_at = NOW(),
                preview_seen = TRUE,
                preview_seen_at = COALESCE(preview_seen_at, NOW()),
                teacher_username = :t,
                updated_at = NOW()
            WHERE class_code = :code AND session_id = :sid
        """), {"code": class_code, "sid": sid, "t": teacher_username})

        return {"ok": True, "finalized": True}


def build_research_session_summary(class_code: str, sid: str) -> Dict[str, Any]:
    """Build one research/admin summary row for Google Sheets."""
    students = db_get_students_in_class(class_code)
    total = len(students)
    submitted_map = db_get_submitted_map(class_code, sid)
    submitted = sum(1 for s in students if submitted_map.get((s.get("name") or "").strip(), False))
    similarity = teacher_student_similarity_summary(class_code, sid)

    with engine.connect() as conn:
        cls = conn.execute(text("""
            SELECT name, teacher_username
            FROM classes
            WHERE code = :code
            LIMIT 1
        """), {"code": class_code}).fetchone()
        tr = conn.execute(text("""
            SELECT COUNT(*) AS n
            FROM teacher_placement_runs
            WHERE class_code = :code AND session_id = :sid AND submitted = TRUE
        """), {"code": class_code, "sid": sid}).fetchone()
        fin = conn.execute(text("""
            SELECT exclusions_resolved, survey_submitted, finalized
            FROM session_finalizations
            WHERE class_code = :code AND session_id = :sid
            LIMIT 1
        """), {"code": class_code, "sid": sid}).fetchone()
        survey = conn.execute(text("""
            SELECT 1
            FROM teacher_surveys
            WHERE class_code = :code AND session_id = :sid
            LIMIT 1
        """), {"code": class_code, "sid": sid}).fetchone()

    return {
        "class_code": class_code,
        "class_name": cls.name if cls else "",
        "teacher": cls.teacher_username if cls else "",
        "session": str(sid),
        "visible_round": int(sid) - 1 if str(sid).isdigit() else sid,
        "total_students": total,
        "submitted_students": submitted,
        "teacher_placement_done": int(tr.n or 0) > 0 if tr else False,
        "exclusions_resolved": bool(getattr(fin, "exclusions_resolved", False)) if fin else False,
        "post_survey_done": (bool(getattr(fin, "survey_submitted", False)) if fin else False) or bool(survey),
        "finalized": bool(getattr(fin, "finalized", False)) if fin else False,
        "similarity_label": similarity.get("label"),
        "similarity_score": similarity.get("score"),
        "similarity_pairs": similarity.get("n_pairs"),
        "generated_at": datetime.utcnow().isoformat() + "Z",
    }


def db_archive_and_reset_session(class_code: str, sid: str, reset_by: str, reason: str = "", scope: str = "all") -> int:
    """Archive selected session data as JSON, then clear that active data."""
    if not engine:
        raise RuntimeError("DB engine not initialized")

    scope = (scope or "all").strip()
    allowed_scopes = {"all", "student", "teacher", "pre_survey", "post_survey"}
    if scope not in allowed_scopes:
        scope = "all"

    with engine.begin() as conn:
        teacher_runs = conn.execute(text("""
            SELECT *
            FROM teacher_placement_runs
            WHERE class_code = :code AND session_id = :sid
            ORDER BY id ASC
        """), {"code": class_code, "sid": sid}).mappings().all()
        run_ids = [int(r["id"]) for r in teacher_runs]

        teacher_decisions = []
        if run_ids:
            teacher_decisions = conn.execute(text("""
                SELECT *
                FROM teacher_decisions
                WHERE run_id = ANY(:run_ids)
                ORDER BY run_id ASC, priority_rank ASC
            """), {"run_ids": run_ids}).mappings().all()

        def rows(sql: str) -> List[Dict[str, Any]]:
            return [dict(r) for r in conn.execute(text(sql), {"code": class_code, "sid": sid}).mappings().all()]

        payload = {
            "class_code": class_code,
            "session_id": sid,
            "reset_by": reset_by,
            "reset_reason": reason,
            "reset_scope": scope,
            "archived_at": datetime.utcnow().isoformat() + "Z",
            "student_sessions": rows("SELECT * FROM student_sessions WHERE class_code=:code AND sid=:sid ORDER BY id ASC") if scope in {"all", "student"} else [],
            "teacher_placement_runs": [dict(r) for r in teacher_runs] if scope in {"all", "teacher", "pre_survey"} else [],
            "teacher_decisions": [dict(r) for r in teacher_decisions] if scope in {"all", "teacher", "pre_survey"} else [],
            "session_exclusions": rows("SELECT * FROM session_exclusions WHERE class_code=:code AND session_id=:sid ORDER BY id ASC") if scope in {"all", "student"} else [],
            "session_finalizations": rows("SELECT * FROM session_finalizations WHERE class_code=:code AND session_id=:sid ORDER BY id ASC") if scope in {"all", "student", "post_survey"} else [],
            "teacher_surveys": rows("SELECT * FROM teacher_surveys WHERE class_code=:code AND session_id=:sid ORDER BY id ASC") if scope in {"all", "post_survey"} else [],
            "analysis_cache": rows("SELECT * FROM analysis_cache WHERE class_code=:code AND session_id=:sid ORDER BY id ASC") if scope in {"all", "student", "teacher"} else [],
        }
        payload = json.loads(json.dumps(payload, ensure_ascii=False, default=str))

        res = conn.execute(text("""
            INSERT INTO archived_session_resets
            (class_code, session_id, reset_by, reset_reason, archived_payload)
            VALUES (:code, :sid, :reset_by, :reason, CAST(:payload AS jsonb))
            RETURNING id
        """), {
            "code": class_code,
            "sid": sid,
            "reset_by": reset_by,
            "reason": reason,
            "payload": json.dumps(payload, ensure_ascii=False),
        }).fetchone()
        archive_id = int(res.id)

        if scope in {"all", "teacher", "pre_survey"} and run_ids:
            conn.execute(text("DELETE FROM teacher_decisions WHERE run_id = ANY(:run_ids)"), {"run_ids": run_ids})
        if scope == "pre_survey":
            conn.execute(text("""
                UPDATE teacher_placement_runs
                SET confidence_score = NULL
                WHERE class_code=:code AND session_id=:sid
            """), {"code": class_code, "sid": sid})
        if scope in {"all", "teacher"}:
            conn.execute(text("DELETE FROM teacher_placement_runs WHERE class_code=:code AND session_id=:sid"), {"code": class_code, "sid": sid})
        if scope in {"all", "student"}:
            conn.execute(text("DELETE FROM student_sessions WHERE class_code=:code AND sid=:sid"), {"code": class_code, "sid": sid})
            conn.execute(text("DELETE FROM session_exclusions WHERE class_code=:code AND session_id=:sid"), {"code": class_code, "sid": sid})
        if scope in {"all", "post_survey"}:
            conn.execute(text("DELETE FROM teacher_surveys WHERE class_code=:code AND session_id=:sid"), {"code": class_code, "sid": sid})
        if scope == "all":
            conn.execute(text("DELETE FROM session_finalizations WHERE class_code=:code AND session_id=:sid"), {"code": class_code, "sid": sid})
        elif scope == "student":
            conn.execute(text("""
                UPDATE session_finalizations
                SET exclusions_resolved = FALSE,
                    finalized = FALSE,
                    updated_at = NOW()
                WHERE class_code=:code AND session_id=:sid
            """), {"code": class_code, "sid": sid})
        elif scope == "post_survey":
            conn.execute(text("""
                UPDATE session_finalizations
                SET survey_submitted = FALSE,
                    finalized = FALSE,
                    updated_at = NOW()
                WHERE class_code=:code AND session_id=:sid
            """), {"code": class_code, "sid": sid})
        if scope in {"all", "student", "teacher"}:
            conn.execute(text("DELETE FROM analysis_cache WHERE class_code=:code AND session_id=:sid"), {"code": class_code, "sid": sid})

    try:
        sheet_append_archived_reset({
            "archive_id": archive_id,
            "class_code": class_code,
            "session": str(sid),
            "reset_by": reset_by,
            "reset_reason": reason,
            "reset_scope": scope,
            "archived_payload": payload,
        })
    except Exception:
        app.logger.exception("sheet archive reset append failed")

    return archive_id


def db_get_student_session(class_code: str, student_name: str, sid: str) -> Optional[Dict[str, Any]]:
    if not engine:
        return None

    with engine.connect() as conn:
        row = conn.execute(text("""
            SELECT placements, placements_json, submitted
            FROM student_sessions
            WHERE class_code = :code
              AND student_name = :name
              AND sid = :sid
            LIMIT 1
        """), {"code": class_code, "name": student_name, "sid": sid}).fetchone()

    if not row:
        return None

    placements_obj: Dict[str, Any] = {}
    if row.placements is not None:
        placements_obj = row.placements if isinstance(row.placements, dict) else _json_load_maybe(row.placements)
    elif row.placements_json:
        placements_obj = _json_load_maybe(row.placements_json)

    return {"placements": placements_obj, "submitted": bool(row.submitted)}


def db_list_submitted_student_sessions(class_code: str, sid: str) -> List[Dict[str, Any]]:
    if not engine:
        return []

    with engine.connect() as conn:
        rows = conn.execute(text("""
            SELECT student_name, placements, placements_json
            FROM student_sessions
            WHERE class_code = :code
              AND sid = :sid
              AND submitted = TRUE
            ORDER BY id ASC
        """), {"code": class_code, "sid": sid}).fetchall()

    out: List[Dict[str, Any]] = []
    for r in rows:
        placements_obj: Dict[str, Any] = {}
        if r.placements is not None:
            placements_obj = r.placements if isinstance(r.placements, dict) else _json_load_maybe(r.placements)
        elif r.placements_json:
            placements_obj = _json_load_maybe(r.placements_json)
        out.append({"student_name": r.student_name, "placements": placements_obj})
    return out

def db_upsert_student_session(class_code: str, student_name: str, sid: str, placements: Dict[str, Any], submitted: bool) -> None:
    if not engine:
        raise RuntimeError("DB engine not initialized")

    placements_str = json.dumps(placements, ensure_ascii=False)

    with engine.begin() as conn:
        # 운영 DB 스키마가 서로 다를 수 있어 session_id 컬럼 존재 여부를 확인한다.
        has_session_id = conn.execute(text("""
            SELECT 1
            FROM information_schema.columns
            WHERE table_name = 'student_sessions' AND column_name = 'session_id'
            LIMIT 1
        """)).fetchone() is not None

        # session_id 타입이 integer가 아닌(text 등) 환경이 실제로 존재하므로
        # CAST를 제거하고 텍스트 비교로 처리한다.
        if has_session_id:
            row = conn.execute(text("""
                SELECT id
                FROM student_sessions
                WHERE class_code = :code
                  AND student_name = :name
                  AND (sid = :sid OR session_id = :sid)
                LIMIT 1
            """), {"code": class_code, "name": student_name, "sid": sid}).fetchone()
        else:
            row = conn.execute(text("""
                SELECT id
                FROM student_sessions
                WHERE class_code = :code AND student_name = :name AND sid = :sid
                LIMIT 1
            """), {"code": class_code, "name": student_name, "sid": sid}).fetchone()

        if row:
            if has_session_id:
                conn.execute(text("""
                    UPDATE student_sessions
                    SET sid = :sid,
                        session_id = :sid,
                        placements = CAST(:placements AS jsonb),
                        placements_json = :placements_json,
                        submitted = :submitted
                    WHERE id = :id
                """), {
                    "sid": sid,
                    "placements": placements_str,
                    "placements_json": placements_str,
                    "submitted": submitted,
                    "id": row.id
                })
            else:
                conn.execute(text("""
                    UPDATE student_sessions
                    SET placements = CAST(:placements AS jsonb),
                        placements_json = :placements_json,
                        submitted = :submitted
                    WHERE id = :id
                """), {
                    "placements": placements_str,
                    "placements_json": placements_str,
                    "submitted": submitted,
                    "id": row.id
                })
        else:
            if has_session_id:
                conn.execute(text("""
                    INSERT INTO student_sessions (
                        class_code, sid, session_id, student_name,
                        placements, placements_json, submitted
                    )
                    VALUES (
                        :code, :sid, :sid, :name,
                        CAST(:placements AS jsonb), :placements_json, :submitted
                    )
                """), {
                    "code": class_code,
                    "sid": sid,
                    "name": student_name,
                    "placements": placements_str,
                    "placements_json": placements_str,
                    "submitted": submitted
                })
            else:
                conn.execute(text("""
                    INSERT INTO student_sessions (class_code, sid, student_name, placements, placements_json, submitted)
                    VALUES (:code, :sid, :name, CAST(:placements AS jsonb), :placements_json, :submitted)
                """), {
                    "code": class_code,
                    "sid": sid,
                    "name": student_name,
                    "placements": placements_str,
                    "placements_json": placements_str,
                    "submitted": submitted
                })

    if submitted:
        cache_clear_session_analysis(class_code, sid)


      
def db_create_teacher_run(class_code: str, teacher_username: str, sid: str, condition: str, tool_run_id: Optional[int] = None) -> int:
    if not engine:
        raise RuntimeError("DB engine not initialized")

    with engine.begin() as conn:
        row = conn.execute(text("""
            INSERT INTO teacher_placement_runs
            (class_code, teacher_username, session_id, condition, tool_run_id, placements, placements_json, submitted, started_at)
            VALUES (:code, :t, :sid, :cond, :tool_run_id, CAST(:placements AS jsonb), :placements_json, FALSE, NOW())
            RETURNING id
        """), {
            "code": class_code,
            "t": teacher_username,
            "sid": sid,
            "cond": condition,
            "tool_run_id": tool_run_id,
            "placements": json.dumps({}, ensure_ascii=False),
            "placements_json": json.dumps({}, ensure_ascii=False),
        }).fetchone()

    return int(row[0])




def db_get_teacher_run(run_id: int) -> Optional[Dict[str, Any]]:
    if not engine:
        return None

    with engine.connect() as conn:
        row = conn.execute(text("""
            SELECT id, class_code, teacher_username, session_id, condition, tool_run_id,
                   placements, placements_json, submitted, started_at, ended_at, duration_ms, confidence_score
            FROM teacher_placement_runs
            WHERE id = :id
            LIMIT 1
        """), {"id": run_id}).fetchone()

    if not row:
        return None

    placements_obj: Dict[str, Any] = {}
    if row.placements is not None:
        placements_obj = row.placements if isinstance(row.placements, dict) else _json_load_maybe(row.placements)
    elif row.placements_json:
        placements_obj = _json_load_maybe(row.placements_json)

    return {
        "id": row.id,
        "class_code": row.class_code,
        "teacher_username": row.teacher_username,
        "session_id": row.session_id,
        "condition": row.condition,
        "tool_run_id": row.tool_run_id,
        "placements": placements_obj,
        "submitted": bool(row.submitted),
        "started_at": row.started_at,
        "ended_at": row.ended_at,
        "duration_ms": row.duration_ms,
        "confidence_score": row.confidence_score,
    }


def db_update_teacher_run_placements(run_id: int, placements: Dict[str, Any]) -> None:
    if not engine:
        raise RuntimeError("DB engine not initialized")

    placements_str = json.dumps(placements, ensure_ascii=False)

    with engine.begin() as conn:
        conn.execute(text("""
            UPDATE teacher_placement_runs
            SET placements = CAST(:placements AS jsonb),
                placements_json = :placements_json
            WHERE id = :id
        """), {"placements": placements_str, "placements_json": placements_str, "id": run_id})



def db_complete_teacher_run(run_id: int, confidence_score: int) -> None:
    if not engine:
        raise RuntimeError("DB engine not initialized")

    with engine.begin() as conn:
        conn.execute(text("""
            UPDATE teacher_placement_runs
            SET ended_at = NOW(),
                confidence_score = :confidence_score,
                submitted = TRUE
            WHERE id = :id
        """), {
            "confidence_score": int(confidence_score),
            "id": run_id
        })

def sync_results_from_sheet_to_db(class_code: str, sid: str, teacher_username: str) -> Dict[str, int]:
    """
    Google Sheets Results의 데이터를 Postgres로 동기화한다.
    - 학생: student_sessions (submitted = TRUE)
    - 교사: teacher_placement_runs (completed)
    """
    rows = sheet_list_results(class_code, sid)

    synced_students = 0
    synced_teacher = 0

    for r in rows:
        student = (r.get("student") or "").strip()
        placements_raw = r.get("placements") or "{}"

        try:
            placements = json.loads(placements_raw) if isinstance(placements_raw, str) else placements_raw
        except Exception:
            placements = {}

        # 교사 관찰
        if student == "teacher":
            run_id = db_create_teacher_run(
                class_code=class_code,
                teacher_username=teacher_username,
                sid=str(sid),
                condition="sheet_import"   # 테스트/연구용 명시
            )
            db_update_teacher_run_placements(run_id, placements)
            db_complete_teacher_run(run_id, confidence_score=0)
            synced_teacher += 1
            continue

        # 학생 인식
        if student:
            db_upsert_student_session(
                class_code=class_code,
                student_name=student,
                sid=str(sid),
                placements=placements,
                submitted=True
            )
            synced_students += 1

    return {
        "teacher_runs": synced_teacher,
        "student_rows": synced_students
    }


def db_replace_teacher_decisions(run_id: int, decisions: List[Dict[str, Any]]) -> None:
    if not engine:
        raise RuntimeError("DB engine not initialized")

    with engine.begin() as conn:
        conn.execute(text("DELETE FROM teacher_decisions WHERE run_id = :run_id"), {"run_id": run_id})
        for d in decisions:
            conn.execute(text("""
                INSERT INTO teacher_decisions
                (run_id, target_student_name, priority_rank, decision_confidence, reason_tags)
                VALUES (:run_id, :name, :rank, :conf, CAST(:tags AS jsonb))
            """), {
                "run_id": run_id,
                "name": (d.get("name") or "").strip(),
                "rank": int(d.get("rank") or 0),
                "conf": int(d.get("confidence") or 0) if d.get("confidence") is not None else None,
                "tags": json.dumps(d.get("tags") or [], ensure_ascii=False),
            })



# -------------------------
# Cache helpers (analysis_cache)
# -------------------------

def cache_get(class_code: str, sid: str, key: str) -> Optional[Dict[str, Any]]:
    if not engine:
        return None
    with engine.connect() as conn:
        row = conn.execute(text("""
            SELECT payload
            FROM analysis_cache
            WHERE class_code = :c AND session_id = :s AND cache_key = :k
            LIMIT 1
        """), {"c": class_code, "s": sid, "k": key}).fetchone()
    if not row:
        return None
    if isinstance(row.payload, dict):
        return row.payload
    return _json_load_maybe(row.payload)


def cache_set(class_code: str, sid: str, key: str, payload: Dict[str, Any]) -> None:
    if not engine:
        return
    payload_str = json.dumps(payload, ensure_ascii=False)
    with engine.begin() as conn:
        conn.execute(text("""
            INSERT INTO analysis_cache (class_code, session_id, cache_key, payload, updated_at)
            VALUES (:c, :s, :k, CAST(:p AS jsonb), NOW())
            ON CONFLICT (class_code, session_id, cache_key)
            DO UPDATE SET payload = EXCLUDED.payload, updated_at = NOW()
        """), {"c": class_code, "s": sid, "k": key, "p": payload_str})


def cache_clear_session_analysis(class_code: str, sid: str) -> None:
    if not engine:
        return
    with engine.begin() as conn:
        conn.execute(text("""
            DELETE FROM analysis_cache
            WHERE class_code = :c
              AND session_id = :s
              AND (
                cache_key = :avg_key
                OR cache_key = :dbscan_key
                OR cache_key = :spm_key
                OR cache_key LIKE :kmeans_prefix
                OR cache_key LIKE :student_vs_prefix
              )
        """), {
            "c": class_code,
            "s": sid,
            "avg_key": f"student_avg_{sid}",
            "dbscan_key": f"dbscan_structure_{sid}",
            "spm_key": f"spm_result_{sid}_v1",
            "kmeans_prefix": f"kmeans_summary_{sid}_k%",
            "student_vs_prefix": f"student_vs_avg_{sid}_%",
        })



# -------------------------
# Context: current class (topbar)
# -------------------------

def get_current_class() -> Optional[Dict[str, str]]:
    code = None
    if "teacher" in session and session.get("selected_class"):
        code = session.get("selected_class")
    elif session.get("code"):
        code = session.get("code")

    if not code:
        return None

    if engine:
        name = db_get_class_name(code)
        if not name:
            return None
        return {"name": name, "code": code}

    d = load_data()
    cls = d.get("classes", {}).get(code)
    if not cls:
        return None
    return {"name": cls.get("name", ""), "code": code}


@app.context_processor
def inject_globals() -> Dict[str, Any]:
    return {"current_class": get_current_class()}


# -------------------------
# Debug routes
# -------------------------



def build_student_pin_pdf(class_name: str, sid: str, students):
    # ---- Lazy imports: prevent whole app from failing if reportlab not installed
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.pdfgen import canvas
    from reportlab.pdfbase.pdfmetrics import stringWidth
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    import os

    # ---- Font (Korean) setup
    # Put font file at: static/fonts/NotoSansKR-Regular.ttf
    FONT_REG = "NotoSansKR"
    FONT_BOLD = "NotoSansKR-Bold"  # optional; if you don't have bold file, we'll reuse regular

    base_dir = os.path.dirname(os.path.abspath(__file__))
    font_path = os.path.join(base_dir, "static", "fonts", "NotoSansKR-Regular.ttf")
    bold_path = os.path.join(base_dir, "static", "fonts", "NotoSansKR-Bold.ttf")

    # Register once (safe even if called multiple times)
    try:
        if FONT_REG not in pdfmetrics.getRegisteredFontNames():
            pdfmetrics.registerFont(TTFont(FONT_REG, font_path))
        if os.path.exists(bold_path) and (FONT_BOLD not in pdfmetrics.getRegisteredFontNames()):
            pdfmetrics.registerFont(TTFont(FONT_BOLD, bold_path))
        has_korean_font = True
    except Exception:
        # Fallback: Helvetica (will show □ for Korean, but prevents crash)
        has_korean_font = False

    # Choose fonts for use below
    F_REG = FONT_REG if has_korean_font else "Helvetica"
    F_BOLD = (FONT_BOLD if (has_korean_font and os.path.exists(bold_path)) else F_REG) if has_korean_font else "Helvetica-Bold"


    buf = BytesIO()
    page_w, page_h = landscape(A4)
    c = canvas.Canvas(buf, pagesize=(page_w, page_h))

    tz = ZoneInfo("Asia/Seoul")
    date_str = datetime.now(tz).strftime("%Y.%m.%d")
    """
    Builds a paged PDF (10 students per page) for student login PIN codes.

    Layout requirements (your spec):
    - A4 landscape
    - Top 60%: title + class/sid/date + "학생 로그인 PIN 코드 인" only (no student codes here)
    - Bottom 40%: one row containing 10 students per page
    - No solid outlines / no horizontal lines
    - Font sizes ~200% (already reflected)
    - PIN is rotated 90 degrees and has underline to avoid 6/9 confusion
    """
    buf = BytesIO()
    page_w, page_h = landscape(A4)
    c = canvas.Canvas(buf, pagesize=(page_w, page_h))

    tz = ZoneInfo("Asia/Seoul")
    date_str = datetime.now(tz).strftime("%Y.%m.%d")

    # Regions
    top_h = page_h * 0.60
    bottom_h = page_h * 0.40

    # Margins
    margin_x = 36
    margin_top = 28
    bottom_y0 = 24
    bottom_y1 = bottom_y0 + bottom_h

    # ~200% sizing
    title_fs = 34
    meta_fs = 18
    sub_fs = 20
    name_fs = 22
    pin_fs = 26

    per_page = 10

    def draw_pin_cell(x0, x1, y0, y1, name, pin_code):
        cell_w = x1 - x0
        cell_h = y1 - y0

        # ---- Name (horizontal)
        name_text = (name or "").strip()
        c.setFont(F_BOLD, name_fs)

        max_w = cell_w * 0.90
        if stringWidth(name_text, F_BOLD, name_fs) > max_w:
            fs = name_fs
            while fs > 12 and stringWidth(name_text, F_BOLD, fs) > max_w:
                fs -= 1
            c.setFont(F_BOLD, fs)

        name_y = y0 + cell_h * 0.70
        c.drawCentredString((x0 + x1) / 2, name_y, name_text)

        # ---- PIN (rotated 90deg) + underline
        pin = (pin_code or "").strip()

        cx = (x0 + x1) / 2
        cy = y0 + cell_h * 0.28

        c.saveState()
        c.translate(cx, cy)
        c.rotate(90)
        c.setFont(F_BOLD, pin_fs)

        text_w = stringWidth(pin, F_BOLD, pin_fs)
        c.drawString(-text_w / 2, 0, pin)


        # underline under rotated PIN (to distinguish 6/9)
        underline_y = -3
        c.setLineWidth(1.5)
        c.line(-text_w / 2, underline_y, text_w / 2, underline_y)
        c.restoreState()

    def draw_header():
        # Title
        c.setFont(F_BOLD, title_fs)
        c.drawCentredString(page_w / 2, page_h - margin_top - 6, "<우리반 관계 지도>")

        # Class + SID + Date
        c.setFont(F_BOLD, meta_fs)
        meta_line = f"{class_name}  |  {sid}회차  |  {date_str}"
        c.drawCentredString(page_w / 2, page_h - margin_top - 42, meta_line)

        # Sub label
        c.setFont(F_BOLD, sub_fs)
        c.drawCentredString(page_w / 2, page_h - margin_top - 70, "학생 로그인 PIN 코드")

    # Defensive cleanup + stable order
    # students: [{no, name, pin_code}, ...]
    safe_students = []
    for s in (students or []):
        nm = (s.get("name") or "").strip()
        if not nm:
            continue
        safe_students.append({
            "no": s.get("no"),
            "name": nm,
            "pin_code": (s.get("pin_code") or "").strip(),  # <-- KEY FIX: pin_code 고정
        })

    # Chunk per page
    total = len(safe_students)
    pages = (total + per_page - 1) // per_page
    if pages == 0:
        pages = 1

    for p in range(pages):
        draw_header()

        # Bottom row geometry: 10 equal cells
        x0 = margin_x
        x1 = page_w - margin_x
        full_w = x1 - x0
        cell_w = full_w / per_page

        # Use full bottom region height
        y0 = bottom_y0
        y1 = bottom_y1

        start = p * per_page
        end = min(start + per_page, total)
        page_items = safe_students[start:end]

        # Draw 10 slots even if less students on last page (empty slots remain blank)
        for i in range(per_page):
            left = x0 + i * cell_w
            right = left + cell_w
            if i < len(page_items):
                it = page_items[i]
                draw_pin_cell(left, right, y0, y1, it["name"], it["pin_code"])
            else:
                # Empty slot: intentionally draw nothing (no outlines/lines)
                pass

        # --- Vertical dashed separators between students (like your screenshot)
        # Draw separators at each cell boundary (except outer edges)
        c.saveState()
        c.setStrokeColorRGB(0.6, 0.6, 0.6)  # light gray
        c.setLineWidth(1)
        c.setDash(1, 3)  # dotted/dashed pattern (dash, gap)

        sep_y0 = y0 + 6
        sep_y1 = y1 - 6

        for i in range(1, per_page):
            x = x0 + i * cell_w
            c.line(x, sep_y0, x, sep_y1)

        c.restoreState()

        c.showPage()

    c.save()
    buf.seek(0)
    return buf


@app.route("/debug/db")
def debug_db():
    if not engine:
        return "DATABASE_URL not set"
    try:
        with engine.connect() as conn:
            conn.execute(text("SELECT 1"))
        return "DB connection OK"
    except Exception as e:
        return f"DB connection failed: {e}", 500


if DEBUG_MODE:

    @app.route("/debug/schema_version")
    def debug_schema_version():
        if not engine:
            return jsonify({"error": "DATABASE_URL not set"}), 500
        try:
            with engine.connect() as conn:
                row = conn.execute(
                    text("SELECT version, updated_at FROM schema_migrations WHERE id = 1")
                ).fetchone()
            if not row:
                return jsonify({"version": None, "updated_at": None})
            return jsonify({
                "version": int(row.version),
                "updated_at": row.updated_at.isoformat() if row.updated_at else None,
            })
        except Exception as e:
            return jsonify({"error": str(e)}), 500


    @app.route("/debug/versions")
    def debug_versions():
        import sys

        out = {
            "python_version": sys.version,
            "python_version_info": list(sys.version_info),
            "openpyxl_available": bool(OPENPYXL_AVAILABLE and (Workbook is not None)),
        }

        # openpyxl 버전도 확인 (설치돼 있을 때만)
        if OPENPYXL_AVAILABLE:
            try:
                import openpyxl  # type: ignore
                out["openpyxl_version"] = getattr(openpyxl, "__version__", None)
            except Exception as e:
                out["openpyxl_version_error"] = str(e)

        return jsonify(out)



# ---------- 헬스 체크 (콜드 스타트 방지용) ----------
@app.route("/health")
def health():
    # 기본 응답은 가볍게 유지하되, 내부적으로는 핵심 의존성 상태를 함께 확인 가능
    status = {
        "ok": True,
        "openpyxl_available": bool(OPENPYXL_AVAILABLE and (Workbook is not None)),
    }

    # DB는 설정돼 있으면 가볍게 ping (실패해도 앱 전체는 살리고 상태만 표시)
    if engine:
        try:
            with engine.connect() as conn:
                conn.execute(text("SELECT 1"))
            status["db_ok"] = True
        except Exception as e:
            status["db_ok"] = False
            status["db_error"] = str(e)
    else:
        status["db_ok"] = False
        status["db_error"] = "DATABASE_URL not set"

    # openpyxl이 “import 됐고 Workbook이 살아있는지”가 핵심
    http_code = 200 if status["openpyxl_available"] else 500
    return jsonify(status), http_code




# -------------------------
# Home
# -------------------------

@app.route("/")
def home():
    # 로그인 유지 중이면 home -> dashboard 이동으로 고정
    if session.get("teacher"):
        return redirect("/teacher/dashboard")

    return render_template("home.html", site_title=SITE_TITLE)



# -------------------------
# Teacher auth
# -------------------------

@app.route("/teacher/signup", methods=["GET", "POST"])
def teacher_signup():
    if request.method == "POST":
        username = request.form.get("username", "").strip()
        pw = request.form.get("password", "")
        pw2 = request.form.get("password2", "")

        if not username or not pw:
            return render_template("teacher_signup.html", error="아이디/비밀번호를 입력해 주세요.")
        if pw != pw2:
            return render_template("teacher_signup.html", error="비밀번호가 서로 다릅니다.")

        # =========================================================
        # RESEARCH_ONLY_BEGIN
        # 논문 작성용 추가 수집(추후 삭제 예정)
        # =========================================================
        teaching_years = request.form.get("teaching_years", "").strip()
        current_grade = request.form.get("current_grade", "").strip()
        research_name = request.form.get("research_name", "").strip()
        research_school = request.form.get("research_school", "").strip()

        if not teaching_years:
            return render_template("teacher_signup.html", error="교직경력을 선택해 주세요.")
        if not current_grade:
            return render_template("teacher_signup.html", error="담당 학년을 선택해 주세요.")
        if not research_name:
            return render_template("teacher_signup.html", error="이름을 입력해 주세요.")
        if not research_school:
            return render_template("teacher_signup.html", error="학교를 입력해 주세요.")
        # =========================================================
        # RESEARCH_ONLY_END
        # =========================================================

        pw_hash = generate_password_hash(pw)
        profile = {
            "teaching_years": teaching_years,
            "current_grade": current_grade,
            "research_name": research_name,
            "research_school": research_school,
        }

        if not GOOGLE_SECRET:
            resp = local_create_teacher(username, pw_hash, profile)
            if resp.get("status") == "ok":
                return redirect("/teacher/login")
            if resp.get("status") == "exists":
                return render_template("teacher_signup.html", error="이미 존재하는 아이디입니다.")
            return render_template("teacher_signup.html", error=f"회원가입 실패: {resp}")

        try:
            resp = post_to_sheet({
                "action": "teacher_signup",
                "username": username,
                "pw_hash": pw_hash,

                # =========================================================
                # RESEARCH_ONLY_BEGIN
                # 논문 작성용 추가 수집(추후 삭제 예정)
                # =========================================================
                "teaching_years": teaching_years,      # "1-5" / "6-10" / "11-15" / "16+"
                "current_grade": current_grade,        # "4" / "5" / "6"
                "research_name": research_name,        # 필수
                "research_school": research_school,    # 필수
                # =========================================================
                # RESEARCH_ONLY_END
                # =========================================================
            })
        except Exception as e:
            return render_template("teacher_signup.html", error=f"서버 통신 오류: {e}")

        status = resp.get("status")
        if status == "ok":
            return redirect("/teacher/login")
        if status == "exists":
            return render_template("teacher_signup.html", error="이미 존재하는 아이디입니다.")
        if status == "blocked":
            return render_template("teacher_signup.html", error="blocked: GOOGLE_SECRET(비밀키) 불일치 또는 누락")
        return render_template("teacher_signup.html", error=f"회원가입 실패: {resp}")

    return render_template("teacher_signup.html")


@app.route("/teacher/login", methods=["GET", "POST"])
def teacher_login():
    # 로그인 유지 중이면 login 화면을 거치지 않고 home으로
    # (home이 다시 dashboard로 보내므로 경로가 home -> dashboard로 고정됨)
    if request.method == "GET" and session.get("teacher"):
        return redirect("/")

    if request.method == "POST":
        username = request.form.get("username", "").strip()
        pw = request.form.get("password", "")

        if not GOOGLE_SECRET:
            resp = local_get_teacher(username)
        else:
            try:
                resp = post_to_sheet({"action": "teacher_get", "username": username})
            except Exception as e:
                return render_template("teacher_login.html", error=f"서버 통신 오류: {e}")

        try:
            if resp.get("status") != "ok":
                return render_template("teacher_login.html", error=f"로그인 실패: {resp}")

            pw_hash = resp.get("pw_hash") or ""
            if check_password_hash(pw_hash, pw):
                session.clear()
                session["teacher"] = username

                # 바로 dashboard로 보내지 않고 home으로 보냄
                # (주소 이동: home -> dashboard)
                return redirect("/")

            return render_template("teacher_login.html", error="로그인 실패: 비밀번호 불일치")
        except Exception as e:
            return render_template("teacher_login.html", error=f"로그인 처리 중 오류: {e} / resp={resp}")

    return render_template("teacher_login.html")

@app.route("/teacher/class/<class_code>/session/<sid>/resume")
def teacher_resume_session(class_code: str, sid: str):
    if "teacher" not in session:
        return redirect("/teacher/login")

    # 1) 해당 teacher/class/sid의 최신 run 찾기 (없으면 생성)
    run_id = db_get_latest_teacher_run_id(class_code, session["teacher"], sid)
    if not run_id:
       run_id = db_create_teacher_run(class_code, session["teacher"], sid, condition="BASELINE", tool_run_id=None)

    run = db_get_teacher_run(run_id)
    if not run:
        return redirect(f"/teacher/class/{class_code}?sid={sid}")

    # 2) 상태에 따라 이어갈 곳 자동 결정
    if run.get("submitted"):
        return redirect(f"/teacher/class/{class_code}?sid={sid}")

    # placements_complete 판단은 teacher_write 라우트에서 쓰는 로직과 동일하게 계산
    students = db_get_students_in_class(class_code) if engine else []
    all_names = [s["name"] for s in students]
    placements = run.get("placements") or {}

    placements_complete = False
    try:
        placements_complete = all((n in placements) for n in all_names) and len(all_names) > 0
    except Exception:
        placements_complete = False

    if placements_complete:
        return redirect(f"/teacher/placement/{run_id}/complete")

    return redirect(f"/teacher/placement/{run_id}")



@app.route("/teacher/logout")
def teacher_logout():
    session.clear()
    return redirect("/")

# -------------------------
# Research admin pages (owner-only)
# -------------------------

@app.route("/research")
def research_admin():
    guard = require_admin()
    if guard is not None:
        return guard

    if not engine:
        return render_template("research_admin.html", db_ready=False, overview=[])

    overview = db_fetch_class_overview()
    return render_template("research_admin.html", db_ready=True, overview=overview)


@app.route("/research/class/<code>/session/<sid>/sync_sheet", methods=["POST"])
def research_sync_session_sheet(code: str, sid: str):
    guard = require_admin()
    if guard is not None:
        return guard
    if not engine:
        return jsonify({"ok": False, "error": "DB_NOT_CONFIGURED"}), 400

    code = (code or "").upper().strip()
    sid = normalize_visible_session_id(sid)
    payload = build_research_session_summary(code, sid)
    resp = sheet_upsert_research_session_summary(payload)
    if resp.get("status") != "ok":
        return jsonify({"ok": False, "error": "SHEET_ERROR", "sheet": resp}), 500
    return jsonify({"ok": True, "sheet": resp, "payload": payload})


@app.route("/research/class/<code>/rename", methods=["POST"])
def research_rename_class(code: str):
    guard = require_admin()
    if guard is not None:
        return guard
    if not engine:
        return jsonify({"ok": False, "error": "DB_NOT_CONFIGURED"}), 400

    code = (code or "").upper().strip()
    new_name = (request.form.get("class_name") or "").strip()
    if not new_name:
        return jsonify({"ok": False, "error": "학급 이름을 입력해 주세요."}), 400

    ok = db_update_class_name_admin(code, new_name, session.get("teacher") or "admin")
    if not ok:
        return jsonify({"ok": False, "error": "학급을 찾을 수 없습니다."}), 404
    return jsonify({"ok": True, "class_code": code, "class_name": new_name})


@app.route("/research/class/<code>/archive", methods=["POST"])
def research_archive_class(code: str):
    guard = require_admin()
    if guard is not None:
        return guard
    if not engine:
        return jsonify({"ok": False, "error": "DB_NOT_CONFIGURED"}), 400

    code = (code or "").upper().strip()
    archived_raw = (request.form.get("archived") or "1").strip().lower()
    archived = archived_raw in {"1", "true", "yes", "on"}
    reason = (request.form.get("reason") or "").strip()
    ok = db_set_class_archived_admin(code, archived, session.get("teacher") or "admin", reason=reason)
    if not ok:
        return jsonify({"ok": False, "error": "학급을 찾을 수 없습니다."}), 404
    return jsonify({"ok": True, "class_code": code, "archived": archived})


@app.route("/research/class/<code>/session/<sid>/reset", methods=["POST"])
def research_reset_session(code: str, sid: str):
    guard = require_admin()
    if guard is not None:
        return guard
    if not engine:
        return jsonify({"ok": False, "error": "DB_NOT_CONFIGURED"}), 400

    code = (code or "").upper().strip()
    sid = normalize_visible_session_id(sid)
    reason = (request.form.get("reason") or "").strip()
    scope = (request.form.get("scope") or "all").strip()
    archive_id = db_archive_and_reset_session(code, sid, session.get("teacher") or "admin", reason, scope=scope)
    return jsonify({"ok": True, "archive_id": archive_id, "scope": scope})


@app.route("/research/export/student_sessions.xlsx")
def export_student_sessions_xlsx():
    guard = require_admin()
    if guard is not None:
        return guard

    # XLSX 기능은 openpyxl 의존: 미설치면 라우트만 500으로 종료
    if not OPENPYXL_AVAILABLE or Workbook is None:
        return "openpyxl not installed on server", 500

    if not engine:
        return "DB not configured", 400

    class_code = (request.args.get("class_code") or "").strip().upper()
    sid = (request.args.get("sid") or "").strip()
    if not class_code or not sid:
        return "class_code and sid are required", 400

    try:
        with engine.connect() as conn:
            rows = conn.execute(text("""
                SELECT class_code, sid, student_name, submitted, confidence, priority, created_at, placements
                FROM student_sessions
                WHERE class_code = :code AND sid = :sid
                ORDER BY student_name ASC
            """), {"code": class_code, "sid": sid}).fetchall()
    except Exception as e:
        return f"DB query failed: {e}", 500

    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "student_sessions"

        headers = ["class_code", "sid", "student_name", "submitted", "confidence", "priority", "created_at", "placements_json"]
        ws.append(headers)

        for r in rows:
            placements_json = ""
            try:
                placements_json = json.dumps(
                    r.placements if isinstance(r.placements, dict) else (r.placements or {}),
                    ensure_ascii=False
                )
            except Exception:
                placements_json = ""

            ws.append([
                r.class_code,
                r.sid,
                r.student_name,
                bool(r.submitted),
                r.confidence,
                r.priority,
                r.created_at.isoformat() if r.created_at else None,
                placements_json,
            ])

        _autosize_columns(ws)
        return _xlsx_response(wb, f"student_sessions_{class_code}_sid{sid}.xlsx")

    except Exception as e:
        return f"xlsx generation failed: {e}", 500



@app.route("/research/export/teacher_runs.xlsx")
def export_teacher_runs_xlsx():
    guard = require_admin()
    if guard is not None:
        return guard

    if not engine:
        return "DB not configured", 400

    class_code = (request.args.get("class_code") or "").strip().upper()
    session_id = (request.args.get("session_id") or "").strip()
    if not class_code or not session_id:
        return "class_code and session_id are required", 400

    with engine.connect() as conn:
        rows = conn.execute(text("""
            SELECT id, class_code, session_id, teacher_username, condition, submitted,
                   started_at, ended_at, duration_ms, confidence_score, created_at, placements
            FROM teacher_placement_runs
            WHERE class_code = :code AND session_id = :sid
            ORDER BY created_at DESC
        """), {"code": class_code, "sid": session_id}).fetchall()

        run_ids = [int(r.id) for r in rows]
        decisions = []
        if run_ids:
            decisions = conn.execute(text("""
                SELECT id, run_id, target_student_name, priority_rank, decision_confidence, reason_tags, created_at
                FROM teacher_decisions
                WHERE run_id = ANY(:run_ids)
                ORDER BY run_id DESC, priority_rank ASC
            """), {"run_ids": run_ids}).fetchall()

    wb = Workbook()

    ws1 = wb.active
    ws1.title = "teacher_runs"
    ws1.append([
        "run_id", "class_code", "session_id", "teacher_username", "condition", "submitted",
        "started_at", "ended_at", "duration_ms", "confidence_score", "created_at", "placements_json"
    ])

    for r in rows:
        placements_json = ""
        try:
            placements_json = json.dumps(
                r.placements if isinstance(r.placements, dict) else (r.placements or {}),
                ensure_ascii=False
            )
        except Exception:
            placements_json = ""

        ws1.append([
            int(r.id),
            r.class_code,
            r.session_id,
            r.teacher_username,
            r.condition,
            bool(r.submitted),
            r.started_at.isoformat() if r.started_at else None,
            r.ended_at.isoformat() if r.ended_at else None,
            r.duration_ms,
            r.confidence_score,
            r.created_at.isoformat() if r.created_at else None,
            placements_json,
        ])

    _autosize_columns(ws1)

    ws2 = wb.create_sheet("teacher_decisions")
    ws2.append(["decision_id", "run_id", "target_student_name", "priority_rank", "decision_confidence", "reason_tags_json", "created_at"])

    for d in decisions:
        reason_json = ""
        try:
            reason_json = json.dumps(d.reason_tags if isinstance(d.reason_tags, (dict, list)) else (d.reason_tags or []), ensure_ascii=False)
        except Exception:
            reason_json = ""

        ws2.append([
            int(d.id),
            int(d.run_id),
            d.target_student_name,
            d.priority_rank,
            d.decision_confidence,
            reason_json,
            d.created_at.isoformat() if d.created_at else None,
        ])

    _autosize_columns(ws2)
    return _xlsx_response(wb, f"teacher_runs_{class_code}_session{session_id}.xlsx")


# -------------------------
# Teacher dashboard + class management
# -------------------------

@app.route("/teacher/dashboard")
def dashboard():
    if "teacher" not in session:
        return redirect("/teacher/login")

    classes: Dict[str, Any] = {}
    try:
        if engine:
            classes = db_list_classes_for_teacher(session["teacher"])
        else:
            d = load_data()
            classes = {c: v for c, v in d.get("classes", {}).items() if v.get("teacher") == session["teacher"]}
    except Exception:
        d = load_data()
        classes = {c: v for c, v in d.get("classes", {}).items() if v.get("teacher") == session["teacher"]}

    if classes and not session.get("selected_class"):
        session["selected_class"] = next(iter(classes.keys()))

    return render_template("dashboard.html", classes=classes)


@app.route("/teacher/create", methods=["GET", "POST"])
def create_class():
    if "teacher" not in session:
        return redirect("/teacher/login")

    if request.method == "POST":
        code = make_code()
        class_name = request.form.get("class_name", "").strip()
        students_raw = request.form.get("students", "")

        parsed: List[Dict[str, str]] = []
        auto_no = 1
        line_no = 0

        for line in students_raw.splitlines():
            line_no += 1
            line = line.strip()
            if not line:
                continue

            parts = [p.strip() for p in line.split("\t")]
            if len(parts) == 1:
                parts = [p.strip() for p in line.split(",")]

            # 허용 형식:
            # - (이름만) / (번호, 이름) / (번호, 이름, 성별)
            name = ""
            no = ""
            gender_raw = ""

            if len(parts) == 1:
                name = parts[0]
                no = str(auto_no)
            elif len(parts) == 2:
                no, name = parts[0], parts[1]
            else:
                no, name, gender_raw = parts[0], parts[1], (parts[2] or "").strip()

            if not name:
                continue

            gender = normalize_gender(gender_raw)

            if gender not in ["남", "여"]:
                flash(f"{line_no}번째 줄: 성별이 비어있거나 인식되지 않습니다. (남/여) 입력 필요", "error")
                return redirect(request.path)

            parsed.append({"no": str(no or auto_no), "name": name, "gender": gender})
            auto_no += 1




        if engine:
            db_create_class(
                teacher_username=session["teacher"],
                class_code=code,
                class_name=class_name or f"학급 {code}",
                students=parsed,
            )
        else:
            d = load_data()
            d.setdefault("classes", {})
            d["classes"][code] = ensure_class_schema({
                "name": class_name or f"학급 {code}",
                "teacher": session["teacher"],
                "students": parsed,
                "students_data": {s["name"]: {"sessions": {}} for s in parsed},
                "sessions": {}
            })
            save_data_safely(d)

        session["selected_class"] = code
        return redirect("/teacher/dashboard")

    return render_template("create_class.html")


@app.route("/teacher/class/delete", methods=["POST"])
def teacher_delete_class():
    if "teacher" not in session:
        return redirect("/teacher/login")

    code = (request.form.get("code") or "").upper().strip()
    if not code:
        return redirect("/teacher/dashboard")

    try:
        if engine:
            db_delete_class_for_teacher(code, session["teacher"])
        else:
            d = load_data()
            cls = d.get("classes", {}).get(code)
            if cls and cls.get("teacher") == session["teacher"]:
                d["classes"].pop(code, None)
                save_data_safely(d)
    except Exception:
        pass

    # If deleted selected class, clear
    if session.get("selected_class") == code:
        session.pop("selected_class", None)
    return redirect("/teacher/dashboard")

@app.route("/teacher/class/<code>/students", methods=["GET", "POST"])
def teacher_manage_students(code: str):
    if "teacher" not in session:
        return redirect("/teacher/login")

    code = (code or "").upper().strip()
    cls = db_get_class_for_viewer(code, session["teacher"])
    if not cls or cls.get("_forbidden"):
        return "학급을 찾을 수 없거나 접근 권한이 없습니다.", 404

    if request.method == "POST":
        action = (request.form.get("action") or "").strip()

        # 1) 전입(추가)
        if action == "add":
            student_no = (request.form.get("student_no") or "").strip()
            name = (request.form.get("name") or "").strip()
            gender = normalize_gender(request.form.get("gender") or "")

            if name:
                import random
                # 학급 내 pin 중복 없이 생성
                with engine.begin() as conn:
                    existing = conn.execute(text("""
                        SELECT pin_code FROM students WHERE class_code=:c
                    """), {"c": code}).fetchall()
                    used = {r.pin_code for r in existing if r.pin_code}

                    pin = None
                    for _ in range(2000):
                        cand = str(random.randint(100000, 999999))
                        if cand not in used:
                            pin = cand
                            break
                    if not pin:
                        pin = "100000"
                        while pin in used:
                            pin = str(int(pin) + 1)

                    conn.execute(text("""
                        INSERT INTO students (class_code, student_no, name, gender, pin_code, active, joined_at)
                        VALUES (:c, :no, :n, :g, :p, TRUE, NOW())
                    """), {"c": code, "no": student_no, "n": name, "g": gender, "p": pin})

        # 2) 전출/복귀(활성 토글)
        if action == "toggle_active":
            name = (request.form.get("name") or "").strip()
            make_inactive = (request.form.get("set") or "") == "0"
            if name:
                with engine.begin() as conn:
                    if make_inactive:
                        conn.execute(text("""
                            UPDATE students
                            SET active = FALSE, left_at = NOW()
                            WHERE class_code = :c AND name = :n
                        """), {"c": code, "n": name})
                    else:
                        conn.execute(text("""
                            UPDATE students
                            SET active = TRUE, left_at = NULL
                            WHERE class_code = :c AND name = :n
                        """), {"c": code, "n": name})

        # 3) 성별 수정
        if action == "set_gender":
            name = (request.form.get("name") or "").strip()
            gender = normalize_gender(request.form.get("gender") or "")

            if name:
                with engine.begin() as conn:
                    conn.execute(text("""
                        UPDATE students
                        SET gender = :g
                        WHERE class_code = :c AND name = :n
                    """), {"c": code, "n": name, "g": gender})

        return redirect(f"/teacher/class/{code}/students")

    # GET: 학생 목록 표시
    with engine.connect() as conn:
        rows = conn.execute(text("""
            SELECT student_no, name, gender, pin_code, active
            FROM students
            WHERE class_code = :c
            ORDER BY id ASC
        """), {"c": code}).fetchall()

    students = []
    for r in rows:
        students.append({
            "no": r.student_no or "",
            "name": r.name,
            "gender": r.gender or "",
            "pin": r.pin_code or "",
            "active": bool(r.active),
        })

    return render_template("teacher_manage_students.html", cls=cls, code=code, students=students)



@app.route("/teacher/class/<code>")
def class_detail(code):
    # v2 is now the primary class detail page
    if "teacher" not in session:
        return redirect("/teacher/login")

    code = (code or "").upper().strip()
    sid = normalize_visible_session_id(request.args.get("sid") or session.get("selected_session"))
    session["selected_session"] = sid
    session["selected_class"] = code
    return redirect(f"/teacher/class/{code}/v2?sid={sid}&open=1")


@app.route("/teacher/class/<code>/legacy")
def class_detail_legacy(code):
    if "teacher" not in session:
        return redirect("/teacher/login")

    code = (code or "").upper().strip()

    sid = normalize_visible_session_id(request.args.get("sid") or session.get("selected_session"))
    session["selected_session"] = sid

    if engine:
        cls = db_get_class_for_teacher(code, session["teacher"])
        if not cls or cls.get("_forbidden"):
            return "학급을 찾을 수 없거나 접근 권한이 없습니다.", 404

        cls = ensure_class_schema(cls)
        students = db_get_students_in_class(code)
        cls["students"] = students

        submitted_map = db_get_submitted_map(code, sid)
        session["selected_class"] = code

        rows: List[Dict[str, Any]] = []
        for i, item in enumerate(students, start=1):
            no = str(item.get("no", "") or i)
            name = (item.get("name") or "").strip()
            if not name:
                continue
            submitted = bool(submitted_map.get(name, False))
            status = "완료" if submitted else "미완료"
            rows.append({"no": no, "name": name, "status": status, "url_name": quote(name)})

        session_links: List[Dict[str, str]] = []
        for _sid, meta in sorted(cls.get("sessions", {}).items(), key=lambda x: int(x[0])):
            session_links.append({"sid": _sid, "label": meta.get("label", f"{_sid}차"), "url": f"/s/{code}/{_sid}"})

        return render_template(
            "class_detail_v2.html",
            cls=cls,
            code=code,
            rows=rows,
            sid=sid,
            session_links=session_links,
            open_panel=True,
            teacher_run=None,
        )

    # JSON fallback
    d = load_data()
    cls = ensure_class_schema(d.get("classes", {}).get(code))
    if not cls or cls.get("teacher") != session["teacher"]:
        return "학급을 찾을 수 없거나 접근 권한이 없습니다.", 404

    session["selected_class"] = code

    rows = []
    for i, item in enumerate(cls.get("students", []), start=1):
        if isinstance(item, dict):
            no = str(item.get("no", "") or i)
            name = (item.get("name") or "").strip()
        else:
            no = str(i)
            name = (item or "").strip()

        if not name:
            continue

        submitted = bool(cls.get("students_data", {}).get(name, {}).get("sessions", {}).get(sid, {}).get("submitted", False))
        status = "완료" if submitted else "미완료"
        rows.append({"no": no, "name": name, "status": status, "url_name": quote(name)})

    session_links = []
    for _sid, meta in sorted(cls.get("sessions", {}).items(), key=lambda x: int(x[0])):
        session_links.append({"sid": _sid, "label": meta.get("label", f"{_sid}차"), "url": f"/s/{code}/{_sid}"})

    return render_template(
        "class_detail_v2.html",
        cls=cls,
        code=code,
        rows=rows,
        sid=sid,
        session_links=session_links,
        open_panel=True,
        teacher_run=None,
    )


@app.route("/teacher/class/<code>/student_pin_pdf")
def teacher_download_student_pin_pdf(code):
    if "teacher" not in session:
        return redirect(url_for("teacher_login"))

    sid = request.args.get("sid") or DEFAULT_VISIBLE_SESSION_ID

    # 권한 확인 + 학급명 확보
    cls = db_get_class_for_viewer(code, session["teacher"])
    if not cls:
        flash("해당 학급에 접근 권한이 없습니다.", "error")
        return redirect("/teacher/dashboard")

    class_name = cls.get("name") or code

    # 학생 + pin_code
    students = db_get_students_with_pin(code)

    pdf_io = build_student_pin_pdf(class_name=class_name, sid=sid, students=students)

    filename = "학생_로그인_PIN.pdf"
    return send_file(
        pdf_io,
        mimetype="application/pdf",
        as_attachment=True,
        download_name=filename
    )

@app.route("/teacher/class/<code>/v2")
def class_detail_v2(code):
    """Parallel rebuild: new class detail UI (v2).
    Keeps existing v1 intact for safe rollout.

    Notes:
    - v2 currently supports research plan default of 4 rounds (sid 1-4).
    - v2 uses `open=1` query param to expand the selected round panel.
    """
    if "teacher" not in session:
        return redirect("/teacher/login")

    code = (code or "").upper().strip()
    sid = normalize_visible_session_id(request.args.get("sid") or session.get("selected_session"))
    session["selected_session"] = sid

    open_panel = (request.args.get("open") or "").strip() == "1"

    if not engine:
        return "DB가 연결되어 있지 않습니다.", 500

    cls = db_get_class_for_teacher(code, session["teacher"])
    if not cls or cls.get("_forbidden"):
        return "학급을 찾을 수 없거나 접근 권한이 없습니다.", 404

    cls = ensure_class_schema(cls)
    class_teacher_username = cls.get("teacher") or session["teacher"]

    students = db_get_students_in_class(code)
    cls["students"] = students

    submitted_map = db_get_submitted_map(code, sid)
    session["selected_class"] = code

    # --- teacher placement run (for "완료/수정" UI) ---
    teacher_run = None
    try:
        with engine.begin() as conn:
            r = conn.execute(text("""
                SELECT id, submitted
                FROM teacher_placement_runs
                WHERE class_code = :code
                  AND session_id = :sid
                  AND teacher_username = :t
                ORDER BY id DESC
                LIMIT 1
            """), {"code": code, "sid": sid, "t": class_teacher_username}).fetchone()

        if r:
            teacher_run = {"id": int(r.id), "submitted": bool(r.submitted)}
    except Exception:
        # 테이블이 없거나(마이그레이션 누락 등) 조회 실패 시: UI는 기본 "시작"으로 표시
        teacher_run = None

    rows: List[Dict[str, Any]] = []
    for i, item in enumerate(students, start=1):
        no = str(item.get("no", "") or i)
        name = (item.get("name") or "").strip()
        if not name:
            continue
        submitted = bool(submitted_map.get(name, False))
        status = "완료" if submitted else "미완료"
        rows.append({"no": no, "name": name, "status": status, "url_name": quote(name)})

    session_links: List[Dict[str, str]] = []
    for _sid, meta in sorted(cls.get("sessions", {}).items(), key=lambda x: int(x[0])):
        session_links.append({
            "sid": _sid,
            "label": meta.get("label", f"{_sid}차"),
            "url": f"/s/{code}/{_sid}"
        })

    return render_template(
        "class_detail_v2.html",
        cls=cls,
        code=code,
        rows=rows,
        sid=sid,
        session_links=session_links,
        open_panel=open_panel,
        teacher_run=teacher_run,
    )



@app.route("/teacher/class/<code>/sync_from_sheet")
def teacher_sync_from_sheet(code):
    if "teacher" not in session:
        return redirect("/teacher/login")

    code = code.upper().strip()
    sid = normalize_visible_session_id(request.args.get("sid") or session.get("selected_session"))

    cls = db_get_class_for_teacher(code, session["teacher"])
    if not cls or cls.get("_forbidden"):
        return "학급을 찾을 수 없거나 접근 권한이 없습니다.", 404

    try:
        sync_results_from_sheet_to_db(
            class_code=code,
            sid=sid,
            teacher_username=session["teacher"]
        )
    except Exception as e:
        # Render 로그에도 남기고, 브라우저에서도 원인을 바로 보이게 반환
        app.logger.exception("sync_from_sheet failed: code=%s sid=%s teacher=%s", code, sid, session.get("teacher"))
        return jsonify({
            "status": "error",
            "where": "teacher_sync_from_sheet",
            "class_code": code,
            "sid": sid,
            "message": str(e)
        }), 500

    return redirect(f"/teacher/class/{code}/v2?sid={sid}&open=1")


      
@app.route("/teacher/class/<code>/analysis_compare")
def teacher_analysis_compare(code):
    if "teacher" not in session:
        return redirect("/teacher/login")

    code = (code or "").upper().strip()

    cls = db_get_class_for_teacher(code, session["teacher"])
    if not cls or cls.get("_forbidden"):
        return "학급을 찾을 수 없거나 접근 권한이 없습니다.", 404

    cls = ensure_class_schema(cls)

    # 3월 말(1회차)은 이번 운영에서 수집하지 않으므로 비교 화면에서는 숨긴다.
    sids = list(VISIBLE_SESSION_IDS)

    return render_template(
        "teacher_analysis_compare.html",
        cls=cls,
        code=code,
        sids=sids,
    )


@app.route("/teacher/class/<code>/analysis")
def teacher_analysis(code):
    if "teacher" not in session:
        return redirect("/teacher/login")

    code = (code or "").upper().strip()
    sid = (request.args.get("sid") or DEFAULT_VISIBLE_SESSION_ID).strip()
    if sid not in ["1", "2", "3", "4", "5"]:
        sid = DEFAULT_VISIBLE_SESSION_ID

    cls = db_get_class_for_teacher(code, session["teacher"])
    if not cls or cls.get("_forbidden"):
        return "학급을 찾을 수 없거나 접근 권한이 없습니다.", 404

    return redirect(f"/teacher/class/{code}/result/{sid}")
  
    # JSON fallback
    d = load_data()
    cls = ensure_class_schema(d.get("classes", {}).get(code))
    if not cls or cls.get("teacher") != session["teacher"]:
        return "학급을 찾을 수 없거나 접근 권한이 없습니다.", 404

    session["selected_class"] = code

    rows = []
    for i, item in enumerate(cls.get("students", []), start=1):
        if isinstance(item, dict):
            no = str(item.get("no", "") or i)
            name = (item.get("name") or "").strip()
        else:
            no = str(i)
            name = (item or "").strip()

        if not name:
            continue

        submitted = bool(cls.get("students_data", {}).get(name, {}).get("sessions", {}).get(sid, {}).get("submitted", False))
        status = "완료" if submitted else "미완료"
        rows.append({"no": no, "name": name, "status": status, "url_name": quote(name)})

    session_links = []
    for _sid, meta in sorted(cls.get("sessions", {}).items(), key=lambda x: int(x[0])):
        session_links.append({"sid": _sid, "label": meta.get("label", f"{_sid}차"), "url": f"/s/{code}/{_sid}"})

    return render_template(
        "class_detail_v2.html",
        cls=cls,
        code=code,
        rows=rows,
        sid=sid,
        session_links=session_links,
        open_panel=open_panel,
    )




# -------------------------
# v2 API: finalize flow
# -------------------------

def _require_teacher() -> Optional[str]:
    t = session.get("teacher")
    if not t:
        return None
    return str(t)


def _require_class_access(code: str, teacher_username: str) -> Optional[Dict[str, Any]]:
    code = (code or "").upper().strip()
    cls = db_get_class_for_viewer(code, teacher_username)
    if not cls or cls.get("_forbidden"):
        return None
    return ensure_class_schema(cls)


def _class_owner_username(cls: Dict[str, Any], fallback_username: str) -> str:
    """Return the class owner's username for state lookups."""
    return str(cls.get("teacher") or fallback_username)


def _deny_admin_write_if_needed(cls: Dict[str, Any]) -> Optional[Any]:
    """Keep research-admin cross-class access read-only on teacher workflow APIs."""
    if cls.get("admin_view"):
        return jsonify({"ok": False, "error": "ADMIN_VIEW_READ_ONLY"}), 403
    return None


@app.route("/api/v2/class/<code>/session/<sid>/state")
def api_v2_state(code: str, sid: str):
    t = _require_teacher()
    if not t:
        return jsonify({"ok": False, "error": "UNAUTHORIZED"}), 401

    code = (code or "").upper().strip()
    sid = (sid or "").strip()
    if sid not in ["1", "2", "3", "4"]:
        sid = "1"

    cls = _require_class_access(code, t)
    if not cls:
        return jsonify({"ok": False, "error": "FORBIDDEN"}), 403

    students = db_get_students_in_class(code)
    submitted_map = db_get_submitted_map(code, sid)
    incomplete = []
    for item in students:
        name = (item.get("name") or "").strip()
        if not name:
            continue
        if not bool(submitted_map.get(name, False)):
            incomplete.append(name)

    state = db_get_session_state_v2(code, sid, _class_owner_username(cls, t))
    state["ok"] = True
    state["incomplete_students"] = incomplete
    return jsonify(state)


@app.route("/api/v2/class/<code>/session/<sid>/preview_seen", methods=["POST"])
def api_v2_preview_seen(code: str, sid: str):
    t = _require_teacher()
    if not t:
        return jsonify({"ok": False, "error": "UNAUTHORIZED"}), 401
    code = (code or "").upper().strip()
    sid = (sid or "").strip()
    if sid not in ["1", "2", "3", "4"]:
        sid = "1"

    cls = _require_class_access(code, t)
    if not cls:
        return jsonify({"ok": False, "error": "FORBIDDEN"}), 403
    denied = _deny_admin_write_if_needed(cls)
    if denied is not None:
        return denied

    db_set_preview_seen_v2(code, sid, t)
    return jsonify({"ok": True, "preview_seen": True})


@app.route("/api/v2/class/<code>/session/<sid>/exclusions", methods=["POST"])
def api_v2_exclusions(code: str, sid: str):
    t = _require_teacher()
    if not t:
        return jsonify({"ok": False, "error": "UNAUTHORIZED"}), 401
    code = (code or "").upper().strip()
    sid = (sid or "").strip()
    if sid not in ["1", "2", "3", "4"]:
        sid = "1"

    cls = _require_class_access(code, t)
    if not cls:
        return jsonify({"ok": False, "error": "FORBIDDEN"}), 403
    denied = _deny_admin_write_if_needed(cls)
    if denied is not None:
        return denied

    payload = request.get_json(silent=True) or {}
    mode = (payload.get("mode") or "").strip()
    items = payload.get("items") or []
    if mode not in ["no_exclude", "exclude"]:
        return jsonify({"ok": False, "error": "BAD_REQUEST"}), 400
    if mode == "exclude":
        # validate reasons present
        norm_items = []
        for it in items:
            if not isinstance(it, dict):
                continue
            nm = (it.get("student_name") or "").strip()
            rs = (it.get("reason") or "").strip()
            if not nm:
                continue
            if not rs:
                return jsonify({"ok": False, "error": "REASON_REQUIRED"}), 400
            norm_items.append({"student_name": nm, "reason": rs})
        items = norm_items

    db_set_exclusions_v2(code, sid, t, mode, items)
    cache_clear_session_analysis(code, sid)
    return jsonify({"ok": True, "exclusions_resolved": True})


@app.route("/api/v2/class/<code>/session/<sid>/exclusions/remove", methods=["POST"])
def api_v2_remove_exclusion(code: str, sid: str):
    t = _require_teacher()
    if not t:
        return jsonify({"ok": False, "error": "UNAUTHORIZED"}), 401
    code = (code or "").upper().strip()
    sid = (sid or "").strip()
    if sid not in ["1", "2", "3", "4"]:
        sid = "1"

    cls = _require_class_access(code, t)
    if not cls:
        return jsonify({"ok": False, "error": "FORBIDDEN"}), 403
    denied = _deny_admin_write_if_needed(cls)
    if denied is not None:
        return denied

    payload = request.get_json(silent=True) or {}
    student_name = (payload.get("student_name") or "").strip()
    if not student_name:
        return jsonify({"ok": False, "error": "BAD_REQUEST"}), 400

    db_remove_session_exclusion_v2(code, sid, t, student_name)
    cache_clear_session_analysis(code, sid)
    return jsonify({"ok": True, "removed": student_name})


@app.route("/api/v2/class/<code>/session/<sid>/survey", methods=["POST"])
def api_v2_survey(code: str, sid: str):
    t = _require_teacher()
    if not t:
        return jsonify({"ok": False, "error": "UNAUTHORIZED"}), 401
    code = (code or "").upper().strip()
    sid = (sid or "").strip()
    if sid not in ["1", "2", "3", "4"]:
        sid = "1"

    cls = _require_class_access(code, t)
    if not cls:
        return jsonify({"ok": False, "error": "FORBIDDEN"}), 403
    denied = _deny_admin_write_if_needed(cls)
    if denied is not None:
        return denied

    payload = request.get_json(silent=True) or {}
    db_upsert_survey_v2(code, sid, t, payload)
    sheet_resp = sheet_upsert_teacher_survey(code, sid, t, "post", payload)
    return jsonify({
        "ok": True,
        "survey_submitted": True,
        "sheet_status": sheet_resp.get("status"),
        "sheet_message": sheet_resp.get("message"),
    })


@app.route("/api/v2/class/<code>/session/<sid>/survey", methods=["GET"])
def api_v2_get_survey(code: str, sid: str):
    t = _require_teacher()
    if not t:
        return jsonify({"ok": False, "error": "UNAUTHORIZED"}), 401
    code = (code or "").upper().strip()
    sid = (sid or "").strip()
    if sid not in ["1", "2", "3", "4"]:
        sid = "1"

    cls = _require_class_access(code, t)
    if not cls:
        return jsonify({"ok": False, "error": "FORBIDDEN"}), 403

    survey = db_get_survey_v2(code, sid)
    return jsonify({
        "ok": True,
        "survey_submitted": bool(survey),
        "survey": survey,
    })


@app.route("/api/v2/class/<code>/session/<sid>/finalize", methods=["POST"])
def api_v2_finalize(code: str, sid: str):
    t = _require_teacher()
    if not t:
        return jsonify({"ok": False, "error": "UNAUTHORIZED"}), 401
    code = (code or "").upper().strip()
    sid = (sid or "").strip()
    if sid not in ["1", "2", "3", "4"]:
        sid = "1"

    cls = _require_class_access(code, t)
    if not cls:
        return jsonify({"ok": False, "error": "FORBIDDEN"}), 403
    denied = _deny_admin_write_if_needed(cls)
    if denied is not None:
        return denied

    res = db_finalize_session_v2(code, sid, t)
    if not res.get("ok"):
        return jsonify(res), 400

    return jsonify({
        "ok": True,
        "finalized": True,
        "result_url": f"/teacher/class/{code}/result/{sid}"
    })


# -------------------------
# v2: session-level result page (minimal)
# -------------------------
@app.route("/teacher/class/<code>/result/<sid>")
def teacher_session_result(code: str, sid: str):
    if "teacher" not in session:
        return redirect("/teacher/login")

    code = (code or "").upper().strip()
    sid = (sid or "").strip()
    if sid not in ["1", "2", "3", "4"]:
        sid = "1"

    cls = db_get_class_for_viewer(code, session["teacher"])
    if not cls or cls.get("_forbidden"):
        return "학급을 찾을 수 없거나 접근 권한이 없습니다.", 404
    cls = ensure_class_schema(cls)
    class_teacher_username = _class_owner_username(cls, session["teacher"])

    students = db_get_students_in_class(code)
    submitted_map = db_get_submitted_map(code, sid)
    done = 0
    total = 0
    for it in students:
        nm = (it.get("name") or "").strip()
        if not nm:
            continue
        total += 1
        if bool(submitted_map.get(nm, False)):
            done += 1

    # teacher placement existence
    teacher_has = False
    if engine:
        with engine.connect() as conn:
            r = conn.execute(text("""
                SELECT 1 FROM teacher_placement_runs
                WHERE class_code=:code AND teacher_username=:t AND session_id=:sid
                LIMIT 1
            """), {"code": code, "t": class_teacher_username, "sid": sid}).fetchone()
            teacher_has = bool(r)

    state = db_get_session_state_v2(code, sid, class_teacher_username)
    spm_payload = None
    spm_error = None
    try:
        spm_payload = build_spm_result_payload(code, sid)
        cache_set(code, sid, f"spm_result_{sid}_v1", spm_payload)
    except Exception as e:
        app.logger.exception("SPM result payload failed for class=%s sid=%s", code, sid)
        spm_error = str(e)

    return render_template(
        "session_result.html",
        cls=cls,
        code=code,
        sid=sid,
        total=total,
        done=done,
        teacher_has=teacher_has,
        state=state,
        spm_payload=spm_payload,
        spm_error=spm_error,
    )

@app.route("/teacher/class/<code>/result/<sid>/<url_name>")
def teacher_view_student_result(code, sid, url_name):
    if "teacher" not in session:
        return redirect("/teacher/login")

    code = (code or "").upper().strip()
    sid = (sid or "1").strip()
    if sid not in ["1", "2", "3", "4", "5"]:
        sid = "1"

    student_name = (unquote(url_name) or "").strip()

    cls = db_get_class_for_teacher(code, session["teacher"])
    if not cls or cls.get("_forbidden"):
        return "학급을 찾을 수 없거나 접근 권한이 없습니다.", 404

    students = db_get_students_in_class(code)
    all_names = [s["name"] for s in students]
    if student_name not in all_names:
        return "학생을 찾을 수 없습니다.", 404

    friends = [n for n in all_names if n != student_name]
    sess = db_get_student_session(code, student_name, sid)
    placements = (sess.get("placements") if sess else {}) or {}

    student_session = {"placements": placements, "submitted": True}

    cls_for_view = ensure_class_schema({
        "code": code,
        "name": db_get_class_name(code) or code,
        "teacher": session["teacher"],
        "sessions": {}
    }) or {}
    session_meta = (cls_for_view.get("sessions") or {}).get(sid, {"label": f"{sid}차"})

    return render_template(
        "student_write.html",
        name=student_name,
        friends=friends,
        placements=placements,
        student_session=student_session,
        sid=sid,
        session_meta=session_meta,
        teacher_view=True,
    )


# -------------------------
# Student entry
# -------------------------

from urllib.parse import quote  # app.py 상단에 없으면 추가하세요.

@app.route("/s/<code>/<sid>", methods=["GET", "POST"])
def student_enter_session(code, sid):
    code = (code or "").upper().strip()
    sid = (sid or "1").strip()

    # 1) 학급 조회 + 스키마 구성
    if engine:
        with engine.connect() as conn:
            row = conn.execute(text("""
                SELECT code, name, teacher_username
                FROM classes
                WHERE code = :code
                LIMIT 1
            """), {"code": code}).fetchone()

        if not row:
            return "학급을 찾을 수 없습니다.", 404

        cls = ensure_class_schema({
            "code": code,
            "name": row.name,
            "teacher": row.teacher_username,
            "sessions": {}
        })

        students = db_get_students_in_class(code)  # 이 함수가 active 컬럼을 쓰는지 확인 필요
        cls["students"] = students
        cls["students_data"] = {s["name"]: {"sessions": {}} for s in students}

    else:
        d = load_data()
        cls = ensure_class_schema(d.get("classes", {}).get(code))
        if not cls:
            return "학급을 찾을 수 없습니다.", 404

    # 2) sid 보정
    if sid not in (cls.get("sessions") or {}):
        sid = "1"

    session_label = (cls.get("sessions") or {}).get(sid, {}).get("label", f"{sid}차")

    # 3) POST: 이름 + PIN 검증
    if request.method == "POST":
        name = request.form.get("name", "").strip()
        pin = (request.form.get("pin") or "").strip()

        # (1) 이름 검증: 명단 기반
        if not name or name not in (cls.get("students_data") or {}):
            return render_template(
                "student_enter_session.html",
                error="학생 명단에 없는 이름입니다.",
                code=code,
                sid=sid,
                session_label=session_label,
            )

        # (2) PIN 형식 검증: 6자리 숫자
        if not (len(pin) == 6 and pin.isdigit()):
            return render_template(
                "student_enter_session.html",
                error="개인 코드는 6자리 숫자여야 합니다.",
                code=code,
                sid=sid,
                session_label=session_label,
            )

        # (3) DB에서 PIN 일치 검증 (active 학생만)
        if engine:
            with engine.connect() as conn:
                r = conn.execute(text("""
                    SELECT pin_code
                    FROM students
                    WHERE class_code = :code
                      AND name = :name
                      AND active = TRUE
                    LIMIT 1
                """), {"code": code, "name": name}).fetchone()

            db_pin = (r.pin_code if r else None)
            if not db_pin or str(db_pin).strip() != pin:
                return render_template(
                    "student_enter_session.html",
                    error="개인 코드(PIN)가 올바르지 않습니다.",
                    code=code,
                    sid=sid,
                    session_label=session_label,
                )

        # 4) 세션에 저장 (프로젝트 다른 라우트들과 호환되게)
        session["name"] = name
        session["code"] = code
        session["sid"] = sid
        session["selected_class"] = code
        session["selected_session"] = sid

        # 5) 학생 작성 화면으로 이동
        # 프로젝트에 /student/write 라우트가 있다고 가정하는 "가장 안전한" 형태
        # 만약 기존 라우트가 querystring을 요구하면 ?code=...&sid=... 형태로 바꾸세요.
        return redirect("/student/write")

    # 6) GET: 진입 페이지 렌더
    return render_template(
        "student_enter_session.html",
        error=None,
        code=code,
        sid=sid,
        session_label=session_label,
    )


@app.route("/qr/<code>/<sid>.png")
def qr_session_link(code, sid):
    code = (code or "").upper().strip()
    sid = str(sid).strip()
    if sid not in ["1", "2", "3", "4", "5"]:
        sid = "1"

    if engine:
        with engine.connect() as conn:
            row = conn.execute(text("SELECT 1 FROM classes WHERE code = :code LIMIT 1"), {"code": code}).fetchone()
        if not row:
            return "학급을 찾을 수 없습니다.", 404
    else:
        d = load_data()
        cls = d.get("classes", {}).get(code)
        if not cls:
            return "학급을 찾을 수 없습니다.", 404

    base = request.url_root.rstrip("/")
    target = f"{base}/s/{code}/{sid}"

    try:
        import qrcode
    except ModuleNotFoundError:
        return "QR 코드 생성을 위해 qrcode 라이브러리가 필요합니다.", 500

    img = qrcode.make(target)
    buf = io.BytesIO()
    img.save(buf, format="PNG")
    buf.seek(0)

    resp = send_file(buf, mimetype="image/png")
    resp.headers["Cache-Control"] = "public, max-age=86400"
    return resp


# Legacy student enter (optional)
@app.route("/student", methods=["GET", "POST"])
def student_enter():
    # Keep legacy entry route as a fallback to sid=1
    if request.method == "POST":
        code = request.form.get("code", "").upper().strip()
        name = request.form.get("name", "").strip()

        if engine:
            with engine.connect() as conn:
                c_row = conn.execute(text("SELECT 1 FROM classes WHERE code=:code LIMIT 1"), {"code": code}).fetchone()
            if not c_row:
                return render_template("student_enter.html", error="입장 실패")

            pin = (request.form.get("pin") or "").strip()

            # 6자리 숫자만 허용(아이들 입력 실수 방지)
            if not (len(pin) == 6 and pin.isdigit()):
                return render_template("student_enter.html", error="개인 코드는 6자리 숫자여야 합니다.")

            with engine.connect() as conn:
                s_row = conn.execute(text("""
                    SELECT 1
                    FROM students
                    WHERE class_code = :code
                      AND name = :name
                      AND pin_code = :pin
                      AND active = TRUE
                    LIMIT 1
                """), {"code": code, "name": name, "pin": pin}).fetchone()

            if not s_row:
                return render_template("student_enter.html", error="입장 실패")

            session["code"] = code
            session["name"] = name
            session["sid"] = DEFAULT_VISIBLE_SESSION_ID
            session["selected_class"] = code
            session["selected_session"] = DEFAULT_VISIBLE_SESSION_ID
            return redirect("/student/write")

        d = load_data()
        cls = d.get("classes", {}).get(code)
        if not cls or name not in (cls.get("students_data") or {}):
            return render_template("student_enter.html", error="입장 실패")

        session["code"] = code
        session["name"] = name
        session["sid"] = DEFAULT_VISIBLE_SESSION_ID
        session["selected_class"] = code
        session["selected_session"] = DEFAULT_VISIBLE_SESSION_ID
        return redirect("/student/write")

    return render_template("student_enter.html")


# -------------------------
# Student write
# -------------------------

@app.route("/student/write", methods=["GET", "POST"])
def student_write():
    if "code" not in session or "name" not in session:
        return redirect("/student")

    code = (session.get("code") or "").upper().strip()
    name = (session.get("name") or "").strip()

    sid = (session.get("sid") or DEFAULT_VISIBLE_SESSION_ID).strip()
    if sid not in ["1", "2", "3", "4", "5"]:
        sid = DEFAULT_VISIBLE_SESSION_ID
        session["sid"] = sid

    if engine:
        with engine.connect() as conn:
            row = conn.execute(text("""
                SELECT code, name, teacher_username
                FROM classes
                WHERE code = :code
                LIMIT 1
            """), {"code": code}).fetchone()
        if not row:
            return redirect("/student")

        students = db_get_students_in_class(code)
        all_names = [s["name"] for s in students]
        if name not in all_names:
            return redirect("/student")

        friends = [n for n in all_names if n != name]

        cls = ensure_class_schema({"code": code, "name": row.name, "teacher": row.teacher_username, "sessions": {}})
        cls["students"] = students
        cls["students_data"] = {n: {"sessions": {}} for n in all_names}

        db_sess = db_get_student_session(code, name, sid)
        placements = (db_sess.get("placements") if db_sess else {}) or {}
        submitted = bool(db_sess.get("submitted")) if db_sess else False

        if request.method == "POST":
            if submitted:
                return redirect("/student/submitted")

            placements_json = (request.form.get("placements_json") or "{}").strip()
            try:
                placements_obj = json.loads(placements_json) if placements_json else {}
            except Exception:
                placements_obj = {}

            resp = post_to_sheet({
                "action": "result_upsert",
                "teacher": row.teacher_username,
                "class_code": code,
                "student": name,
                "session": str(sid),
                "placements": placements_obj,
                "ip": request.headers.get("X-Forwarded-For", request.remote_addr) or "",
            })


            if resp.get("status") != "ok":
                return render_template(
                    "student_write.html",
                    error=f"저장 실패(구글 시트): {resp}",
                    name=name,
                    friends=friends,
                    placements=placements_obj,
                    student_session={"placements": placements_obj, "submitted": False},
                    sid=sid,
                    session_meta=(cls.get("sessions") or {}).get(sid, {}),
                )

            db_upsert_student_session(code, name, sid, placements_obj, submitted=True)
            return redirect("/student/submitted")

        return render_template(
            "student_write.html",
            name=name,
            friends=friends,
            placements=placements,
            student_session={"placements": placements, "submitted": submitted},
            sid=sid,
            session_meta=(cls.get("sessions") or {}).get(sid, {}),
        )

    # JSON fallback
    d = load_data()
    cls = ensure_class_schema(d.get("classes", {}).get(code))
    if not cls:
        return redirect("/student")
    if name not in (cls.get("students_data") or {}):
        return redirect("/student")
    if sid not in (cls.get("sessions") or {}):
        sid = "1"
        session["sid"] = sid

    student = cls["students_data"][name]
    student.setdefault("sessions", {})
    student["sessions"].setdefault(sid, {"placements": {}, "submitted": False})
    ssession = student["sessions"][sid]

    friends = [s["name"] for s in cls.get("students", []) if isinstance(s, dict) and s.get("name") != name]
    placements = ssession.get("placements") or {}

    if request.method == "POST":
        if ssession.get("submitted"):
            return redirect("/student/submitted")

        placements_json = (request.form.get("placements_json") or "{}").strip()
        try:
            placements_obj = json.loads(placements_json) if placements_json else {}
        except Exception:
            placements_obj = {}

        resp = post_to_sheet({
            "action": "result_append",
            "teacher": cls.get("teacher", ""),
            "class_code": code,
            "student": name,
            "session": sid,
            "placements": placements_obj,
            "ip": request.headers.get("X-Forwarded-For", request.remote_addr) or "",
        })

        if resp.get("status") != "ok":
            return render_template(
                "student_write.html",
                error=f"저장 실패(구글 시트): {resp}",
                name=name,
                friends=friends,
                placements=placements_obj,
                student_session=ssession,
                sid=sid,
                session_meta=(cls.get("sessions") or {}).get(sid, {}),
            )

        ssession["placements"] = placements_obj
        ssession["submitted"] = True
        student["sessions"][sid] = ssession
        d["classes"][code] = ensure_class_schema(cls)
        save_data_safely(d)
        return redirect("/student/submitted")

    return render_template(
        "student_write.html",
        name=name,
        friends=friends,
        placements=placements,
        student_session=ssession,
        sid=sid,
        session_meta=(cls.get("sessions") or {}).get(sid, {}),
    )


@app.route("/student/submitted")
def student_submitted():
    return render_template("student_submitted.html")


# -------------------------
# Teacher placement flow
# -------------------------

@app.route("/teacher/class/<code>/placement/start")
def teacher_placement_start(code):
    if "teacher" not in session:
        return redirect("/teacher/login")

    code = (code or "").upper().strip()
    sid = normalize_visible_session_id(request.args.get("sid") or session.get("selected_session"))
    if sid not in ["1", "2", "3", "4", "5"]:
        sid = DEFAULT_VISIBLE_SESSION_ID

    # ✅ 새로 작성(덮어쓰기 시작)은 ?new=1 로만 한다
    force_new = (request.args.get("new") or "").strip() == "1"

    # ✅ 1) 기본은 "최신 run으로 재진입" (submitted여부와 무관)
    #    - 저장 후 다시 들어왔을 때, 이전 배치가 보이게 하는 핵심
    if not force_new:
        latest_id = db_get_latest_teacher_run_id(code, session["teacher"], sid)
        if latest_id:
            return redirect(f"/teacher/placement/{latest_id}")

    # ✅ 2) 기존 run이 없거나, new=1이면 새 run 생성
    condition = (request.args.get("condition") or "BASELINE").strip()
    if condition not in ["BASELINE", "TOOL_ASSISTED"]:
        condition = "BASELINE"

    tool_run_id = request.args.get("tool_run_id")
    tool_run_id_val = int(tool_run_id) if (tool_run_id and str(tool_run_id).isdigit()) else None

    run_id = db_create_teacher_run(code, session["teacher"], sid, condition, tool_run_id=tool_run_id_val)
    return redirect(f"/teacher/placement/{run_id}")




@app.route("/teacher/placement/<int:run_id>", methods=["GET", "POST"])
def teacher_placement_write(run_id: int):
    if "teacher" not in session:
        return redirect("/teacher/login")

    run = db_get_teacher_run(run_id)
    if not run or run["teacher_username"] != session["teacher"]:
        return "접근 권한이 없습니다.", 403

    code = run["class_code"]
    sid = run["session_id"]

    students = db_get_students_in_class(code)
    all_names = [s["name"] for s in students]
    placements = run.get("placements") or {}

    # ✅ placements가 "모든 학생 포함"이면 완료로 판단
    placements_complete = False
    try:
        placements_complete = all((n in placements) for n in all_names) and len(all_names) > 0
    except Exception:
        placements_complete = False

    # ✅ 기본은 읽기 전용(완료/마무리된 경우). 수정하고 싶으면 ?edit=1
    edit_mode = (request.args.get("edit") or "").strip() == "1"
    readonly = (not edit_mode) and (placements_complete or bool(run.get("submitted")))

    if request.method == "POST":
        placements_json = (request.form.get("placements_json") or "{}").strip()
        try:
            placements_obj = json.loads(placements_json) if placements_json else {}
        except Exception:
            placements_obj = {}

        # 1) DB에 저장(기존 동작 유지)
        db_update_teacher_run_placements(run_id, placements_obj)

        # 2) 구글 시트 Results 탭에도 저장(원자료)
        #    - teacher/학생 저장과 동일하게 result_append 사용
        #    - teacher는 session["teacher"]
        #    - student에는 teacher_username을 넣어서 "교사 1줄"로 구분
        name_to_no = {s.get("name", ""): str(s.get("no", "")).strip() for s in students}

        placements_for_sheet = {}
        for name, v in (placements_obj or {}).items():
            key = name_to_no.get(name) or name  # 가능하면 학생번호, 아니면 이름
            if isinstance(v, (list, tuple)) and len(v) >= 2:
                placements_for_sheet[key] = [v[0], v[1]]
            elif isinstance(v, dict) and ("x" in v) and ("y" in v):
                placements_for_sheet[key] = [v.get("x"), v.get("y")]
            else:
                placements_for_sheet[key] = v

        resp = post_to_sheet({
            "action": "result_append",
            "teacher": session["teacher"],
            "class_code": code,
            "student": session["teacher"],  # 교사 1줄(teacher=student) 규칙
            "session": str(sid),
            "placements": placements_for_sheet,
            "ip": request.headers.get("X-Forwarded-For", request.remote_addr) or "",
        })

        if resp.get("status") != "ok":
            app.logger.error("teacher placement save to sheet failed: %s", resp)
            return jsonify({
                "status": "error",
                "where": "teacher_placement_write",
                "message": f"저장 실패(구글 시트): {resp}",
                "class_code": code,
                "sid": str(sid),
                "run_id": run_id,
            }), 500

        return redirect(f"/teacher/placement/{run_id}/complete")

    return render_template(
        "teacher_write.html",
        run=run,
        code=code,
        sid=sid,
        friends=all_names,
        placements=placements,
        placements_complete=placements_complete,
        readonly=readonly,
        edit_mode=edit_mode,
    )


@app.route("/teacher/placement/<int:run_id>/complete", methods=["GET", "POST"])
def teacher_placement_complete(run_id: int):
    if "teacher" not in session:
        return redirect("/teacher/login")

    run = db_get_teacher_run(run_id)
    if not run or run["teacher_username"] != session["teacher"]:
        return "접근 권한이 없습니다.", 403

    if request.method == "POST":
        duration_ms = 0  # duration은 수집/분석하지 않음
        confidence_raw = (request.form.get("confidence_score") or "").strip()

        # 슬라이더: 0~100 int (범위 밖이면 자동 보정)
        try:
            confidence_score = int(confidence_raw) if confidence_raw != "" else 50
        except Exception:
            confidence_score = 50
        confidence_score = max(0, min(100, confidence_score))

        # 우선순위(hidden input: priority_1~3)
        decisions: List[Dict[str, Any]] = []
        for rank in [1, 2, 3]:
            nm = (request.form.get(f"priority_{rank}") or "").strip()
            if nm:
                decisions.append({"name": nm, "rank": rank})

        db_replace_teacher_decisions(run_id, decisions)
        db_complete_teacher_run(run_id, confidence_score=confidence_score)
        sheet_upsert_teacher_survey(
            run["class_code"],
            run["session_id"],
            session["teacher"],
            "pre",
            {
                "run_id": run_id,
                "confidence_score": confidence_score,
                "priority_students": decisions,
            },
        )

        return redirect(f"/teacher/class/{run['class_code']}?sid={run['session_id']}")

    return redirect(f"/teacher/placement/{run_id}")



# -------------------------
# Analysis helpers
# -------------------------

def _extract_point(v: Any, canvas_w: Optional[float] = None, canvas_h: Optional[float] = None) -> Optional[Tuple[float, float, str]]:
    """
    placements item v -> (x,y,mode_tag)
    - abs: {x,y,w,h,mode:'abs'} normalized by w/h if present
    - rel: {x,y,...} used as-is, later bbox-normalized
    """
    if not isinstance(v, dict):
        return None

    mode = v.get("mode")
    x = v.get("x")
    y = v.get("y")
    if not (isinstance(x, (int, float)) and isinstance(y, (int, float))):
        return None

    if mode == "abs":
        w = v.get("w") if isinstance(v.get("w"), (int, float)) else canvas_w
        h = v.get("h") if isinstance(v.get("h"), (int, float)) else canvas_h
        if not (isinstance(w, (int, float)) and w > 0 and isinstance(h, (int, float)) and h > 0):
            return (float(x), float(y), "abs_raw")
        return (float(x) / float(w), float(y) / float(h), "abs_norm")

    return (float(x), float(y), "rel")


def points_from_placements_all_students(placements: Dict[str, Any], names: List[str]) -> Tuple[List[Tuple[float, float]], List[bool]]:
    pts: List[Tuple[float, float]] = []
    valid: List[bool] = []
    for nm in names:
        p = _extract_point(placements.get(nm))
        if p is None:
            pts.append((0.0, 0.0))
            valid.append(False)
        else:
            pts.append((p[0], p[1]))
            valid.append(True)

    xs = [pts[i][0] for i in range(len(pts)) if valid[i]]
    ys = [pts[i][1] for i in range(len(pts)) if valid[i]]
    if len(xs) >= 2 and len(ys) >= 2:
        minx, maxx = min(xs), max(xs)
        miny, maxy = min(ys), max(ys)
        rx = (maxx - minx) if (maxx - minx) > 1e-9 else 1.0
        ry = (maxy - miny) if (maxy - miny) > 1e-9 else 1.0
        pts = [((x - minx) / rx, (y - miny) / ry) for (x, y) in pts]

    return pts, valid


def points_from_student_session(placements: Dict[str, Any], names: List[str], self_name: str) -> Tuple[List[Tuple[float, float]], List[bool]]:
    """
    Student: self_name is included as (0,0) and valid=True.
    Others: from placements; missing -> invalid.
    Then bbox-normalize based on valid points.
    """
    pts: List[Tuple[float, float]] = []
    valid: List[bool] = []

    for nm in names:
        if nm == self_name:
            pts.append((0.0, 0.0))
            valid.append(True)
            continue

        v = placements.get(nm)
        if not isinstance(v, dict):
            pts.append((0.0, 0.0))
            valid.append(False)
            continue

        x = v.get("x")
        y = v.get("y")
        if not (isinstance(x, (int, float)) and isinstance(y, (int, float))):
            pts.append((0.0, 0.0))
            valid.append(False)
            continue

        pts.append((float(x), float(y)))
        valid.append(True)

    xs = [pts[i][0] for i in range(len(pts)) if valid[i]]
    ys = [pts[i][1] for i in range(len(pts)) if valid[i]]
    if len(xs) >= 2 and len(ys) >= 2:
        minx, maxx = min(xs), max(xs)
        miny, maxy = min(ys), max(ys)
        rx = (maxx - minx) if (maxx - minx) > 1e-9 else 1.0
        ry = (maxy - miny) if (maxy - miny) > 1e-9 else 1.0
        pts = [((x - minx) / rx, (y - miny) / ry) for (x, y) in pts]

    return pts, valid


def distance_matrix(points: List[Tuple[float, float]], valid: List[bool]) -> List[List[Optional[float]]]:
    n = len(points)
    D: List[List[Optional[float]]] = [[None] * n for _ in range(n)]
    for i in range(n):
        if not valid[i]:
            continue
        D[i][i] = 0.0
        xi, yi = points[i]
        for j in range(i + 1, n):
            if not valid[j]:
                continue
            xj, yj = points[j]
            d = math.hypot(xi - xj, yi - yj)
            d = round(float(d), 6)
            D[i][j] = d
            D[j][i] = d
    return D


def mean_distance_matrix(mats: List[List[List[Optional[float]]]]) -> List[List[Optional[float]]]:
    if not mats:
        return []
    n = len(mats[0])
    avg: List[List[Optional[float]]] = [[None] * n for _ in range(n)]
    for i in range(n):
        for j in range(n):
            vals: List[float] = []
            for M in mats:
                v = M[i][j]
                if isinstance(v, (int, float)):
                    vals.append(float(v))
            if vals:
                avg[i][j] = round(sum(vals) / len(vals), 6)
    for i in range(n):
        avg[i][i] = 0.0
    return avg


def classical_mds_2d(D: List[List[Optional[float]]]) -> List[Tuple[float, float]]:
    n = len(D)
    if n == 0:
        return []

    filled: List[List[float]] = [[0.0] * n for _ in range(n)]
    for i in range(n):
        row_vals = [D[i][j] for j in range(n) if isinstance(D[i][j], (int, float)) and i != j]
        row_mean = (sum(row_vals) / len(row_vals)) if row_vals else 0.0
        for j in range(n):
            v = D[i][j]
            filled[i][j] = float(v) if isinstance(v, (int, float)) else float(row_mean)

    D2 = [[filled[i][j] ** 2 for j in range(n)] for i in range(n)]
    row_mean = [sum(D2[i]) / n for i in range(n)]
    col_mean = [sum(D2[i][j] for i in range(n)) / n for j in range(n)]
    total_mean = sum(row_mean) / n

    B = [[-0.5 * (D2[i][j] - row_mean[i] - col_mean[j] + total_mean) for j in range(n)] for i in range(n)]

    def matvec(M: List[List[float]], v: List[float]) -> List[float]:
        return [sum(M[i][k] * v[k] for k in range(n)) for i in range(n)]

    def dot(a: List[float], b: List[float]) -> float:
        return sum(a[i] * b[i] for i in range(n))

    def norm(v: List[float]) -> float:
        return math.sqrt(dot(v, v)) + 1e-12

    def power_iter(M: List[List[float]], iters: int = 80) -> Tuple[float, List[float]]:
        v = [1.0 / math.sqrt(n)] * n
        for _ in range(iters):
            w = matvec(M, v)
            nv = norm(w)
            v = [x / nv for x in w]
        lam = dot(v, matvec(M, v))
        return lam, v

    lam1, v1 = power_iter(B)
    B2 = [[B[i][j] - lam1 * v1[i] * v1[j] for j in range(n)] for i in range(n)]
    lam2, v2 = power_iter(B2)

    lam1 = max(lam1, 0.0)
    lam2 = max(lam2, 0.0)
    s1 = math.sqrt(lam1)
    s2 = math.sqrt(lam2)

    coords = [(round(v1[i] * s1, 6), round(v2[i] * s2, 6)) for i in range(n)]
    return coords


def teacher_run_distance_payload(class_code: str, sid: str, run_id: int) -> Optional[Dict[str, Any]]:
    run = db_get_teacher_run(run_id)
    if not run:
        return None

    students = db_get_students_in_class(class_code)
    names = [s["name"] for s in students]
    placements = run.get("placements") or {}

    pts, valid = points_from_placements_all_students(placements, names)
    D = distance_matrix(pts, valid)
    X = classical_mds_2d(D)

    return {
        "class_code": class_code,
        "session_id": sid,
        "run_id": run_id,
        "names": names,
        "points_norm": [{"x": pts[i][0], "y": pts[i][1], "valid": bool(valid[i])} for i in range(len(names))],
        "distance_matrix": D,
        "mds_2d": [{"x": X[i][0], "y": X[i][1]} for i in range(len(names))],
        "generated_at": datetime.utcnow().isoformat() + "Z",
    }


def student_avg_distance_payload(class_code: str, sid: str) -> Dict[str, Any]:
    students = db_get_students_in_class(class_code)
    names = [s["name"] for s in students]
    n_total = len(names)

    if n_total == 0:
        return {
            "class_code": class_code,
            "session_id": sid,
            "names": [],
            "n_total": 0,
            "n_submitted": 0,
            "submitted_students": [],
            "avg_distance_matrix": [],
            "mds_2d": [],
            "generated_at": datetime.utcnow().isoformat() + "Z",
        }

    submitted = db_list_submitted_student_sessions(class_code, sid)

    mats: List[List[List[Optional[float]]]] = []
    used_students: List[str] = []

    for item in submitted:
        self_name = item.get("student_name")
        placements = item.get("placements") or {}

        if self_name not in names:
            continue

        pts, valid = points_from_student_session(placements, names, self_name=self_name)
        D = distance_matrix(pts, valid)
        mats.append(D)
        used_students.append(self_name)

    avgD = mean_distance_matrix(mats)
    X = classical_mds_2d(avgD) if avgD else []

    return {
        "class_code": class_code,
        "session_id": sid,
        "names": names,
        "n_total": n_total,
        "n_submitted": len(used_students),
        "submitted_students": used_students,
        "avg_distance_matrix": avgD,
        "mds_2d": [{"x": X[i][0], "y": X[i][1]} for i in range(len(X))],
        "generated_at": datetime.utcnow().isoformat() + "Z",
    }


def student_vs_avg_distance_payload(class_code: str, sid: str, student_name: str) -> Optional[Dict[str, Any]]:
    students = db_get_students_in_class(class_code)
    names = [s["name"] for s in students]
    if not names or student_name not in names:
        return None

    avg_cache_key = f"student_avg_{sid}"
    avg_payload = cache_get(class_code, sid, avg_cache_key)
    if not avg_payload:
        avg_payload = student_avg_distance_payload(class_code, sid)

    avgD = avg_payload.get("avg_distance_matrix") or []
    if not avgD:
        return {
            "class_code": class_code,
            "session_id": sid,
            "student_name": student_name,
            "error": "avg_distance_matrix empty",
            "generated_at": datetime.utcnow().isoformat() + "Z",
        }

    ss = db_get_student_session(class_code, student_name, sid)
    if not ss:
        return None
    if not ss.get("submitted"):
        return {
            "class_code": class_code,
            "session_id": sid,
            "student_name": student_name,
            "error": "student session not submitted",
            "generated_at": datetime.utcnow().isoformat() + "Z",
        }

    placements = ss.get("placements") or {}
    pts, valid = points_from_student_session(placements, names, self_name=student_name)
    Ds = distance_matrix(pts, valid)

    n = len(names)
    diffs: List[float] = []
    abs_diffs: List[float] = []

    used_pairs = 0
    total_pairs = 0

    for i in range(n):
        for j in range(i + 1, n):
            total_pairs += 1
            v_s = Ds[i][j]
            v_a = avgD[i][j] if (i < len(avgD) and j < len(avgD[i])) else None
            if isinstance(v_s, (int, float)) and isinstance(v_a, (int, float)):
                d = float(v_s) - float(v_a)
                diffs.append(d)
                abs_diffs.append(abs(d))
                used_pairs += 1

    if used_pairs == 0:
        return {
            "class_code": class_code,
            "session_id": sid,
            "student_name": student_name,
            "n_total_students": n,
            "n_pairs_total": total_pairs,
            "n_pairs_used": 0,
            "error": "no comparable pairs",
            "generated_at": datetime.utcnow().isoformat() + "Z",
        }

    mean_abs = sum(abs_diffs) / len(abs_diffs)
    mean_signed = sum(diffs) / len(diffs)
    var_abs = sum((x - mean_abs) ** 2 for x in abs_diffs) / len(abs_diffs)

    self_idx = names.index(student_name)
    self_peer_diffs: List[Dict[str, Any]] = []
    for j in range(n):
        if j == self_idx:
            continue
        v_s = Ds[self_idx][j]
        v_a = avgD[self_idx][j] if (self_idx < len(avgD) and j < len(avgD[self_idx])) else None
        if isinstance(v_s, (int, float)) and isinstance(v_a, (int, float)):
            self_peer_diffs.append({
                "peer": names[j],
                "student_dist": round(float(v_s), 6),
                "avg_dist": round(float(v_a), 6),
                "diff": round(float(v_s) - float(v_a), 6),
                "abs_diff": round(abs(float(v_s) - float(v_a)), 6),
            })
    self_peer_diffs.sort(key=lambda x: x["abs_diff"], reverse=True)

    return {
        "class_code": class_code,
        "session_id": sid,
        "student_name": student_name,
        "n_total_students": n,
        "n_pairs_total": total_pairs,
        "n_pairs_used": used_pairs,
        "mean_abs_diff": round(mean_abs, 6),
        "var_abs_diff": round(var_abs, 6),
        "mean_signed_diff": round(mean_signed, 6),
        "self_peer_diffs_top": self_peer_diffs[:10],
        "generated_at": datetime.utcnow().isoformat() + "Z",
    }


def kmeans_2d(points: List[Tuple[float, float]], k: int, n_init: int = 10, max_iter: int = 60, seed: int = 42) -> Tuple[List[int], List[Tuple[float, float]], float]:
    if not points:
        return [], [], 0.0

    n = len(points)
    k = max(1, min(int(k), n))

    def dist2(a: Tuple[float, float], b: Tuple[float, float]) -> float:
        dx = a[0] - b[0]
        dy = a[1] - b[1]
        return dx * dx + dy * dy

    best_labels: List[int] = []
    best_centers: List[Tuple[float, float]] = []
    best_inertia: Optional[float] = None

    for t in range(max(1, int(n_init))):
        rnd = random.Random(int(seed) + t)

        idx = list(range(n))
        rnd.shuffle(idx)
        idx = idx[:k]
        centers = [points[i] for i in idx]
        labels = [0] * n

        for _ in range(int(max_iter)):
            changed = False

            for i in range(n):
                p = points[i]
                best_c = 0
                best_d = dist2(p, centers[0])
                for ci in range(1, k):
                    d = dist2(p, centers[ci])
                    if d < best_d:
                        best_d = d
                        best_c = ci
                if labels[i] != best_c:
                    labels[i] = best_c
                    changed = True

            new_centers: List[Tuple[float, float]] = []
            for ci in range(k):
                members = [points[i] for i in range(n) if labels[i] == ci]
                if not members:
                    new_centers.append(points[rnd.randrange(0, n)])
                    continue
                mx = sum(p[0] for p in members) / len(members)
                my = sum(p[1] for p in members) / len(members)
                new_centers.append((mx, my))

            centers = new_centers
            if not changed:
                break

        inertia = 0.0
        for i in range(n):
            inertia += dist2(points[i], centers[labels[i]])

        if best_inertia is None or inertia < best_inertia:
            best_inertia = inertia
            best_labels = labels[:]
            best_centers = centers[:]

    return best_labels, best_centers, float(best_inertia if best_inertia is not None else 0.0)

# -------------------------
# DBSCAN (core analysis for SPM v2)
# -------------------------

def _standardize_2d(points):
    if not points:
        return [], {"mx": 0.0, "my": 0.0, "sx": 1.0, "sy": 1.0}

    xs = [p[0] for p in points]
    ys = [p[1] for p in points]
    mx = sum(xs) / len(xs)
    my = sum(ys) / len(ys)

    vx = sum((x - mx) ** 2 for x in xs) / len(xs)
    vy = sum((y - my) ** 2 for y in ys) / len(ys)
    sx = math.sqrt(vx) if vx > 1e-12 else 1.0
    sy = math.sqrt(vy) if vy > 1e-12 else 1.0

    z = [((x - mx) / sx, (y - my) / sy) for x, y in points]
    return z, {"mx": mx, "my": my, "sx": sx, "sy": sy}


def _kth_neighbor_distances(points, k):
    n = len(points)
    kk = max(1, min(k, n - 1))
    out = []

    for i in range(n):
        xi, yi = points[i]
        ds = []
        for j in range(n):
            if i == j:
                continue
            xj, yj = points[j]
            ds.append(math.hypot(xi - xj, yi - yj))
        ds.sort()
        out.append(ds[kk - 1])

    return out


def _elbow_epsilon(kdists):
    ys = sorted(kdists)
    if len(ys) < 3:
        return ys[-1]

    y0, y1 = ys[0], ys[-1]
    ys_n = [(y - y0) / (y1 - y0) for y in ys]

    ax, ay = 0, 0
    bx, by = len(ys) - 1, 1

    def dist(px, py):
        return abs((by - ay) * px - (bx - ax) * py + bx * ay - by * ax) / math.hypot(by - ay, bx - ax)

    best_i, best_d = 0, -1
    for i, y in enumerate(ys_n):
        d = dist(i, y)
        if d > best_d:
            best_i, best_d = i, d

    return ys[best_i]


def _dbscan_2d(points, eps, min_samples):
    n = len(points)
    neigh = [[] for _ in range(n)]

    for i in range(n):
        for j in range(n):
            if math.hypot(points[i][0] - points[j][0], points[i][1] - points[j][1]) <= eps:
                neigh[i].append(j)

    is_core = [len(neigh[i]) >= min_samples for i in range(n)]
    labels = [-1] * n
    visited = [False] * n
    cid = 0

    for i in range(n):
        if visited[i] or not is_core[i]:
            continue

        visited[i] = True
        labels[i] = cid
        seeds = list(neigh[i])

        k = 0
        while k < len(seeds):
            j = seeds[k]
            if not visited[j]:
                visited[j] = True
                if is_core[j]:
                    seeds.extend([x for x in neigh[j] if x not in seeds])
            if labels[j] == -1:
                labels[j] = cid
            k += 1

        cid += 1

    return labels, is_core


def _is_finite_number(v: Any) -> bool:
    """Return True for finite int/float values."""
    return isinstance(v, (int, float)) and math.isfinite(float(v))


def _percentile(values: List[float], q: float) -> Optional[float]:
    """Small dependency-free percentile helper."""
    vals = sorted(float(v) for v in values if _is_finite_number(v))
    if not vals:
        return None
    if len(vals) == 1:
        return vals[0]
    q = max(0.0, min(1.0, float(q)))
    pos = (len(vals) - 1) * q
    lo = int(math.floor(pos))
    hi = int(math.ceil(pos))
    if lo == hi:
        return vals[lo]
    frac = pos - lo
    return vals[lo] * (1.0 - frac) + vals[hi] * frac


def _normalize_matrix_01(M: List[List[Optional[float]]]) -> List[List[Optional[float]]]:
    """Normalize off-diagonal matrix values to 0..1 while preserving None."""
    vals: List[float] = []
    for i in range(len(M)):
        for j in range(len(M)):
            if i != j and _is_finite_number(M[i][j]):
                vals.append(float(M[i][j]))

    if not vals:
        return [[0.0 if i == j else None for j in range(len(M))] for i in range(len(M))]

    mn = min(vals)
    mx = max(vals)
    span = mx - mn
    out: List[List[Optional[float]]] = []
    for i in range(len(M)):
        row: List[Optional[float]] = []
        for j in range(len(M)):
            if i == j:
                row.append(0.0)
            elif _is_finite_number(M[i][j]):
                row.append(0.0 if span <= 1e-12 else round((float(M[i][j]) - mn) / span, 6))
            else:
                row.append(None)
        out.append(row)
    return out


def _fill_symmetric_distance_matrix(M: List[List[Optional[float]]]) -> Tuple[List[List[float]], int, float]:
    """Fill missing symmetric distances with the off-diagonal median."""
    n = len(M)
    vals = [float(M[i][j]) for i in range(n) for j in range(i + 1, n) if _is_finite_number(M[i][j])]
    median = _percentile(vals, 0.5)
    if median is None:
        median = 1.0

    missing_count = 0
    out: List[List[float]] = [[0.0] * n for _ in range(n)]
    for i in range(n):
        for j in range(i + 1, n):
            v = M[i][j]
            if not _is_finite_number(v):
                v = median
                missing_count += 1
            vv = max(0.0, float(v))
            out[i][j] = vv
            out[j][i] = vv
    return out, missing_count, float(median)


def _normalize_points_01(points: List[Tuple[float, float]]) -> List[Tuple[float, float]]:
    """Scale 2D points to 0..1 for screen display."""
    if not points:
        return []
    xs = [p[0] for p in points]
    ys = [p[1] for p in points]
    minx, maxx = min(xs), max(xs)
    miny, maxy = min(ys), max(ys)
    rx = (maxx - minx) if (maxx - minx) > 1e-12 else 1.0
    ry = (maxy - miny) if (maxy - miny) > 1e-12 else 1.0
    return [((x - minx) / rx, (y - miny) / ry) for x, y in points]


def _placement_xy_normalized(v: Any, fallback_self_mode: str = "center") -> Optional[Tuple[float, float]]:
    """Normalize one placement item using w/h when available."""
    if not isinstance(v, dict):
        return None
    x = v.get("x")
    y = v.get("y")
    if not (_is_finite_number(x) and _is_finite_number(y)):
        return None

    w = v.get("w")
    h = v.get("h")
    if _is_finite_number(w) and float(w) > 0 and _is_finite_number(h) and float(h) > 0:
        return (max(0.0, min(1.0, float(x) / float(w))), max(0.0, min(1.0, float(y) / float(h))))

    # Older saved data has raw canvas coordinates without w/h. Keep it usable by
    # normalizing later within the respondent's map.
    return (float(x), float(y))


def _respondent_points(names: List[str], respondent: str, placements: Dict[str, Any]) -> Tuple[List[Tuple[float, float]], List[bool]]:
    """Build one respondent's full map, including self, with safe normalization."""
    raw: List[Tuple[float, float]] = []
    valid: List[bool] = []
    used_explicit_canvas = False

    for nm in names:
        if nm == respondent:
            raw.append((0.5, 0.5))
            valid.append(True)
            continue
        p = _placement_xy_normalized((placements or {}).get(nm))
        if p is None:
            raw.append((0.0, 0.0))
            valid.append(False)
            continue
        raw.append(p)
        valid.append(True)
        v = (placements or {}).get(nm)
        if isinstance(v, dict) and _is_finite_number(v.get("w")) and _is_finite_number(v.get("h")):
            used_explicit_canvas = True

    if used_explicit_canvas:
        return raw, valid

    # Backward compatible path for legacy raw pixel coordinates. It matches the
    # old app's self-at-origin idea, but returns a stable 0..1 map.
    legacy_raw: List[Tuple[float, float]] = []
    for idx, nm in enumerate(names):
        legacy_raw.append((0.0, 0.0) if nm == respondent else raw[idx])

    xs = [legacy_raw[i][0] for i in range(len(names)) if valid[i]]
    ys = [legacy_raw[i][1] for i in range(len(names)) if valid[i]]
    if len(xs) >= 2 and len(ys) >= 2:
        minx, maxx = min(xs), max(xs)
        miny, maxy = min(ys), max(ys)
        rx = (maxx - minx) if (maxx - minx) > 1e-12 else 1.0
        ry = (maxy - miny) if (maxy - miny) > 1e-12 else 1.0
        return [((x - minx) / rx, (y - miny) / ry) for x, y in legacy_raw], valid
    return raw, valid


def _cosine_distance_pair(v1: List[Optional[float]], v2: List[Optional[float]], min_common: int = 2) -> Optional[float]:
    """Compute 0..1 cosine distance on shared non-missing dimensions."""
    a: List[float] = []
    b: List[float] = []
    for x, y in zip(v1, v2):
        if _is_finite_number(x) and _is_finite_number(y):
            a.append(float(x))
            b.append(float(y))
    if len(a) < min_common:
        return None
    na = math.sqrt(sum(x * x for x in a))
    nb = math.sqrt(sum(y * y for y in b))
    if na <= 1e-12 or nb <= 1e-12:
        return None
    sim = sum(x * y for x, y in zip(a, b)) / (na * nb)
    sim = max(-1.0, min(1.0, sim))
    return round((1.0 - sim) / 2.0, 6)


def build_spm_result_payload(class_code: str, sid: str, alpha: float = 0.6, beta: float = 0.4) -> Dict[str, Any]:
    """Build the research-facing SPM result payload from student placement data."""
    students = db_get_students_in_class(class_code)
    excluded_names = db_get_excluded_student_names(class_code, sid)
    names = [
        (s.get("name") or "").strip()
        for s in students
        if (s.get("name") or "").strip() and (s.get("name") or "").strip() not in excluded_names
    ]
    n = len(names)
    submitted = db_list_submitted_student_sessions(class_code, sid)
    submitted = [s for s in submitted if (s.get("student_name") or "").strip() in names]
    n_submitted = len(submitted)
    generated_at = datetime.utcnow().isoformat() + "Z"

    method = {
        "main": "DBSCAN",
        "embedding": "classical_MDS",
        "distance_model": "direct_structural_mixed",
        "alpha": float(alpha),
        "beta": float(beta),
        "missing_policy": "pairwise_then_median_imputation",
        "kmeans_role": "optional_summary_only",
    }

    if n == 0:
        return {
            "ok": True,
            "status": "no_students",
            "class_code": class_code,
            "session_id": sid,
            "n_students": 0,
            "n_submitted": 0,
            "generated_at": generated_at,
            "method": method,
            "students": [],
            "clusters": [],
            "outliers": [],
            "distributed": [],
            "boundaries": [],
            "distance_matrix": {"names": [], "values": []},
            "diagnostics": {"missing_count": 0},
            "summary": spm_result_summary(0, 0, 0, 0, []),
            "kmeans_optional": {"enabled": False, "reason": "no_students"},
        }

    name_to_idx = {nm: i for i, nm in enumerate(names)}
    respondent_maps: Dict[str, Dict[str, Any]] = {}
    for item in submitted:
        respondent = (item.get("student_name") or "").strip()
        pts, valid = _respondent_points(names, respondent, item.get("placements") or {})
        respondent_maps[respondent] = {"points": pts, "valid": valid}

    map_mats: Dict[str, List[List[Optional[float]]]] = {}
    for respondent, data in respondent_maps.items():
        map_mats[respondent] = distance_matrix(data["points"], data["valid"])

    direct_raw: List[List[Optional[float]]] = [[None] * n for _ in range(n)]
    for i in range(n):
        direct_raw[i][i] = 0.0
        for j in range(i + 1, n):
            vals: List[float] = []
            mi = map_mats.get(names[i])
            mj = map_mats.get(names[j])
            if mi and _is_finite_number(mi[i][j]):
                vals.append(float(mi[i][j]))
            if mj and _is_finite_number(mj[j][i]):
                vals.append(float(mj[j][i]))
            if vals:
                direct_raw[i][j] = direct_raw[j][i] = round(sum(vals) / len(vals), 6)

    perception_vectors: List[List[Optional[float]]] = []
    for i, nm in enumerate(names):
        M = map_mats.get(nm)
        if not M:
            perception_vectors.append([None] * n)
            continue
        row: List[Optional[float]] = []
        for k in range(n):
            row.append(None if k == i else M[i][k])
        perception_vectors.append(row)

    vector_struct: List[List[Optional[float]]] = [[None] * n for _ in range(n)]
    third_party_struct: List[List[Optional[float]]] = [[None] * n for _ in range(n)]
    for i in range(n):
        vector_struct[i][i] = 0.0
        third_party_struct[i][i] = 0.0
        for j in range(i + 1, n):
            cd = _cosine_distance_pair(perception_vectors[i], perception_vectors[j])
            vector_struct[i][j] = vector_struct[j][i] = cd

            vals: List[float] = []
            for respondent, M in map_mats.items():
                # Prefer third-party maps, but allow all maps as a fallback when
                # the class has few submissions.
                if respondent in [names[i], names[j]] and n_submitted >= 3:
                    continue
                if _is_finite_number(M[i][j]):
                    vals.append(float(M[i][j]))
            if vals:
                third_party_struct[i][j] = third_party_struct[j][i] = round(sum(vals) / len(vals), 6)

    vector_norm = _normalize_matrix_01(vector_struct)
    third_norm = _normalize_matrix_01(third_party_struct)
    structural_raw: List[List[Optional[float]]] = [[None] * n for _ in range(n)]
    for i in range(n):
        structural_raw[i][i] = 0.0
        for j in range(i + 1, n):
            vals = [v for v in [vector_norm[i][j], third_norm[i][j]] if _is_finite_number(v)]
            if vals:
                structural_raw[i][j] = structural_raw[j][i] = round(sum(float(v) for v in vals) / len(vals), 6)

    direct_norm = _normalize_matrix_01(direct_raw)
    struct_norm = _normalize_matrix_01(structural_raw)

    final_optional: List[List[Optional[float]]] = [[None] * n for _ in range(n)]
    direct_present = struct_present = 0
    for i in range(n):
        final_optional[i][i] = 0.0
        for j in range(i + 1, n):
            d = direct_norm[i][j]
            s = struct_norm[i][j]
            if _is_finite_number(d):
                direct_present += 1
            if _is_finite_number(s):
                struct_present += 1

            if _is_finite_number(d) and _is_finite_number(s):
                v = alpha * float(d) + beta * float(s)
            elif _is_finite_number(d):
                v = float(d)
            elif _is_finite_number(s):
                v = float(s)
            else:
                v = None
            final_optional[i][j] = final_optional[j][i] = round(v, 6) if _is_finite_number(v) else None

    final_D, missing_count, imputed_value = _fill_symmetric_distance_matrix(final_optional)
    mds = classical_mds_2d(final_D)
    display_pts = _normalize_points_01(mds)

    labels = [-1] * n
    is_core = [False] * n
    dbscan_status = "not_run"
    eps: Optional[float] = None
    min_samples: Optional[int] = None

    if n >= 3 and n_submitted >= 3 and len(display_pts) == n:
        zpts, zstats = _standardize_2d(display_pts)
        min_samples = max(2, min(3, n))
        kd = _kth_neighbor_distances(zpts, max(1, min_samples - 1))
        eps = _percentile(kd, 0.75)
        if eps is None or not math.isfinite(eps) or eps <= 1e-12:
            pair_ds = [math.hypot(zpts[i][0] - zpts[j][0], zpts[i][1] - zpts[j][1]) for i in range(n) for j in range(i + 1, n)]
            eps = _percentile(pair_ds, 0.5) or 0.5
        eps = max(float(eps), 1e-6)
        labels, is_core = _dbscan_2d(zpts, eps, min_samples)
        dbscan_status = "ok"
    else:
        zstats = {"mx": 0.0, "my": 0.0, "sx": 1.0, "sy": 1.0}

    cluster_members: Dict[int, List[str]] = {}
    outliers: List[str] = []
    boundaries: List[str] = []
    student_rows: List[Dict[str, Any]] = []
    for i, nm in enumerate(names):
        label = int(labels[i]) if i < len(labels) else -1
        if dbscan_status != "ok":
            status = "not_enough_data"
            status_label = "자료 부족"
        elif label == -1:
            status = "outlier"
            status_label = "분산형"
            outliers.append(nm)
        elif is_core[i]:
            status = "core"
            status_label = "밀집형"
            cluster_members.setdefault(label, []).append(nm)
        else:
            status = "boundary"
            status_label = "경계형"
            boundaries.append(nm)
            cluster_members.setdefault(label, []).append(nm)

        x, y = display_pts[i] if i < len(display_pts) else (0.5, 0.5)
        student_rows.append({
            "name": nm,
            "x": round(float(x), 6),
            "y": round(float(y), 6),
            "cluster_id": label,
            "status": status,
            "status_label": status_label,
            "is_noise": dbscan_status == "ok" and label == -1,
        })

    clusters: List[Dict[str, Any]] = []
    for cid in sorted(cluster_members):
        members = sorted(cluster_members[cid])
        idxs = [name_to_idx[m] for m in members if m in name_to_idx]
        cx = sum(student_rows[i]["x"] for i in idxs) / len(idxs) if idxs else 0.0
        cy = sum(student_rows[i]["y"] for i in idxs) / len(idxs) if idxs else 0.0
        clusters.append({
            "cluster_id": int(cid),
            "size": len(members),
            "members": members,
            "center_x": round(cx, 6),
            "center_y": round(cy, 6),
            "density_summary": f"{len(members)}명의 학생이 비교적 가까운 관계 영역으로 묶였습니다.",
        })

    kmeans_optional = build_spm_kmeans_optional(names, display_pts)
    summary = spm_result_summary(
        n_students=n,
        n_submitted=n_submitted,
        cluster_count=len(clusters),
        boundary_count=len(boundaries),
        outlier_names=outliers,
    )

    return {
        "ok": True,
        "status": dbscan_status,
        "class_code": class_code,
        "session_id": sid,
        "n_students": n,
        "n_submitted": n_submitted,
        "generated_at": generated_at,
        "method": method,
        "students": student_rows,
        "clusters": clusters,
        "outliers": outliers,
        "distributed": outliers,
        "boundaries": boundaries,
        "distance_matrix": {
            "names": names,
            "values": [[round(float(v), 6) for v in row] for row in final_D],
        },
        "diagnostics": {
            "missing_count": missing_count,
            "imputed_value": round(imputed_value, 6),
            "direct_pairs": direct_present,
            "structural_pairs": struct_present,
            "dbscan": {
                "eps": round(float(eps), 6) if eps is not None else None,
                "min_samples": min_samples,
                "standardize": zstats,
            },
        },
        "summary": summary,
        "kmeans_optional": kmeans_optional,
    }


def build_spm_kmeans_optional(names: List[str], points: List[Tuple[float, float]]) -> Dict[str, Any]:
    """Build optional k-means grouping for compact comparison only."""
    n = len(points)
    if n < 2:
        return {"enabled": False, "reason": "not_enough_points"}
    k = 2 if n < 6 else 3
    k = max(2, min(4, min(k, n)))
    labels, centers, inertia = kmeans_2d(points, k=k, n_init=10, max_iter=60, seed=42)
    groups: List[Dict[str, Any]] = []
    for cid in range(k):
        members = [names[i] for i, lb in enumerate(labels) if int(lb) == cid]
        cx, cy = centers[cid] if cid < len(centers) else (0.0, 0.0)
        groups.append({
            "group_no": cid + 1,
            "size": len(members),
            "members": sorted(members),
            "center_x": round(float(cx), 6),
            "center_y": round(float(cy), 6),
        })
    return {
        "enabled": True,
        "k": k,
        "inertia": round(float(inertia), 6),
        "groups": groups,
        "points": [{"name": names[i], "x": round(points[i][0], 6), "y": round(points[i][1], 6), "group_no": int(labels[i]) + 1} for i in range(n)],
    }


def spm_result_summary(n_students: int, n_submitted: int, cluster_count: int, boundary_count: int, outlier_names: List[str]) -> Dict[str, str]:
    """Create teacher-facing, non-diagnostic interpretation text."""
    if n_students <= 0:
        short = "아직 분석할 학생 명단이 없습니다."
    elif n_submitted <= 0:
        short = "아직 제출된 학생 배치 데이터가 없어 관계 구조를 생성하지 않았습니다."
    elif n_submitted < 3:
        short = "제출 학생 수가 적어 관계 지도 요약은 참고 수준으로만 표시됩니다."
    else:
        parts = [f"이번 회차에서는 학생 인식상 {cluster_count}개의 밀집된 관계 영역이 나타났습니다."]
        if boundary_count > 0:
            parts.append(f"{boundary_count}명의 학생은 여러 관계 영역 사이의 경계 위치에 가깝게 나타났습니다.")
        if outlier_names:
            parts.append(f"{len(outlier_names)}명의 학생은 특정 밀집 영역에 강하게 포함되지 않는 분산형으로 나타났습니다.")
        short = " ".join(parts)

    return {
        "short": short,
        "structure": "이 결과는 직접 거리와 제3자 배치 패턴을 함께 반영한 학급 전체 관계 구조 요약입니다.",
        "cautions": "이 결과는 학생 개인을 진단하거나 관계 상태를 확정하기 위한 것이 아닙니다. 학생 인식 기반 공간 배치 결과를 바탕으로 학급 관계 구조를 이해하기 위한 참고 자료입니다. 실제 생활지도 판단은 교사의 관찰, 상담, 또래지명 결과 등을 함께 고려하여 이루어져야 합니다.",
    }


def _anon_id_map(names: List[str]) -> Dict[str, str]:
    """Map student names to stable anonymous IDs for external explanation APIs."""
    return {name: f"S{i + 1:02d}" for i, name in enumerate(names)}


def _distance_bucket(value: Optional[float]) -> Optional[str]:
    """Convert a normalized distance into an easy teacher-facing level."""
    if not _is_finite_number(value):
        return None
    v = max(0.0, min(1.0, float(value)))
    if v <= 0.33:
        return "가까움"
    if v <= 0.66:
        return "중간"
    return "멀음"


def _relation_type_from_status(status: str) -> str:
    """Convert internal DBSCAN status into non-stigmatizing wording."""
    if status == "core":
        return "밀집형"
    if status == "boundary":
        return "경계형"
    if status == "outlier":
        return "분산형"
    return "자료 부족"


def build_privacy_safe_relation_payload(class_code: str, sid: str, student_id: Optional[str] = None) -> Dict[str, Any]:
    """Build an anonymized relation payload that is safe to send to an AI API."""
    result = build_spm_result_payload(class_code, sid)
    students = result.get("students") or []
    names = [str(s.get("name") or "") for s in students if str(s.get("name") or "")]
    anon = _anon_id_map(names)

    dense_count = sum(1 for s in students if s.get("status") == "core")
    boundary_count = sum(1 for s in students if s.get("status") == "boundary")
    distributed_count = sum(1 for s in students if s.get("status") == "outlier")
    cluster_count = len(result.get("clusters") or [])
    n_students = int(result.get("n_students") or 0)
    n_submitted = int(result.get("n_submitted") or 0)
    overall_density = round(dense_count / n_students, 4) if n_students else 0.0

    payload: Dict[str, Any] = {
        "ok": bool(result.get("ok")),
        "session_id": str(sid),
        "n_students": n_students,
        "n_submitted": n_submitted,
        "counts": {
            "dense": dense_count,
            "boundary": boundary_count,
            "distributed": distributed_count,
            "cluster_count": cluster_count,
        },
        "overall_density": overall_density,
        "summary": {
            "short": result.get("summary", {}).get("short", ""),
            "cautions": result.get("summary", {}).get("cautions", ""),
        },
    }

    selected_name: Optional[str] = None
    if student_id:
        sid_upper = str(student_id).strip().upper()
        for nm, aid in anon.items():
            if aid == sid_upper:
                selected_name = nm
                break

    if selected_name and selected_name in names:
        idx = names.index(selected_name)
        selected_row = next((s for s in students if s.get("name") == selected_name), {})
        submitted = db_list_submitted_student_sessions(class_code, sid)
        respondent_maps: Dict[str, List[List[Optional[float]]]] = {}
        for item in submitted:
            respondent = (item.get("student_name") or "").strip()
            if respondent not in names:
                continue
            pts, valid = _respondent_points(names, respondent, item.get("placements") or {})
            respondent_maps[respondent] = distance_matrix(pts, valid)

        def _direction_bucket(respondent: str, i: int, j: int) -> Optional[str]:
            M = respondent_maps.get(respondent)
            if not M or i >= len(M) or j >= len(M[i]) or not _is_finite_number(M[i][j]):
                return None
            return _distance_bucket(float(M[i][j]) / math.sqrt(2.0))

        student_to_peers = {"near": [], "middle": [], "far": []}
        peers_to_student = {"near": [], "middle": [], "far": []}
        asymmetry: List[Dict[str, str]] = []
        for peer_idx, peer_name in enumerate(names):
            if peer_name == selected_name:
                continue
            peer_id = anon[peer_name]
            a = _direction_bucket(selected_name, idx, peer_idx)
            b = _direction_bucket(peer_name, peer_idx, idx)

            if a == "가까움":
                student_to_peers["near"].append(peer_id)
            elif a == "중간":
                student_to_peers["middle"].append(peer_id)
            elif a == "멀음":
                student_to_peers["far"].append(peer_id)

            if b == "가까움":
                peers_to_student["near"].append(peer_id)
            elif b == "중간":
                peers_to_student["middle"].append(peer_id)
            elif b == "멀음":
                peers_to_student["far"].append(peer_id)

            if a and b and a != b:
                asymmetry.append({
                    "peer_id": peer_id,
                    "student_to_peer": a,
                    "peer_to_student": b,
                })

        payload["selected_student"] = {
            "student_id": anon[selected_name],
            "relation_type": _relation_type_from_status(str(selected_row.get("status") or "")),
            "student_to_peers_counts": {k: len(v) for k, v in student_to_peers.items()},
            "peers_to_student_counts": {k: len(v) for k, v in peers_to_student.items()},
            "student_to_peers": student_to_peers,
            "peers_to_student": peers_to_student,
            "asymmetry_count": len(asymmetry),
            "asymmetry": asymmetry[:8],
        }

    return payload


def _fallback_ai_relation_answer(question_type: str, payload: Dict[str, Any]) -> str:
    """Return a local AI-style explanation when OpenAI is not configured."""
    selected = payload.get("selected_student") or {}
    relation_type = selected.get("relation_type") or "학급 전체"
    counts = payload.get("counts") or {}
    dense = counts.get("dense", 0)
    boundary = counts.get("boundary", 0)
    distributed = counts.get("distributed", 0)
    asymmetry_count = selected.get("asymmetry_count", 0)
    student_to = selected.get("student_to_peers_counts") or {}
    peer_to = selected.get("peers_to_student_counts") or {}

    if question_type == "why_distributed" and relation_type == "분산형":
        return "이 학생은 밀집 영역의 중심에 포함되지 않고, 여러 친구와의 거리가 중간 이상으로 넓게 나타나 분산형으로 판단되었습니다."
    if question_type == "student_to_peers":
        return f"이 학생의 배치에서는 가까움 {student_to.get('near', 0)}명, 중간 {student_to.get('middle', 0)}명, 멀음 {student_to.get('far', 0)}명으로 나타났습니다."
    if question_type == "peers_to_student":
        return f"친구들이 이 학생을 배치한 결과는 가까움 {peer_to.get('near', 0)}명, 중간 {peer_to.get('middle', 0)}명, 멀음 {peer_to.get('far', 0)}명입니다."
    elif question_type == "asymmetry":
        return f"서로 다르게 배치한 관계가 {asymmetry_count}건 확인됩니다. 서로 느낀 거리감이 완전히 같지 않은 짝을 살펴볼 수 있습니다."
    return f"이번 회차는 밀집형 {dense}명, 경계형 {boundary}명, 분산형 {distributed}명으로 요약됩니다. 학급 전체 구조를 중심으로 살펴보세요."


def _limit_ai_answer(answer: str, limit: int = 200) -> str:
    """Keep relation helper answers short enough for the result page."""
    text_value = " ".join(str(answer or "").split())
    if len(text_value) <= limit:
        return text_value
    return text_value[: max(0, limit - 1)].rstrip() + "…"


def _join_names_short(names: List[str], limit: int = 6) -> str:
    """Format a short list of student names for teacher-visible answers."""
    clean = [str(n).strip() for n in names if str(n).strip()]
    if not clean:
        return "없음"
    shown = clean[:limit]
    extra = len(clean) - len(shown)
    suffix = f" 외 {extra}명" if extra > 0 else ""
    return ", ".join(shown) + suffix


def _visible_relation_context(class_code: str, sid: str, student_id: str, privacy_payload: Dict[str, Any]) -> Dict[str, Any]:
    """Build teacher-visible name context from anonymized relation payload."""
    result = build_spm_result_payload(class_code, sid)
    names = [str(s.get("name") or "") for s in (result.get("students") or []) if str(s.get("name") or "")]
    anon = _anon_id_map(names)
    id_to_name = {aid: name for name, aid in anon.items()}
    students_by_name = {str(s.get("name") or ""): s for s in (result.get("students") or [])}
    selected_id = (student_id or "").strip().upper()
    selected_name = id_to_name.get(selected_id, "")
    selected = privacy_payload.get("selected_student") or {}

    def _ids_to_names(ids: List[str]) -> List[str]:
        return [id_to_name.get(str(x), str(x)) for x in (ids or [])]

    student_to = selected.get("student_to_peers") or {}
    peers_to = selected.get("peers_to_student") or {}
    asymmetry = []
    for item in selected.get("asymmetry") or []:
        peer_id = str(item.get("peer_id") or "")
        asymmetry.append({
            "peer_name": id_to_name.get(peer_id, peer_id),
            "student_to_peer": item.get("student_to_peer") or "",
            "peer_to_student": item.get("peer_to_student") or "",
        })

    return {
        "selected_id": selected_id,
        "selected_name": selected_name,
        "relation_type": selected.get("relation_type") or _relation_type_from_status(str((students_by_name.get(selected_name) or {}).get("status") or "")),
        "student_to": {
            "near": _ids_to_names(student_to.get("near") or []),
            "middle": _ids_to_names(student_to.get("middle") or []),
            "far": _ids_to_names(student_to.get("far") or []),
        },
        "peers_to": {
            "near": _ids_to_names(peers_to.get("near") or []),
            "middle": _ids_to_names(peers_to.get("middle") or []),
            "far": _ids_to_names(peers_to.get("far") or []),
        },
        "student_to_counts": selected.get("student_to_peers_counts") or {},
        "peers_to_counts": selected.get("peers_to_student_counts") or {},
        "asymmetry_count": int(selected.get("asymmetry_count") or 0),
        "asymmetry": asymmetry,
        "counts": privacy_payload.get("counts") or {},
    }


def _relation_chat_followups(question_type: str) -> List[Dict[str, str]]:
    """Return next suggested questions for the relation helper."""
    followups = {
        "why_type": [
            {"type": "why_type_basis", "label": "어떤 기준으로 판단했나요?"},
            {"type": "peers_to_student", "label": "친구들이 본 결과도 비슷한가요?"},
            {"type": "distance_gap", "label": "거리감 차이가 큰 친구가 있나요?"},
        ],
        "why_distributed": [
            {"type": "why_type_basis", "label": "어떤 기준으로 판단했나요?"},
            {"type": "peers_to_student", "label": "친구들이 본 결과도 비슷한가요?"},
            {"type": "distance_gap", "label": "거리감 차이가 큰 친구가 있나요?"},
        ],
        "student_to_peers": [
            {"type": "student_near_list", "label": "가깝게 둔 친구는 누구인가요?"},
            {"type": "student_far_list", "label": "멀게 둔 친구가 있나요?"},
            {"type": "student_spread", "label": "넓게 배치했나요, 모아서 배치했나요?"},
        ],
        "peers_to_student": [
            {"type": "peers_near_list", "label": "이 학생을 가깝게 둔 친구는 누구인가요?"},
            {"type": "peers_far_list", "label": "이 학생을 멀게 둔 친구가 있나요?"},
            {"type": "mutual_similarity", "label": "이 학생이 본 것과 비슷한가요?"},
        ],
        "distance_gap": [
            {"type": "distance_gap_detail", "label": "어떤 친구와 차이가 컸나요?"},
            {"type": "distance_gap_largest", "label": "차이가 어떻게 나타났나요?"},
            {"type": "distance_gap_effect", "label": "이 차이를 결과에 어떻게 반영했나요?"},
        ],
        "asymmetry": [
            {"type": "distance_gap_detail", "label": "어떤 친구와 차이가 컸나요?"},
            {"type": "distance_gap_largest", "label": "차이가 어떻게 나타났나요?"},
            {"type": "distance_gap_effect", "label": "이 차이를 결과에 어떻게 반영했나요?"},
        ],
    }
    return followups.get(question_type, [
        {"type": "why_type", "label": "왜 이 유형으로 나왔나요?"},
        {"type": "student_to_peers", "label": "이 학생은 친구들을 어떻게 배치했나요?"},
        {"type": "peers_to_student", "label": "친구들은 이 학생을 어떻게 배치했나요?"},
        {"type": "distance_gap", "label": "거리감 차이가 큰 친구가 있나요?"},
    ])


def _relation_chat_local_answer(question_type: str, ctx: Dict[str, Any]) -> str:
    """Answer relation-helper questions from app data, keeping names local."""
    name = ctx.get("selected_name") or "이 학생"
    rel = ctx.get("relation_type") or "자료 부족"
    st = ctx.get("student_to_counts") or {}
    pt = ctx.get("peers_to_counts") or {}
    counts = ctx.get("counts") or {}
    gap_count = int(ctx.get("asymmetry_count") or 0)
    gaps = ctx.get("asymmetry") or []

    if question_type in {"why_type", "why_distributed"}:
        if rel == "분산형":
            return "이 학생은 밀집 영역의 중심에 포함되지 않고, 여러 친구와의 거리가 중간 이상으로 넓게 나타나 분산형으로 판단되었습니다."
        return f"{name} 학생은 관계 지도에서 {rel}으로 나타났습니다. 주변 학생들과의 최종 거리와 영역 안쪽 포함 여부를 함께 반영한 결과입니다."
    if question_type == "why_type_basis":
        return "직접 배치 거리, 친구들이 본 거리, 다른 학생을 배치한 패턴을 함께 섞어 최종 거리를 만들고, 그 지도에서 밀집 영역 포함 정도를 봅니다."
    if question_type == "student_to_peers":
        return f"이 학생의 배치에서는 가까움 {st.get('near', 0)}명, 중간 {st.get('middle', 0)}명, 멀음 {st.get('far', 0)}명으로 나타났습니다."
    if question_type == "student_near_list":
        return f"가깝게 배치한 친구는 {_join_names_short(ctx['student_to'].get('near', []))}입니다."
    if question_type == "student_far_list":
        return f"멀게 배치한 친구는 {_join_names_short(ctx['student_to'].get('far', []))}입니다."
    if question_type == "student_spread":
        return f"가까움 {st.get('near', 0)}명, 중간 {st.get('middle', 0)}명, 멀음 {st.get('far', 0)}명으로, 이 학생의 배치가 얼마나 모였는지 볼 수 있습니다."
    if question_type == "peers_to_student":
        return f"친구들이 이 학생을 배치한 결과는 가까움 {pt.get('near', 0)}명, 중간 {pt.get('middle', 0)}명, 멀음 {pt.get('far', 0)}명입니다."
    if question_type == "peers_near_list":
        return f"이 학생을 가깝게 배치한 친구는 {_join_names_short(ctx['peers_to'].get('near', []))}입니다."
    if question_type == "peers_far_list":
        return f"이 학생을 멀게 배치한 친구는 {_join_names_short(ctx['peers_to'].get('far', []))}입니다."
    if question_type == "mutual_similarity":
        if gap_count == 0:
            return "이 학생이 본 거리감과 친구들이 본 거리감의 단계 차이는 크게 확인되지 않았습니다."
        return f"서로의 거리감 단계가 다르게 나타난 배치가 {gap_count}건 있어, 일부 친구와는 인식 차이를 따로 볼 수 있습니다."
    if question_type in {"distance_gap", "asymmetry"}:
        return f"서로의 거리감 단계가 다르게 나타난 배치가 {gap_count}건 있습니다."
    if question_type in {"distance_gap_detail", "distance_gap_largest"}:
        if not gaps:
            return "거리감 단계가 다르게 나타난 친구는 확인되지 않았습니다."
        parts = [f"{g['peer_name']}: 이 학생→{g['student_to_peer']}, 친구→{g['peer_to_student']}" for g in gaps[:3]]
        return " / ".join(parts)
    if question_type == "distance_gap_effect":
        return "이 차이는 한쪽 응답만으로 판단하지 않고, 양방향 직접 거리와 다른 학생 배치 패턴을 섞은 최종 거리 계산에 함께 반영됩니다."
    return f"이번 회차는 밀집형 {counts.get('dense', 0)}명, 경계형 {counts.get('boundary', 0)}명, 분산형 {counts.get('distributed', 0)}명으로 요약됩니다."


def _classify_relation_free_question(question_text: str) -> str:
    """Map a teacher's free question to a safe internal relation-helper question type."""
    q = " ".join(str(question_text or "").lower().split())
    if not q:
        return "unsupported"

    risky_words = [
        "왕따", "따돌림", "괴롭", "문제 학생", "문제아", "위험", "인기",
        "싫어", "미움", "원인", "잘못", "부적응", "소외", "고립", "진단",
    ]
    if any(word in q for word in risky_words):
        return "unsupported"

    has_gap = any(word in q for word in ["거리감 차이", "차이", "다르게", "서로", "비대칭"])
    asks_who = any(word in q for word in ["누구", "어떤 친구", "어느 친구", "누구와", "명단"])
    if has_gap and asks_who:
        return "distance_gap_detail"
    if has_gap and any(word in q for word in ["반영", "영향", "결과"]):
        return "distance_gap_effect"
    if has_gap:
        return "distance_gap"

    if any(word in q for word in ["기준", "판단", "계산", "근거", "분류"]):
        return "why_type_basis"
    if any(word in q for word in ["왜", "유형", "분산", "밀집", "경계", "나왔", "나온"]):
        return "why_type"

    if any(word in q for word in ["친구들이", "다른 학생들이", "친구가"]) and any(word in q for word in ["이 학생", "해당 학생", "배치", "바라"]):
        if any(word in q for word in ["가깝", "가까"]):
            return "peers_near_list" if asks_who else "peers_to_student"
        if any(word in q for word in ["멀", "떨어"]):
            return "peers_far_list" if asks_who else "peers_to_student"
        return "peers_to_student"

    if any(word in q for word in ["이 학생", "학생이", "친구들을", "배치", "바라"]):
        if any(word in q for word in ["가깝", "가까"]):
            return "student_near_list" if asks_who else "student_to_peers"
        if any(word in q for word in ["멀", "떨어"]):
            return "student_far_list" if asks_who else "student_to_peers"
        if any(word in q for word in ["넓", "흩어", "모여"]):
            return "student_spread"
        return "student_to_peers"

    return "unsupported"


def _safe_relation_question_refusal() -> str:
    """Return a short safe answer for questions outside the relation-map scope."""
    return "이 질문은 원인이나 학생 상태를 단정할 수 있어 답하기 어렵습니다. 거리감, 유형 판단, 배치 차이처럼 관계 지도 안의 정보로 질문해 주세요."


def _openai_relation_answer(question_type: str, payload: Dict[str, Any]) -> Tuple[str, str]:
    """Generate a privacy-safe teacher explanation, falling back locally if needed."""
    api_key = (os.environ.get("OPENAI_API_KEY") or "").strip()
    if not api_key:
        return _fallback_ai_relation_answer(question_type, payload), "fallback"

    model = (os.environ.get("OPENAI_MODEL") or "gpt-5.4-mini").strip()
    system_prompt = (
        "학급 관계 구조 지표를 교사용 설명문으로 바꾸는 보조자입니다. "
        "학생 개인을 진단하지 않고, 결과를 확정적으로 말하지 않습니다. "
        "분산형은 특정 밀집 영역에 강하게 속하지 않은 구조로 설명합니다. "
        "학생에게 낙인이 될 수 있는 표현은 쓰지 않습니다."
    )
    user_prompt = (
        "다음 구조 지표를 바탕으로 교사가 읽기 쉬운 한국어 설명을 작성해 주세요.\n"
        "공백 포함 200자 이내로 답하세요.\n"
        "문단 번호, 제목, 불필요한 주의 문구는 쓰지 마세요.\n"
        "왜 그렇게 판단했는지 계산 근거 중심으로 짧게 말하세요.\n"
        "학생 이름을 추측하거나 만들지 마세요.\n"
        f"질문 유형: {question_type}\n"
        f"구조 지표 JSON: {json.dumps(payload, ensure_ascii=False)}"
    )
    try:
        resp = requests.post(
            "https://api.openai.com/v1/responses",
            headers={"Authorization": f"Bearer {api_key}", "Content-Type": "application/json"},
            json={
                "model": model,
                "input": [
                    {"role": "system", "content": [{"type": "input_text", "text": system_prompt}]},
                    {"role": "user", "content": [{"type": "input_text", "text": user_prompt}]},
                ],
                "max_output_tokens": 160,
            },
            timeout=20,
        )
        resp.raise_for_status()
        data = resp.json()
        answer = (data.get("output_text") or "").strip()
        if not answer:
            chunks: List[str] = []
            for item in data.get("output") or []:
                for content in item.get("content") or []:
                    txt = content.get("text") or ""
                    if txt:
                        chunks.append(str(txt))
            answer = "\n".join(chunks).strip()
        if answer:
            return _limit_ai_answer(answer), "openai"
    except Exception:
        app.logger.exception("privacy-safe relation AI call failed")
    return _limit_ai_answer(_fallback_ai_relation_answer(question_type, payload)), "fallback"


def dbscan_teacher_summary(
    n_total: int,
    dense_count: int,
    boundary_count: int,
    isolated_count: int,
    cluster_sizes: list,
) -> dict:
    """
    DBSCAN 단일 회차 결과를 교사용 관찰 문장으로 요약한다.
    - 관찰 중심
    - 핵심 신호 1개만 강조
    - 청유형은 "~해보는 것 어떨까요?" 톤으로 고정
    """

    n_total = int(n_total or 0)
    dense_count = int(dense_count or 0)
    boundary_count = int(boundary_count or 0)
    isolated_count = int(isolated_count or 0)
    cluster_sizes = cluster_sizes or []

    # 안전 처리
    if n_total <= 0:
        return {
            "structure_summary": "이번 회차 결과를 요약할 수 없습니다.",
            "key_signal": "참여 학생 수가 확인되지 않습니다.",
            "reflection_prompt": "학생 참여 현황을 먼저 확인해 보는 것 어떨까요?",
            "note": "이 결과는 학생 인식 기반 배치에 따른 구조 요약이며, 개별 관계를 확정하지 않습니다.",
            "rule": "none",
        }

    isolated_ratio = isolated_count / n_total
    boundary_ratio = boundary_count / n_total

    cluster_count = len(cluster_sizes)
    largest_cluster_ratio = (max(cluster_sizes) / n_total) if cluster_sizes else 0.0

    structure_summary = (
        f"이번 회차에서는 밀집형 {dense_count}명, 경계형 {boundary_count}명, 분산형 {isolated_count}명으로 나타났습니다."
    )

    note = "이 결과는 학생 인식 기반 배치에 따른 구조 요약이며, 개별 관계를 확정하지 않습니다."

    # 우선순위: A(분산형) > C(쏠림) > B(경계) > D(분산) > none
    # A: 분산형 신호
    if (isolated_ratio >= 0.15) or (isolated_count >= 3):
        return {
            "structure_summary": structure_summary,
            "key_signal": "분산형으로 나타난 학생이 상대적으로 많아, 특정 밀집 영역에 강하게 포함되지 않는 관계 양상을 함께 살펴볼 필요가 있어 보입니다.",
            "reflection_prompt": "해당 학생들의 최근 교내외 활동 변화나 학교생활 경험을 함께 떠올려 보는 것 어떨까요?",
            "note": note,
            "rule": "A_distributed",
        }

    # C: 한 집단 쏠림
    if largest_cluster_ratio >= 0.55:
        return {
            "structure_summary": structure_summary,
            "key_signal": "한 관계 중심이 비교적 크게 형성되어, 관계 구조가 중심–주변 형태로 나타날 가능성이 있습니다.",
            "reflection_prompt": "관계 중심 바깥에 있는 학생들의 참여 경험이 어떻게 형성되고 있는지 한 번 돌아보는 것 어떨까요?",
            "note": note,
            "rule": "C_center_dominance",
        }

    # B: 경계 신호
    if boundary_ratio >= 0.30:
        return {
            "structure_summary": structure_summary,
            "key_signal": "경계 위치의 학생이 비교적 많아, 관계 구조의 경계가 넓게 형성된 상태로 보입니다.",
            "reflection_prompt": "최근 교내외 활동 변화가 관계 경계에 어떤 영향을 주었는지 생각해 보는 것도 도움이 될 수 있습니다.",
            "note": note,
            "rule": "B_boundary",
        }

    # D: 분산/다중 집단
    if (cluster_count >= 3) and (largest_cluster_ratio < 0.45):
        return {
            "structure_summary": structure_summary,
            "key_signal": "관계가 여러 영역으로 나뉘어 형성되어 있는 모습이 관찰됩니다.",
            "reflection_prompt": "집단 간 교류가 자연스럽게 이루어질 수 있는 경험이 있었는지 떠올려 보는 것 어떨까요?",
            "note": note,
            "rule": "D_multi_flow",
        }

    # none
    return {
        "structure_summary": structure_summary,
        "key_signal": "관계 구조에서 뚜렷한 특이 신호는 크게 나타나지 않습니다.",
        "reflection_prompt": "현재의 관계 구조가 비교적 안정적으로 유지되고 있는지 지켜보는 것도 의미가 있을 수 있습니다.",
        "note": note,
        "rule": "none",
    }


def dbscan_change_summary(prev_counts: dict, curr_counts: dict) -> dict:
    """
    DBSCAN 회차 간 변화 서술(구조 지표 증감 기반).
    - 개인 추적 없음
    - 방향성만 제시
    - 핵심 변화 신호 1개만 강조
    """

    def _safe_int(x):
        try:
            return int(x)
        except Exception:
            return 0

    prev_n = _safe_int(prev_counts.get("n_total"))
    curr_n = _safe_int(curr_counts.get("n_total"))

    # 참여 인원이 0이면 비교 불가
    if prev_n <= 0 or curr_n <= 0:
        return {
            "change_summary": "이전 회차와의 비교 요약을 만들기 어렵습니다.",
            "change_signal": "참여 학생 수가 확인되지 않습니다.",
            "reflection_prompt": "학생 참여 현황을 먼저 확인해 보는 것 어떨까요?",
            "note": "이 변화는 학생 인식 기반 배치 결과의 비교에 따른 구조적 변화 요약이며, 개별 관계의 변화로 단정하지 않습니다.",
            "rule": "none",
        }

    prev_dense = _safe_int(prev_counts.get("dense"))
    prev_boundary = _safe_int(prev_counts.get("boundary"))
    prev_isolated = _safe_int(prev_counts.get("isolated"))
    prev_sizes = prev_counts.get("cluster_sizes") or []

    curr_dense = _safe_int(curr_counts.get("dense"))
    curr_boundary = _safe_int(curr_counts.get("boundary"))
    curr_isolated = _safe_int(curr_counts.get("isolated"))
    curr_sizes = curr_counts.get("cluster_sizes") or []

    d_isolated = curr_isolated - prev_isolated
    d_boundary = curr_boundary - prev_boundary
    d_dense = curr_dense - prev_dense

    prev_cluster_count = len(prev_sizes)
    curr_cluster_count = len(curr_sizes)

    prev_largest_ratio = (max(prev_sizes) / prev_n) if prev_sizes else 0.0
    curr_largest_ratio = (max(curr_sizes) / curr_n) if curr_sizes else 0.0

    note = "이 변화는 학생 인식 기반 배치 결과의 비교에 따른 구조적 변화 요약이며, 개별 관계의 변화로 단정하지 않습니다."

    change_summary = "이전 회차와 비교할 때, 학급 내 관계 구조에 일부 변화가 관찰됩니다."

    # 우선순위: 분산형 변화 > 경계 변화 > 중심 쏠림 변화 > 분산/수렴 > 미미
    # A: 분산형 증가/감소
    if (d_isolated >= 2) or ((d_isolated / prev_n) >= 0.10):
        return {
            "change_summary": change_summary,
            "change_signal": "분산형으로 나타난 학생의 수가 이전 회차보다 증가한 것으로 나타났습니다.",
            "reflection_prompt": "최근 교내외 활동 변화나 학교생활 경험의 변화가 있었는지 함께 떠올려 보는 것 어떨까요?",
            "note": note,
            "rule": "A_distributed_up",
            "delta": {"dense": d_dense, "boundary": d_boundary, "isolated": d_isolated},
        }

    if (d_isolated <= -2) or ((d_isolated / prev_n) <= -0.10):
        return {
            "change_summary": change_summary,
            "change_signal": "분산형으로 나타난 학생의 수가 이전 회차보다 줄어든 것으로 보입니다.",
            "reflection_prompt": "이러한 변화가 어떤 경험이나 상호작용과 함께 나타났는지 돌아보는 것도 의미가 있을 수 있습니다.",
            "note": note,
            "rule": "A_distributed_down",
            "delta": {"dense": d_dense, "boundary": d_boundary, "isolated": d_isolated},
        }

    # B: 경계 증가(너가 고정한 짧은 문장)
    if (d_boundary >= 3) or ((d_boundary / prev_n) >= 0.15):
        return {
            "change_summary": change_summary,
            "change_signal": "일부 학생들이 특정 관계 집단에 뚜렷하게 속하기보다는 여러 관계 사이에 위치한 모습이 더 많이 나타났습니다.",
            "reflection_prompt": "최근 교내외 활동 변화가 관계 형성 방식에 어떤 영향을 주었는지 생각해 보는 것도 도움이 될 수 있습니다.",
            "note": note,
            "rule": "B_boundary_up",
            "delta": {"dense": d_dense, "boundary": d_boundary, "isolated": d_isolated},
        }

    # C: 중심 쏠림 변화
    if (curr_largest_ratio - prev_largest_ratio) >= 0.15:
        return {
            "change_summary": change_summary,
            "change_signal": "한 관계 중심이 이전 회차보다 더 두드러지게 형성된 것으로 보입니다.",
            "reflection_prompt": "관계 중심 바깥에 있는 학생들의 참여 경험이 어떻게 변화했는지 한 번 돌아보는 것 어떨까요?",
            "note": note,
            "rule": "C_center_up",
            "delta": {"dense": d_dense, "boundary": d_boundary, "isolated": d_isolated},
        }

    if (prev_largest_ratio - curr_largest_ratio) >= 0.15:
        return {
            "change_summary": change_summary,
            "change_signal": "관계 중심이 이전 회차보다 완화되며, 구조가 보다 분산된 모습으로 나타났습니다.",
            "reflection_prompt": "이러한 변화가 어떤 경험과 함께 나타났는지 떠올려 보는 것도 의미가 있을 수 있습니다.",
            "note": note,
            "rule": "C_center_down",
            "delta": {"dense": d_dense, "boundary": d_boundary, "isolated": d_isolated},
        }

    # D: 집단 수 증가(분산)
    if curr_cluster_count >= prev_cluster_count + 1:
        return {
            "change_summary": change_summary,
            "change_signal": "관계 구조가 이전 회차보다 여러 영역으로 분산된 모습이 관찰됩니다.",
            "reflection_prompt": "집단 간 교류 경험이 어떻게 형성되고 있었는지 떠올려 보는 것 어떨까요?",
            "note": note,
            "rule": "D_more_flows",
            "delta": {"dense": d_dense, "boundary": d_boundary, "isolated": d_isolated},
        }

    # 미미
    return {
        "change_summary": "이전 회차와 비교했을 때, 관계 구조의 큰 변화는 두드러지게 나타나지 않습니다.",
        "change_signal": "관계 구조의 큰 변화는 두드러지게 나타나지 않습니다.",
        "reflection_prompt": "현재의 관계 구조가 비교적 안정적으로 유지되고 있는지 지켜보는 것도 의미가 있을 수 있습니다.",
        "note": note,
        "rule": "none",
        "delta": {"dense": d_dense, "boundary": d_boundary, "isolated": d_isolated},
    }

def dbscan_structure_payload(class_code, sid):
    avg = cache_get(class_code, sid, f"student_avg_{sid}") or \
          student_avg_distance_payload(class_code, sid)

    # KeyError 방지 + 빈 배열 방지
    names = avg.get("names", [])
    raw = [(p["x"], p["y"]) for p in (avg.get("mds_2d") or [])]

    Z, stats = _standardize_2d(raw)
    n = len(Z)

    # 자료 부족(또는 없음) 처리:
    # - n==0이면 math.log(0)로 500이 나던 문제를 차단
    # - n<3이면 구조 해석이 불안정하므로 no_data 처리
    if n < 3:
        points = []
        for i in range(n):
            points.append({
                "name": names[i] if i < len(names) else f"student_{i+1}",
                "x": raw[i][0],
                "y": raw[i][1],
                "state": "isolated"
            })

        counts = {
            "n_total": n,
            "dense": 0,
            "boundary": 0,
            "isolated": n,
            "cluster_sizes": []
        }

        teacher_summary = dbscan_teacher_summary(
            n_total=counts["n_total"],
            dense_count=counts["dense"],
            boundary_count=counts["boundary"],
            isolated_count=counts["isolated"],
            cluster_sizes=counts["cluster_sizes"],
        )

        return {
            "status": "no_data",
            "reason": "not_enough_points",
            "points": points,
            "fog_points": [],
            "counts": counts,
            "teacher_summary": teacher_summary,
            "params": {
                "epsilon": None,
                "min_samples": None,
                "standardize": stats
            }
        }

    min_samples = max(3, round(math.log(n)))
    kd = _kth_neighbor_distances(Z, min_samples)
    eps = max(0.15, min(0.8, _elbow_epsilon(kd)))

    labels, is_core = _dbscan_2d(Z, eps, min_samples)

    points = []
    dense = boundary = isolated = 0
    clusters = {}

    for i in range(n):
        if labels[i] == -1:
            state = "isolated"
            isolated += 1
        elif is_core[i]:
            state = "dense"
            dense += 1
            clusters[labels[i]] = clusters.get(labels[i], 0) + 1
        else:
            state = "boundary"
            boundary += 1
            clusters[labels[i]] = clusters.get(labels[i], 0) + 1

        points.append({
            "name": names[i] if i < len(names) else f"student_{i+1}",
            "x": raw[i][0],
            "y": raw[i][1],
            "state": state
        })

    counts = {
        "n_total": n,
        "dense": dense,
        "boundary": boundary,
        "isolated": isolated,
        "cluster_sizes": list(clusters.values())
    }

    teacher_summary = dbscan_teacher_summary(
        n_total=counts["n_total"],
        dense_count=counts["dense"],
        boundary_count=counts["boundary"],
        isolated_count=counts["isolated"],
        cluster_sizes=counts["cluster_sizes"],
    )

    return {
        "status": "ok",
        "points": points,
        "fog_points": [p for p in points if p["state"] == "dense"],
        "counts": counts,
        "teacher_summary": teacher_summary,
        "params": {
            "epsilon": eps,
            "min_samples": min_samples,
            "standardize": stats
        }
    }



def kmeans_summary_payload(class_code: str, sid: str, k: int) -> Dict[str, Any]:
    avg_cache_key = f"student_avg_{sid}"
    avg_payload = cache_get(class_code, sid, avg_cache_key)
    if not avg_payload:
        avg_payload = student_avg_distance_payload(class_code, sid)

    names = avg_payload.get("names") or []
    pts = avg_payload.get("mds_2d") or []
    if not names or not pts or len(names) != len(pts):
        return {
            "class_code": class_code,
            "session_id": sid,
            "k": int(k),
            "error": "avg mds_2d not available",
            "generated_at": datetime.utcnow().isoformat() + "Z",
        }

    points: List[Tuple[float, float]] = []
    for p in pts:
        x = p.get("x") if isinstance(p, dict) else None
        y = p.get("y") if isinstance(p, dict) else None
        if not (isinstance(x, (int, float)) and isinstance(y, (int, float))):
            x, y = 0.0, 0.0
        points.append((float(x), float(y)))

    kk = int(k)
    if kk < 2:
        kk = 2
    if kk > 4:
        kk = 4
    if kk > len(points):
        kk = max(2, min(4, len(points)))

    labels, centers, inertia = kmeans_2d(points, k=kk, n_init=10, max_iter=60, seed=42)

    cluster_sizes = [0] * kk
    for lb in labels:
        if 0 <= int(lb) < kk:
            cluster_sizes[int(lb)] += 1

    cluster_mean_radius: List[Optional[float]] = []
    for ci in range(kk):
        members_idx = [i for i in range(len(points)) if labels[i] == ci]
        if not members_idx:
            cluster_mean_radius.append(None)
            continue
        cx, cy = centers[ci]
        ds = [math.hypot(points[i][0] - cx, points[i][1] - cy) for i in members_idx]
        cluster_mean_radius.append(round(sum(ds) / len(ds), 6))

    # -------------------------
    # Standardized payload
    # -------------------------
    assignments: Dict[str, int] = {}
    points_std: List[Dict[str, Any]] = []

    for i in range(len(names)):
        cid = int(labels[i]) if i < len(labels) else 0
        group_no = cid + 1  # 1..k
        nm = names[i]
        assignments[nm] = int(group_no)
        points_std.append({
            "name": nm,
            "x": round(points[i][0], 6),
            "y": round(points[i][1], 6),
            "group_no": int(group_no),
        })

    members_by_group: Dict[int, List[str]] = {g: [] for g in range(1, kk + 1)}
    for nm, gno in assignments.items():
        if 1 <= gno <= kk:
            members_by_group[gno].append(nm)

    groups: List[Dict[str, Any]] = []
    for gno in range(1, kk + 1):
        cid = gno - 1
        members = sorted(members_by_group.get(gno, []))
        cx, cy = centers[cid]
        groups.append({
            "group_no": int(gno),
            "size": int(len(members)),
            "members": members,
            "centroid": {"x": round(cx, 6), "y": round(cy, 6)},
            "mean_radius": cluster_mean_radius[cid],
        })

    return {
        "class_code": class_code,
        "session_id": sid,
        "k": int(kk),
        "n_points": int(len(points)),
        "groups": groups,
        "assignments": assignments,
        "points": points_std,
        "meta": {
            "inertia": round(float(inertia), 6),
            "cluster_sizes": cluster_sizes,
            "generated_at": datetime.utcnow().isoformat() + "Z",
        },
    }


# -------------------------
# Analysis routes
# -------------------------

@app.route("/analysis/class/<code>/<sid>/teacher_run/<int:run_id>.json")
def analysis_teacher_run(code, sid, run_id):
    if "teacher" not in session:
        return redirect("/teacher/login")

    code = (code or "").upper().strip()
    sid = (sid or "1").strip()

    run = db_get_teacher_run(run_id)
    if (not run) or run["class_code"] != code or run["session_id"] != sid:
        return jsonify({"error": "run not found"}), 404
    if run["teacher_username"] != session["teacher"]:
        return jsonify({"error": "forbidden"}), 403

    cache_key = f"teacher_run_{run_id}"
    cached = cache_get(code, sid, cache_key)
    if cached:
        return jsonify(cached)

    payload = teacher_run_distance_payload(code, sid, run_id)
    if not payload:
        return jsonify({"error": "failed"}), 500

    cache_set(code, sid, cache_key, payload)
    return jsonify(payload)


@app.route("/analysis/class/<code>/<sid>/student_avg.json")
def analysis_student_avg(code, sid):
    if "teacher" not in session:
        return redirect("/teacher/login")

    code = (code or "").upper().strip()
    sid = (sid or "1").strip()
    if sid not in ["1", "2", "3", "4", "5"]:
        sid = "1"

    cls = db_get_class_for_teacher(code, session["teacher"])
    if not cls or cls.get("_forbidden"):
        return jsonify({"error": "forbidden"}), 403

    cache_key = f"student_avg_{sid}"
    cached = cache_get(code, sid, cache_key)
    if cached:
        return jsonify(cached)

    payload = student_avg_distance_payload(code, sid)
    cache_set(code, sid, cache_key, payload)
    return jsonify(payload)


@app.route("/analysis/class/<code>/<sid>/student/<path:student_name>/vs_avg.json")
def analysis_student_vs_avg(code, sid, student_name):
    if "teacher" not in session:
        return redirect("/teacher/login")

    code = (code or "").upper().strip()
    sid = (sid or "1").strip()
    if sid not in ["1", "2", "3", "4", "5"]:
        sid = "1"

    student_name = unquote(student_name or "").strip()

    cls = db_get_class_for_teacher(code, session["teacher"])
    if not cls or cls.get("_forbidden"):
        return jsonify({"error": "forbidden"}), 403

    if not student_name:
        return jsonify({"error": "student_name required"}), 400

    cache_key = f"student_vs_avg_{sid}_{student_name}"
    cached = cache_get(code, sid, cache_key)
    if cached:
        return jsonify(cached)

    payload = student_vs_avg_distance_payload(code, sid, student_name)
    if not payload:
        return jsonify({"error": "not found"}), 404

    cache_set(code, sid, cache_key, payload)
    return jsonify(payload)


@app.route("/analysis/class/<code>/<sid>/kmeans_summary.json")
def analysis_kmeans_summary(code, sid):
    if "teacher" not in session:
        return redirect("/teacher/login")

    code = (code or "").upper().strip()
    sid = (sid or "1").strip()
    if sid not in ["1", "2", "3", "4", "5"]:
        sid = "1"

    cls = db_get_class_for_teacher(code, session["teacher"])
    if not cls or cls.get("_forbidden"):
        return jsonify({"error": "forbidden"}), 403

    k_raw = request.args.get("k", "3")
    try:
        k = int(k_raw)
    except Exception:
        k = 3
    if k < 2:
        k = 2
    if k > 4:
        k = 4

    cache_key = f"kmeans_summary_{sid}_k{k}"
    cached = cache_get(code, sid, cache_key)
    if cached:
        return jsonify(cached)

    payload = kmeans_summary_payload(code, sid, k)
    cache_set(code, sid, cache_key, payload)
    return jsonify(payload)


@app.route("/analysis/class/<code>/<sid>/spm_result.json")
def analysis_spm_result(code, sid):
    if "teacher" not in session:
        return redirect("/teacher/login")

    code = (code or "").upper().strip()
    sid = (sid or "1").strip()
    if sid not in ["1", "2", "3", "4", "5"]:
        sid = "1"

    cls = db_get_class_for_viewer(code, session["teacher"])
    if not cls or cls.get("_forbidden"):
        return jsonify({"ok": False, "error": "forbidden"}), 403

    refresh = (request.args.get("refresh") or "").strip().lower() in ["1", "true", "yes"]
    cache_key = f"spm_result_{sid}_v1"
    if refresh:
        cache_clear_session_analysis(code, sid)
    cached = None if refresh else cache_get(code, sid, cache_key)
    if cached:
        return jsonify(cached)

    payload = build_spm_result_payload(code, sid)
    cache_set(code, sid, cache_key, payload)
    return jsonify(payload)


@app.route("/analysis/class/<code>/<sid>/privacy_relation_payload.json")
def analysis_privacy_relation_payload(code, sid):
    if "teacher" not in session:
        return jsonify({"ok": False, "error": "unauthorized"}), 401

    code = (code or "").upper().strip()
    sid = (sid or "1").strip()
    if sid not in ["1", "2", "3", "4", "5"]:
        sid = "1"

    cls = db_get_class_for_viewer(code, session["teacher"])
    if not cls or cls.get("_forbidden"):
        return jsonify({"ok": False, "error": "forbidden"}), 403

    student_id = (request.args.get("student_id") or "").strip().upper() or None
    payload = build_privacy_safe_relation_payload(code, sid, student_id=student_id)
    return jsonify(payload)


@app.route("/analysis/class/<code>/<sid>/ai_relation_chat.json", methods=["POST"])
def analysis_ai_relation_chat(code, sid):
    if "teacher" not in session:
        return jsonify({"ok": False, "error": "unauthorized"}), 401

    code = (code or "").upper().strip()
    sid = (sid or "1").strip()
    if sid not in ["1", "2", "3", "4", "5"]:
        sid = "1"

    cls = db_get_class_for_viewer(code, session["teacher"])
    if not cls or cls.get("_forbidden"):
        return jsonify({"ok": False, "error": "forbidden"}), 403

    body = request.get_json(silent=True) or {}
    question_text = (body.get("question_text") or "").strip()
    free_question = bool(question_text)
    question_type = _classify_relation_free_question(question_text) if free_question else (body.get("question_type") or "overview").strip()
    if free_question and question_type == "unsupported":
        out = {
            "ok": True,
            "answer": _limit_ai_answer(_safe_relation_question_refusal()),
            "notice": "",
            "source": "local",
            "followups": _relation_chat_followups("overview"),
        }
        return jsonify(out)

    allowed_questions = {
        "overview",
        "why_type",
        "why_distributed",
        "why_type_basis",
        "student_to_peers",
        "student_near_list",
        "student_far_list",
        "student_spread",
        "peers_to_student",
        "peers_near_list",
        "peers_far_list",
        "mutual_similarity",
        "distance_gap",
        "distance_gap_detail",
        "distance_gap_largest",
        "distance_gap_effect",
        "asymmetry",
        "cautions",
    }
    if question_type not in allowed_questions:
        question_type = "overview"
    student_id = (body.get("student_id") or "").strip().upper()
    if student_id and not (student_id.startswith("S") and student_id[1:].isdigit()):
        student_id = ""

    if free_question and not student_id:
        try:
            result = build_spm_result_payload(code, sid)
            names = [str(s.get("name") or "") for s in (result.get("students") or []) if str(s.get("name") or "")]
            anon = _anon_id_map(names)
            for name in sorted(names, key=len, reverse=True):
                if name and name in question_text:
                    student_id = anon.get(name, "")
                    break
        except Exception:
            student_id = ""

    if question_type != "overview" and not student_id:
        return jsonify({
            "ok": True,
            "answer": "먼저 학생을 선택해 주세요. 학생을 선택하면 이 질문에 답할 수 있습니다.",
            "notice": "",
            "source": "local",
            "followups": _relation_chat_followups("overview"),
        })

    local_question = free_question or question_type != "overview"
    ai_mode = "local" if local_question else ("openai" if (os.environ.get("OPENAI_API_KEY") or "").strip() else "fallback")
    model_key = (os.environ.get("OPENAI_MODEL") or "gpt-5.4-mini").strip() if ai_mode == "openai" else "local"
    cache_key = f"ai_relation_chat_v5_safe_{sid}_{student_id or 'class'}_{question_type}_{ai_mode}_{model_key}"
    cached = cache_get(code, sid, cache_key)
    if cached:
        return jsonify(cached)

    privacy_payload = build_privacy_safe_relation_payload(code, sid, student_id=student_id or None)
    if local_question:
        ctx = _visible_relation_context(code, sid, student_id, privacy_payload)
        answer = _relation_chat_local_answer(question_type, ctx)
        source = "local"
    else:
        answer, source = _openai_relation_answer(question_type, privacy_payload)
    out = {
        "ok": True,
        "answer": _limit_ai_answer(answer),
        "notice": "",
        "source": source,
        "followups": _relation_chat_followups(question_type),
    }
    cache_set(code, sid, cache_key, out)
    return jsonify(out)


@app.route("/analysis/class/<code>/<sid>/dbscan_structure.json")
def analysis_dbscan_structure(code, sid):
    if "teacher" not in session:
        return redirect("/teacher/login")

    code = code.upper()
    sid = sid if sid in ["1", "2", "3", "4", "5"] else "1"

    cls = db_get_class_for_teacher(code, session["teacher"])
    if not cls or cls.get("_forbidden"):
        return jsonify({"error": "forbidden"}), 403

    refresh = (request.args.get("refresh") or "").strip().lower() in ["1", "true", "yes"]
    if refresh:
        cache_clear_session_analysis(code, sid)

    cache_key = f"dbscan_structure_{sid}"
    cached = None if refresh else cache_get(code, sid, cache_key)
    if cached:
        return jsonify(cached)

    payload = dbscan_structure_payload(code, sid)
    cache_set(code, sid, cache_key, payload)
    return jsonify(payload)

@app.route("/analysis/class/<code>/dbscan_change.json")
def analysis_dbscan_change(code):
    if "teacher" not in session:
        return redirect("/teacher/login")

    code = (code or "").upper().strip()

    prev_sid = (request.args.get("prev") or "1").strip()
    curr_sid = (request.args.get("curr") or "2").strip()

    valid_sids = ["1", "2", "3", "4", "5"]
    if prev_sid not in valid_sids:
        prev_sid = "1"
    if curr_sid not in valid_sids:
        curr_sid = "2"

    cls = db_get_class_for_teacher(code, session["teacher"])
    if not cls or cls.get("_forbidden"):
        return jsonify({"error": "forbidden"}), 403

    # prev payload
    prev_cache_key = f"dbscan_structure_{prev_sid}"
    prev_payload = cache_get(code, prev_sid, prev_cache_key)
    if not prev_payload:
        prev_payload = dbscan_structure_payload(code, prev_sid)
        cache_set(code, prev_sid, prev_cache_key, prev_payload)

    # curr payload
    curr_cache_key = f"dbscan_structure_{curr_sid}"
    curr_payload = cache_get(code, curr_sid, curr_cache_key)
    if not curr_payload:
        curr_payload = dbscan_structure_payload(code, curr_sid)
        cache_set(code, curr_sid, curr_cache_key, curr_payload)

    prev_counts = prev_payload.get("counts") or {}
    curr_counts = curr_payload.get("counts") or {}

    change = dbscan_change_summary(prev_counts, curr_counts)

    return jsonify({
        "class_code": code,
        "prev_sid": prev_sid,
        "curr_sid": curr_sid,
        "prev_counts": prev_counts,
        "curr_counts": curr_counts,
        "change": change,
    })


# -------------------------
# Main
# -------------------------

if __name__ == "__main__":
    port = int(os.environ.get("PORT", "5000"))
    app.run(host="0.0.0.0", port=port)


def db_get_students_with_pin(class_code: str):
    """Returns a list of dicts: [{no:int|None, name:str, pin_code:str}, ...]"""
    if not engine:
        raise RuntimeError("DB engine not initialized")

    class_code = (class_code or "").upper().strip()

    with engine.connect() as conn:
        rows = conn.execute(text("""
            SELECT student_no, name, pin_code
            FROM students
            WHERE class_code = :code
              AND active = TRUE
            ORDER BY id ASC
        """), {"code": class_code}).fetchall()

    out = []
    for r in rows:
        out.append({
            "no": int(r[0]) if r[0] is not None else None,
            "name": (r[1] or "").strip(),
            "pin_code": str(r[2] or "").strip(),
        })
    return out
